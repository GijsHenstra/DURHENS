import itertools
import datetime
import pandas as pd
import numpy as np
import re
import os
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from scipy.stats import gaussian_kde
from mpl_toolkits.axes_grid1 import make_axes_locatable
from itertools import product
from openpyxl import load_workbook


plt.rcParams['figure.dpi'] = 1000
plt.style.use('bmh')

colors = plt.rcParams["axes.prop_cycle"].by_key()["color"]

TITLE = True
SAVE = True


#%% functions

time_dict = {
    '24-H (00:00-24:00)': ('00:00', '23:59'),
    'Evening (18:00-20:00)': ('18:00', '20:01'),
    'Before sunrise (04:00)': ('04:00', '04:01'),
    # 'Daytime (09:010-19:00)': ('09:00', '19:01'),
    # 'Midday (14:00)': ('13:50', '13:51'),
    # 'Morning (09:00)': ('09:00', '09:01'),
    # 'Midday (13:00 - 14:00)': ('13:00', '14:01'),
    # 'Hours before midday (11:00 - 14:00)': ('11:00', '14:01'),
    'Around Noon (12:00 - 16:00)': ('12:00', '16:01'),
    # 'daytime': ('06:00', '21:00'),
    # 'nighttime': ('21:00', '06:00'),
    }

line_properties = {
    'albedo': {
        'linestyle':{
            0.1: 'dotted',
            0.2: 'solid',
            0.4: (0, (15, 3)),
            0.6: (0, (2, 2)),
            0.8: (0, (2, 6)),
            },
        'color': {
            0.0: colors[1],
            0.1: colors[2],
            0.2: colors[0],
            0.4: colors[3],
            0.6: colors[4],
            0.8: colors[5],
            1.0: colors[6],
            }
        },
        
    'hw': {
        'linewidth': {
            0.0: 0.75,
            0.5: 1.25,
            1.0: 1.75,
            2.0: 2.25,
            3.0: 2.75,
            4.0: 3.25,
            
            '0.5 - 0.0': 0.5,
            '1.0 - 0.0': 1.0,
            '2.0 - 0.0': 2.0,
            '4.0 - 0.0': 2.5,
            
            '0.5 / 0.0': 0.5,
            '1.0 / 0.0': 1.0,
            '2.0 / 0.0': 2.0,
            '4.0 / 0.0': 2.5,
            },
        'color': {
            1.0: colors[0],
            0.0: colors[1],
            0.5: colors[2],
            1.5: colors[7],
            2.0: colors[4],
            4.0: colors[3],
            3.0: colors[6],
            8.0: colors[5],
            
            }
        },
    
    # [':', '-.',  (0, (3, 2)), '-', (0, (15, 3))]
    
    'location': {
        'color': {
            'east-facing wall': 'tab:blue',
            'west-facing wall':'tab:orange',
            'north-facing wall':'tab:green',
            'south-facing wall':'tab:red',
            'upwards-facing roof': 'tab:purple',
            
            'canyon street (x-slice)': 'tab:blue',
            'canyon street (y-slice)': 'tab:blue',
            'walls': 'tab:orange',
            'bottom (x-slice)': 'tab:green',
            'top (x-slice)': 'tab:olive',
            'total': 'tab:red',
            'bottom (x-slice) / top (x-slice)': 'tab:grey',
            'bottom (y-slice) / top (y-slice)': 'tab:grey',
            'total + bottom (x-slice)': 'tab:grey',
            'north-facing wall (x-slice)': 'tab:purple',
            'south-facing wall (x-slice)': 'tab:brown',
            },
        'linestyle': {
            'east-facing wall': (0, (15, 3)),
            'west-facing wall': 'dotted',
            'north-facing wall': (0, (3, 2)),
            'south-facing wall': 'dashdot',
            
            'east-facing wall (y-slice)': (0, (15, 3)),
            'west-facing wall (y-slice)': 'dotted',
            'north-facing wall (x-slice)': (0, (3, 2)),
            'south-facing wall (x-slice)': 'dashdot',
            
            'upwards-facing roof': 'solid',
            'canyon street (x-slice)': 'solid',
            'canyon street (y-slice)': 'solid',
            'top (x-slice)': 'dotted',
            'top (y-slice)': 'dotted',
            },
        },
    
    'variable': {
        'color': {
            "q_sw['in_from_els']": colors[0],
            "q_sw['in_from_sky']": colors[0],
            "q_sw['in_dir']": colors[0],
            "q_sw['in_diff']": colors[0],
            "q_sw['abs_from_els']": colors[0],
            "q_sw['abs_from_sky']": colors[0],
            "q_sw['abs_dir']": colors[0],
            "q_sw['abs_diff']": colors[0],
            "q_sw['abs']": colors[0],
            "q_sw['in']": colors[0],
            "q_sw['in'] - q_sw['abs']": colors[0],
            "q_sw_sensors['in_from_els']": colors[0],
            "q_sw_sensors['in_from_sky']": colors[0],
            "q_sw_sensors['in_dir']": colors[0],
            "q_sw_sensors['in_diff']": colors[0],
            
            "q_lw['in_from_els']": colors[1],
            "q_lw['in_from_sky']": colors[1],
            "q_lw['in_from_els'] + q_lw['in_from_sky']": colors[1],
            "q_lw['in_from_sky'] + q_lw['in_from_els']": colors[1],
            "q_lw['abs_from_els']": colors[1],
            "q_lw['abs_from_sky']": colors[1],
            "q_lw['abs_from_els'] + q_lw['abs_from_sky']": colors[1],
            "q_lw['abs_from_sky'] + q_lw['abs_from_els']": colors[1],
    
            "q_lw['net']": colors[1],
            "q_lw['out_rad']": colors[1],
            
            "T": colors[2],
    
            "q_lw['net'] + q_sw['abs']": colors[3],
            "q_lw['abs'] + q_sw['abs']": colors[3],
            "q_sw['abs'] + q_lw['abs']": colors[3],
            "q_sw['abs'] + q_lw['net']": colors[3],
            "q_sw['in_from_els'] + q_sw['in_dir'] + q_sw['in_diff'] + q_lw['in_from_els'] + q_lw['in_from_sky']": colors[3],
            
            "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs_from_sky']": colors[3],
            "q_sw['abs_from_els'] + q_lw['abs_from_els']": colors[3],
            "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs']": colors[3],
            "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs']": colors[3],
            "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['net']": colors[3],
            
            "q_convection": colors[4],
            
            },
        'linestyle': {
            "q_sw_sensors['in'] + q_lw_sensors['in']": '-',
            "q_lw_sensors['in']": ':',
            "q_sw_sensors['in']": (0, (15, 3)),
            
            "q_lw['in_from_els']": 'dotted', 
            "q_lw['in_from_sky']": 'dashed', 
            "q_lw['in_from_els'] + q_lw['in_from_sky']": 'solid', 
            "q_lw['net'] - q_lw['in_from_els'] - q_lw['in_from_sky']": 'dashdot',
            
            
            "q_sw['in_from_els'] + q_sw['in_dir'] + q_sw['in_diff']": 'solid', 
            "q_sw['in_from_els']": 'dotted', 
            "q_sw['in_dir']": (0, (5, 1)), 
            "q_sw['in_diff']": (0, (5, 5)), 
            "q_sw['in_dir'] + q_sw['in_diff']": 'dashed', 
            
            
            "q_lw['abs_from_els']": 'dotted', 
            "q_lw['abs_from_sky']": 'dashed', 
            "q_lw['abs_from_els'] + q_lw['abs_from_sky']": 'solid', 
            "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff']": 'solid', 
            "q_sw['abs_from_els']": 'dotted', 
            "q_sw['abs_dir']": (0, (5, 1)), 
            "q_sw['abs_diff']": (0, (5, 5)), 
            "q_sw['abs_dir'] + q_sw['abs_diff']": 'dashed', 
            
            
            
            "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs_from_sky']": 'dashed', 
            "q_sw['abs_from_els'] + q_lw['abs_from_els']": 'dotted', 
            # "q_sw['abs']", 
            # "q_lw['out_rad']"
            
            "q_convection": (0, (3, 1, 1, 1, 1, 1)),
            
            }
        },
    
    'wind speed': {
        'linewidth': {
            0:0.75, 
            1:1, 
            2:1.75,
            5: 2.5,
            },
        },
    
    'skin thickness': {
        'linewidth': {
            0.01: 0.75,
            0.05: 1.25,
            0.1: 1.75,
            0.15: 2.25,
            0.25: 2.75,
            0.5: 3.25,
            },
        },
    }
    
markers_dict = {
    # "q_lw['in_from_els']": '_', # U+21c6 / ⇅
    "q_lw['in_from_els']": '$\u21ff$', # single
    "q_lw['in_from_sky']": '$↓$',
    "q_lw['in_from_els'] + q_lw['in_from_sky']": 'P',
    "q_lw['in_from_sky'] + q_lw['in_from_els']": 'P',
    # "q_lw['in_from_sky'] + q_lw['in_from_els']": '$\u271a$',
    "q_lw['out_rad']": '$↑$',
    "q_lw['out_rad']": '$\u21e7$',
    "q_lw['in_from_sky']": '$\u21e9$',
    "q_sw['abs']": '$\u21a1$',
    # "q_lw['net'] + q_sw['abs']": 'D',
    "q_lw['net']": 'D',
    # "q_lw['out_rad']": '$\u2B06$',
    }

    
markersize_dict = {
    'P': 10,
    'D': 6,
    '$\u21e7$': 15,
    '$↓$': 15,
    '$↑$': 15,
    '$\u21e9$': 15,
    '$\u21f0$': 15,
    '$\u21ff$': 15,
    '$\u21a1$': 15,
    }


markeredgewidth_dict = {
    'P': 1,
    'D': 1,
    '$\u2194$': 0.5,
    '$↓$': 0.5,
    '$↑$': 0.5,
    '$\u1f815$': 0.5,
    }

legend_titles = {
    'albedo': r'$\alpha$',
    'hw': 'H/W'}

labels_dict = {
    'canyon street (x-slice)': 'Street',
    'walls': 'Walls',
    'bottom': 'bottom of sensors',
    
    # "q_lw['in_from_sky']": r'LW$_\mathrm{in from sky}$',
    "q_lw['in']": r'$L^\downarrow$',
    "q_lw['in_from_sky']": r'$L^\downarrow_\mathrm{sky}$',
    "q_lw['in_from_els']": r'$L^\downarrow_\mathrm{env}$',
    "q_lw['out_rad']": r'$L^\uparrow_\mathrm{emm}$',
    "q_lw['net']": r'$L^\ast$',
    
    "q_lw['abs_from_sky']": r'$L^\downarrow_\mathrm{sky,abs}$',
    "q_lw['abs_from_els']": r'$L^\downarrow_\mathrm{env,abs}$',
    "q_lw['abs']": r'$L^\downarrow_\mathrm{abs}$',
    "q_lw['abs_from_els'] + q_lw['abs_from_sky']": r'$L^\downarrow_\mathrm{abs}$',
    "q_lw['abs_from_sky'] + q_lw['abs_from_els']": r'$L^\downarrow_\mathrm{abs}$',
    "q_lw['net'] - q_lw['in_from_els'] - q_lw['in_from_sky']": r'$L^\uparrow$',
    "q_lw['in_from_els'] + q_lw['in_from_sky']": r'$L^\downarrow$',
    
    # "q_lw_sensors['in']": r'$L^{\ast}$',
    # "q_sw_sensors['in']": r'$K^{\ast}$',
    # "q_lw_sensors['in'] + q_sw_sensors['in']": r'$Q^{\ast}$',
    # "q_sw_sensors['in'] + q_lw_sensors['in']": r'$Q^{\ast}$',
    # 
    # "q_lw_sensors['in']": r'$L^{\downarrow}$',
    # "q_sw_sensors['in']": r'$K^{\downarrow}$',
    
    "q_lw_sensors['in_from_els'] + q_lw_sensors['in_from_sky']": r'L$_\downarrow$',
    
    "q_sw['abs']": r'$K^\downarrow_\mathrm{abs}$',
    "q_sw['in']": r'$K^\downarrow$',
    "q_sw['in_diff']": r'$K^\downarrow_\mathrm{diff}$',
    "q_sw['in_dir']": r'$K^\downarrow_\mathrm{dir}$',
    "q_sw['in_from_els']": r'$K^\downarrow_\mathrm{env}$',
    "q_sw['in_diff'] + q_sw['in_dir']": r'$K^\downarrow_\mathrm{sky}$',
    "q_sw['in_dir'] + q_sw['in_diff']": r'$K^\downarrow_\mathrm{sky}$',
    
    "q_sw['abs_diff']": r'$K^\downarrow_\mathrm{diff,abs}$',
    "q_sw['abs_dir']": r'$K^\downarrow_\mathrm{dir,abs}$',
    
    "q_sw['abs_dir'] + q_sw['abs_diff']": r'$K^\downarrow_\mathrm{sky,abs}$',
    "q_sw['abs_diff'] + q_sw['abs_dir']": r'$K^\downarrow_\mathrm{sky,abs}$',
    "q_sw['abs_from_els']": r'$K^\downarrow_\mathrm{env,abs}$',
    "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff']": r'$K^\downarrow_\mathrm{abs}$',
    
    "q_lw['net'] + q_sw['abs']": r'$Q^\ast$',
    "q_sw['in'] - q_sw['abs']": r'$K^\uparrow_{refl}$',
    
    
    # use K+L
    "q_lw['abs'] + q_sw['abs']": r'$K^\downarrow_\mathrm{abs} + L^\downarrow_\mathrm{abs}$',
    "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs']": r'$K^\downarrow_\mathrm{abs}$ $\plus$ $L^\downarrow_\mathrm{abs}$',
    "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs_from_sky']": r'$K^\downarrow_\mathrm{sky,abs}$ $\plus$ $L^\downarrow_\mathrm{sky,abs}$ ',
    "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['net']": r'$K^\ast$ $\plus$ $L^\ast$',
    
    # use Q instead of K+L
    "q_lw['abs'] + q_sw['abs']": r'$Q^\downarrow_\mathrm{abs}$',
    "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs']": r'$Q^\downarrow_\mathrm{abs}$',
    "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs_from_sky']": r'$Q^\downarrow_\mathrm{sky,abs}$',
    "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['net']": r'$Q^\ast$',
    "q_sw['abs_from_els'] + q_lw['abs_from_els']": r'$Q^\downarrow_\mathrm{env,abs}$',
    
    'q_convection': '$Q_H$',
    # "q_sw['in_from_els'] + q_sw['in_dir'] + q_sw['in_diff'] + q_lw['in_from_els'] + q_lw['in_from_sky']": '$K^\downarrow + L^\downarrow$',
    
    'east-facing wall': 'E-wall',
    'west-facing wall': 'W-wall',
    'north-facing wall': 'N-wall',
    'south-facing wall': 'S-wall',
    'upwards-facing roof': 'Roof',
    
    'east-facing wall (y-slice)': 'E-facing wall',
    'west-facing wall (y-slice)': 'W-facing wall',
    'north-facing wall (x-slice)': 'N-facing wall',
    'south-facing wall (x-slice)': 'S-facing wall',
    'street': 'Street',
    'top (x-slice)': 'Canyon top',
    'top (y-slice)': 'Canyon top',
}

variable_dict = {
    'skin thickness': r'$d$',
    'albedo': r'$\alpha$',
    'emissivity': r'$\varepsilon$',
    'wind speed': r'$v$',
    'vol. heat cap.': '$C$'
    }

y_label1 = r'Energy flux density [W m$^{-2}$]'
y_label2 = r'Daily energy flux density [MJ m$^{-2}$ d$^{-1}$]'
y_label_T = r'Surface temperature [K]'
y_label_deg_C = 'Surface temperature [$^{\circ}$C]'
# y_label_ground_deg_C = 'Temperature [$^{\circ}$C]'

x_label = {
    'hw': 'H/W ratio [-]',
    't': 'Local time [H]',
    'svf': 'Sky View Factor [-]',
    'albedo': 'Surface Albedo [-]'}

xlim_hw = [-0.25, 4.25]
colorbar_label_W = r'[W/m$^2$]'

ylim_MJ_normalised = [-12, 30]
ylim_MJ_normalised = [-12, 40]
ylim_W = [-250, 725]
ylim_T = [285, 323]

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

def replace_keys_in_tuple_strings(my_tuple, my_dict, label_not_found='NO_LABEL'):
    new_tuple = []
    for inner_tuple in my_tuple:
        
        # remove '()' around tuple
        if ', ' in inner_tuple:
            inner_tuple = inner_tuple[1:-1]

        inner_list = []
        for item in inner_tuple.split(", "):
            key = item.strip("'")
            value = my_dict.get(key, key)
            
            if isinstance(value, str):
                try:
                    value = my_dict[value]
                except KeyError:
                    split_values = value.split(' + ')
                    if len(split_values) > 1:
                        split_values = [v.strip() for v in split_values]
                        replaced_values = [my_dict.get(v, f'??{v}??"') for v in split_values]
                        value = ' + '.join(replaced_values)
                
            inner_list.append(value)
            
        if len(inner_list) == 1:
            new_tuple.append(inner_list[0])
        else:
            new_tuple.append(tuple(inner_list))
        
    return tuple(new_tuple)

def infer_data_type(value_str):
    try:
        value = float(value_str)
        return value
    except ValueError:
        return value_str
    
def select_nth_element(lst, n):
    if isinstance(lst, list) and len(lst) > n:
        return lst[n]
    else:
        return None

def elementwise_operation(df1, df2, operation):
    list1, list2 = df1.tolist(), df2.tolist()
    if operation == ' + ':
        return [x + y for x, y in zip(df1.values, df2.values)]
    elif operation == ' - ':
        return [x - y for x, y in zip(df1, df2)]
    elif operation == ' * ':
        return [x * y for x, y in zip(df1, df2)]
    elif operation == ' / ':
        return [x / y if y != 0 else float('nan') for x, y in zip(df1, df2)]
    else:
        raise ValueError("Unsupported operation")

def transform_to_filename(title):
    # Remove illegal characters
    filename = re.sub(r'[\\/:"*?<>|]', '', title)
    
    # Replace newline characters with underscores
    filename = filename.replace('\n', '_')
    
    # Replace spaces with underscores
    filename = filename.replace(' ', '_')
    
    # Limit the filename length if necessary
    max_length = 255  # Adjust the value as per your requirements
    if len(filename) > max_length:
        filename = filename[:max_length]
    
    return filename


def filter_dataframe(df, plot_setup, sort=True):

    multiply = None    
    temp_column = None

    if 'date' in plot_setup.keys():
        df_xs = df.loc[df.index.get_level_values('datetime').date == plot_setup['date']]
    else:
        df_xs = df
    
    slicers_dict = {item[0]:item[1] for item in plot_setup['xs'].items() if not isinstance(item[1], (tuple, list, slice))}
    # loc_dict = {key: value[:] for key, value in plot_setup['xs'].items() if isinstance(value, (tuple, list))}
    loc_dict = {key: value[:] if isinstance(value, (tuple, list)) else value for key, value in plot_setup['xs'].items() if isinstance(value, (tuple, list, slice))}
    
    if 'variable' in loc_dict:
        if any(["q_sw['abs']" in var for var in loc_dict['variable']]):
            temp_column = "q_sw['in']"
            loc_dict['variable'].append(temp_column)
            multiply = 'albedo'
            
    operations = {}
    i=0
    for key, value in loc_dict.items():
        if value == slice(None):
            continue
        for item in value:
            for operator in [' / ', ' * ', ' + ', ' - ']:
                if isinstance(item, str) and operator in item:
                    # col1, col2 = [col.strip() for col in item.split(operator)[:2]]
                    # col1 = infer_data_type(col1)
                    # col2 = infer_data_type(col2)
                    # loc_dict[key].append(col1)
                    # loc_dict[key].append(col2)
                    # operations[key+str(i)] = {operator: [col1, col2]}
                    # i+=1
                    
                    # cols = [infer_data_type(col.strip()) for col in item.split(operator)]
                    cols = [col.strip() for col in item.split(operator)]
                    # col1 = infer_data_type(col1)
                    # col2 = infer_data_type(col2)
                    [loc_dict[key].append(infer_data_type(col)) for col in cols]
                    # loc_dict[key].append(col2)
                    
                    cols = [infer_data_type(col) for col in cols]
                    
                    operations[key+str(i)] = {operator: cols}
                    i+=1
                    
    
    # slice the columns using provided floats, drop labels
    if slicers_dict:
        df_xs = df_xs.xs(key=slicers_dict.values(), level=list(slicers_dict.keys()), axis=1)
    
    # slice the columns using provided lists
    try:
        if isinstance(df_xs.columns, pd.MultiIndex):
            names = tuple(loc_dict[name] for name in df_xs.columns.names)
            df_xs = df_xs.loc[:, names]
        else:
            df_xs = df_xs.loc[:, list(loc_dict.values())[0]]
        
    except KeyError:
        return None
    
    if df_xs.empty:
        return None
    
    if plot_setup['x'] == 'svf':
        df_xs.columns = df_xs.columns.remove_unused_levels()
        df_xs.columns = df_xs.columns.set_levels(list(df_xs.columns.get_level_values('hw').map(hw_to_svf)), level='hw', verify_integrity=False).rename(names='svf', level='hw')
        
    
    # if multiply == 'albedo':
    #     if "q_sw['abs']" in loc_dict['variable']:
    #         _df_in = df_xs.xs(key="q_sw['in']", level='variable', axis=1)
    #         _df_abs = pd.concat({"q_sw['abs']": _df_in * (1-_df_in.columns.get_level_values('albedo'))}, names=['variable'], axis=1).reorder_levels(df_xs.columns.names, axis=1)
    #         df_xs = pd.concat([df_xs, _df_abs], axis=1)
        
    if 'droplevel' in plot_setup.keys():
        df_xs = df_xs.droplevel(plot_setup['droplevel'])
        
    if 'droplevel_col' in plot_setup.keys():
        df_xs = df_xs.droplevel(plot_setup['droplevel_col'], axis=1)
        
    if 'between_time' in plot_setup.keys():
        df_xs = df_xs.between_time(plot_setup['between_time'][0], plot_setup['between_time'][1])
        
    if plot_setup['x'] != 't':
        df_xs = df_xs.resample('D').mean().stack(plot_setup['x'])
        df_xs = df_xs.droplevel('datetime')
        
    if 'y_unit' in plot_setup.keys():
        if plot_setup['y_unit'] == 'MJ/m2/d':
            df_xs = df_xs * (24*60*60)
        
    for level, operation_dct in operations.items():
        level = ''.join([i for i in level if not i.isdigit()])
        
        for operator, cols in operation_dct.items():
                
            df_col0 = df_xs.xs(key=cols[0], level=level, axis=1)
            df_col1 = df_xs.xs(key=cols[1], level=level, axis=1)
            
            cols = list(map(str, cols))
            
            # new after 6 oct
            if df_col0.iloc[0].dtype == 'object':
                if len(cols) > 2:
                    raise NotImplementedError('Fix functionality to use multiple operations for raw data.')
                
                for df_col0_cols, df_col0_lst in df_col0.items():
                    try:
                        df_col0_arr = np.array(df_col0_lst.tolist())
                        # df_col0_arr = np.array(df_col0.values.tolist()) # !!! after 28 aug
                        
                        if len(df_col0_lst[0]) == len(df_col1[df_col0_cols].tolist()[0]):
                            df_col1_arr = np.array(df_col1[df_col0_cols].tolist())
                        else:
                            df_col1_arr = np.array(df_col1[df_col0_cols].apply(lambda x: sum(x) / len(x)).values)[:, np.newaxis]
                        
                        
                        if operator == ' / ':
                            col0_operator_col1_arr = df_col0_arr / df_col1_arr
                        elif operator == ' + ':
                            col0_operator_col1_arr = df_col0_arr + df_col1_arr
                        elif operator == ' - ':
                            col0_operator_col1_arr = df_col0_arr - df_col1_arr
                        elif operator == ' * ':
                            col0_operator_col1_arr = df_col0_arr * df_col1_arr
                        else:
                            raise ValueError("Invalid operation")
                            
                        col0_operator_col1 = pd.DataFrame(col0_operator_col1_arr[:, np.newaxis].tolist(), index=df_col0.index, columns=df_col1[[df_col0_cols]].columns)
                            
                        _df_operated = pd.concat({operator.join(cols): col0_operator_col1}, names=[level], axis=1).reorder_levels(df_xs.columns.names, axis=1)
                        df_xs = pd.concat([df_xs, _df_operated], axis=1)
                    except (KeyError, TypeError) as e:
                        # pass
                        print(e)
                
                
            else:
                
                if operator == ' / ':
                    if len(cols) > 2:
                        raise NotImplementedError('Fix functionality to use divide multiple columns for raw data.')
                    col0_operator_col1 = df_col0 / df_col1
                elif operator == ' + ':
                    col0_operator_col1 = sum([df_xs.xs(key=col, level=level, axis=1) for col in cols])
                elif operator == ' - ':
                    col0_operator_col1 = df_col0 - df_col1
                else:
                    raise ValueError("Invalid operation")
            
                _df_operated = pd.concat({operator.join(cols): col0_operator_col1}, names=[level], axis=1).reorder_levels(df_xs.columns.names, axis=1)
                df_xs = pd.concat([df_xs, _df_operated], axis=1)
                
            #     # new after 6 oct
            # if df_col0.iloc[0].dtype == 'object' and df_col0.shape[1] > 1:
            #     for i in range(df_col0.shape[1]):
            #         df_col0_arr = np.array(df_col0.iloc[:, i].values.tolist())
            #         # df_col0_arr = np.array(df_col0.values.tolist()) # !!! after 28 aug
                    
            #         if len(df_col0.iloc[0].values[i]) == len(df_col1.iloc[0].values[i]):
            #             df_col1_arr = np.array(df_col1.iloc[:, i].values.tolist())
            #         else:
            #             df_col1_arr = np.array(df_col1.iloc[:, i].apply(lambda x: sum(x) / len(x)).values)[:, np.newaxis]
                    
                    
            #         if operator == ' / ':
            #             col0_operator_col1_arr = df_col0_arr / df_col1_arr
            #         elif operator == ' + ':
            #             col0_operator_col1_arr = df_col0_arr + df_col1_arr
            #         elif operator == ' - ':
            #             col0_operator_col1_arr = df_col0_arr - df_col1_arr
            #         elif operator == ' * ':
            #             col0_operator_col1_arr = df_col0_arr * df_col1_arr
            #         else:
            #             raise ValueError("Invalid operation")
                        
            #         col0_operator_col1 = pd.DataFrame(col0_operator_col1_arr[:, np.newaxis].tolist(), index=df_col0.index, columns=df_col0.columns[[i]])
                        
            #         _df_operated = pd.concat({f'{cols[0]}{operator}{cols[1]}': col0_operator_col1}, names=[level], axis=1).reorder_levels(df_xs.columns.names, axis=1)
            #         df_xs = pd.concat([df_xs, _df_operated], axis=1)
                
            # else:
            #     if operator == ' / ':
            #         col0_operator_col1 = df_col0 / df_col1
            #     elif operator == ' + ':
            #         col0_operator_col1 = df_col0 + df_col1
            #     elif operator == ' - ':
            #         col0_operator_col1 = df_col0 - df_col1
            #     else:
            #         raise ValueError("Invalid operation")
            
            #     _df_operated = pd.concat({f'{cols[0]}{operator}{cols[1]}': col0_operator_col1}, names=[level], axis=1).reorder_levels(df_xs.columns.names, axis=1)
            #     df_xs = pd.concat([df_xs, _df_operated], axis=1)
    
    loc_dict = {key: value[:] if isinstance(value, (tuple, list)) else value for key, value in plot_setup['xs'].items() if isinstance(value, (tuple, list, slice))}
    if isinstance(df_xs.columns, pd.MultiIndex):
        df_xs = df_xs.loc[:, tuple(loc_dict[name] for name in df_xs.columns.names)]
    else:
        df_xs = df_xs.loc[:, list(loc_dict.values())[0]]

    # df_xs = df_xs.dropna(how='any', axis=1) # pre 7 march 2024
    df_xs = df_xs.dropna(how='all', axis=1) # post 7 march 2024

    for operator in ['subtract', 'divide', 'multiply', 'add']:
        if not operator in plot_setup.keys():
            continue
        
        if isinstance(plot_setup[operator], (str)):
        
            if '=' in plot_setup[operator]:
                col, val = [s.strip() for s in plot_setup[operator].split('=')]
                val = eval(val)
                
                # TODO
                if col in df_xs.columns.names:
                    # df_xs_subtract = df_xs.xs(val, axis=1, level=col)
                    print('needs fix')
                
                if col in df_xs.index.names:
                    
                    if operator == 'divide':
                        df_xs = df_xs.div(df_xs.loc[val]).drop(val)
                    if operator == 'multiply':
                        df_xs = df_xs.mul(df_xs.loc[val]).drop(val)
                    if operator == 'subtract':
                        df_xs = df_xs.sub(df_xs.loc[val]).drop(val)
                         
                
            elif not '=' in plot_setup[operator]:
                if operator == 'divide':
                    df_xs = df_xs / df_xs.columns.get_level_values(plot_setup[operator])
                if operator == 'multiply':
                    df_xs = df_xs * df_xs.columns.get_level_values(plot_setup[operator])
                if operator == 'subtract':
                    df_xs = df_xs - df_xs.columns.get_level_values(plot_setup[operator])
                if operator == 'add':
                    df_xs = df_xs + df_xs.columns.get_level_values(plot_setup[operator])
                    
        elif isinstance(plot_setup[operator], (float, int)):
            if isinstance(df_xs.values.dtype, (int, float)):
                if operator == 'divide':
                    df_xs = df_xs / plot_setup[operator]
                if operator == 'multiply':
                    df_xs = df_xs * plot_setup[operator]
                if operator == 'subtract':
                    df_xs = df_xs - plot_setup[operator]
                if operator == 'add':
                    df_xs = df_xs + plot_setup[operator]
            else:
                is_iterable = isinstance(df_xs.iloc[0, 0], (list, tuple))
                
                if is_iterable:
                    
                    if operator == 'divide':
                        df_xs = df_xs.applymap(lambda x: [item / plot_setup[operator] for item in x])
                    if operator == 'multiply':
                        df_xs = df_xs.applymap(lambda x: [item * plot_setup[operator] for item in x])
                    if operator == 'subtract':
                        df_xs = df_xs.applymap(lambda x: [item - plot_setup[operator] for item in x])
                    if operator == 'add':
                        df_xs = df_xs.applymap(lambda x: [item + plot_setup[operator] for item in x])
                    
                else:
                    if operator == 'divide':
                        df_xs = df_xs / plot_setup[operator]
                    if operator == 'multiply':
                        df_xs = df_xs * plot_setup[operator]
                    if operator == 'subtract':
                        df_xs = df_xs - plot_setup[operator]
                    if operator == 'add':
                        df_xs = df_xs + plot_setup[operator]
                    
                    
    if sort:
        df_xs = df_xs.sort_index(axis=1)
        
    if df_xs.empty:
        print('df empty.')
        
    return df_xs


def plot_heatmap(df, plots_setup):
    for plot_name, plot_setup in plots_setup.items():
        
        df_xs = filter_dataframe(df, plot_setup)
        if df_xs is None or df_xs.empty:
            continue
        
        data_2d = np.array(df_xs.iloc[:, 0].values.tolist()).T
        
        data_2d = data_2d[plot_setup.get('slice', slice(None))] # slice data, if slicer is given
        
        plt.figure(figsize=(7, 2))
        
        ax = plt.gca()
        vmin, vmax = plot_setup.get('ylim', [None, None])
        cmap = plot_setup.get('cmap', 'viridis')
        im = ax.imshow(data_2d, vmin=vmin, vmax=vmax, cmap=cmap, aspect='auto')
        
        title = plot_setup.get('title', plot_name)
        
        num_ticks = 6
        # x_tick_positions = [i * (len(df_xs.index) - 1) / (num_ticks - 1) for i in range(num_ticks)]
        # x_tick_positions = [0, 16, 32, 48, 64, 80]
        # x_tick_positions = [0, 12, 24, 36, 48, 60, 72, 84]
        segments = 4
        # x_tick_positions = [int((i+1/2)*len(df_xs)/segments) for i in range(segments)]
        x_tick_positions = [int((i)*len(df_xs)/segments) for i in range(segments)]
        plt.xticks(x_tick_positions, [df_xs.index.get_level_values('datetime').strftime('%H:%M')[int(pos)] for pos in x_tick_positions])
        
        if ['north' in col.lower() for col in df.columns.unique('location')]:
            y_labels = ('N-side', 'S-side')
        elif ['east' in col.lower() for col in df.columns.unique('location')]:
            y_labels = ('W-side', 'E-side')
        else:
            y_labels = ('A', 'B')
            
        plt.xlabel('Local time [H]')
        plt.ylabel(f'Street location \n(canyon slice)')
        
        # plt.yticks((0-0.5, data_2d.shape[0]-0.5), labels=y_labels)
        plt.yticks((0-0.5+0.5, data_2d.shape[0]-0.5-0.5), labels=y_labels)
        plt.tick_params(
            axis='y',          # changes apply to the x-axis
            which='both',      # both major and minor ticks are affected
            left=False,      # ticks along the bottom edge are off
            # top=False,         # ticks along the top edge are off
            # labelbottom=False
            ) # labels along the bottom edge are off
        plt.grid(None)
        
        # create an axes on the right side of ax. The width of cax will be 5%
        # of ax and the padding between cax and ax will be fixed at 0.05 inch.
        divider = make_axes_locatable(ax)
        cax = divider.append_axes("right", size="5%", pad=0.05)
        
        colorbar_label = plot_setup.get('colorbar_label', '')
        clb = plt.colorbar(im, cax=cax, label=colorbar_label)
        # clb.ax.set_title(df_xs.columns.get_level_values('variable')[0])
        
        
        
        if not plot_name.startswith('db') and SAVE:
            if not os.path.isdir(f'postprocessing plots/{folder}'):
                os.mkdir(f'postprocessing plots/{folder}')
            
            plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(plot_name)}.png', bbox_inches="tight")
            
        # if TITLE:
        ax.set_title(label=plot_name)
            
        plt.show()
        


def plot_lines(df, plots_setup, line_properties, sort=True, add_to_name=False, legend_kwargs={}, **kwargs):
    
    for plot_name, plot_setup in plots_setup.items():
        
        df_xs = filter_dataframe(df, plot_setup, sort=sort)
        
        if df_xs is None: 
            plot_name_0 = plot_name.split("\n")[0]
            print(f'Missing data for plot: "{plot_name_0}"')
            continue
           
        # color = plot_setup.get('color', [line_properties['variable'].get(var, 'tab:grey') for var in df_xs.columns.get_level_values('variable')])
        # color = [line_properties['variable'].get(var, 'tab:grey') for var in df_xs.columns.get_level_values('variable')]
        
        # Set linestyle
        if 'linestyle' in plot_setup.keys():
            if isinstance(plot_setup['linestyle'], (list, tuple)):
                linestyle = plot_setup['linestyle']
            else:
                linestyle = [line_properties[plot_setup['linestyle']]['linestyle'].get(col, '-') for col in df_xs.columns.get_level_values(plot_setup['linestyle'])]
        else: 
            linestyle = ['-'] * df_xs.shape[1]
            
            
        # Set linewidth
        if 'linewidth' in plot_setup.keys():
            if isinstance(plot_setup['linewidth'], (list, tuple)):
                linewidth = plot_setup['linewidth']
            else:
                linewidth = [line_properties[plot_setup['linewidth']]['linewidth'].get(col, 1) for col in df_xs.columns.get_level_values(plot_setup['linewidth'])]
        else: 
            linewidth = [1] * df_xs.shape[1]
            
            
        # Set color
        if 'color' in plot_setup.keys():
            if isinstance(plot_setup['color'], (list, tuple)):
                color = plot_setup['color']
            else:
                color = [line_properties[plot_setup['color']]['color'].get(col, 'tab:blue') for col in df_xs.columns.get_level_values(plot_setup['color'])]
        else: 
            color = ['tab:blue'] * df_xs.shape[1]
            
            
        if plot_setup.get('legend', False): #!!! True to False 6 oct
            df_xs = df_xs.droplevel([l for l in df_xs.columns.names if l not in plot_setup['legend']], axis=1)
    
        
    
        if 'y_scale' in plot_setup.keys():
            df_xs = df_xs * plot_setup['y_scale']
    
        if plot_name.startswith('db'):
            pass
        
        if 'marker' in plot_setup.keys():
            markers = [markers_dict.get(loc, '') for loc in df_xs.columns.get_level_values('variable')]
            markersize = [markersize_dict.get(m, 15) for m in markers]
            markeredgewidth = [markeredgewidth_dict.get(m, 0.5) for m in markers]
            
        figsize = kwargs.get('figsize', (5,3))
        
        fig, ax = plt.subplots(figsize=figsize)
        
        if plot_setup.get('interpolate', False):
            df_xs = df_xs.interpolate()
        
        df_xs.plot(
            ax=ax,
            color=color, 
            markeredgecolor='white',
            markevery=[0, -1],
            fillstyle='full',
            )
        
        if plot_setup.get('fill_between', False):
            fill_color = plot_setup.get('fill_color', colors[::2])
            fill_zorder = plot_setup.get('fill_zorder', [1]*len(plot_setup['fill_between']))
            
            for i, (col1, col2) in enumerate(plot_setup['fill_between']):
                # if isinstance(df_xs.columns, pd.core.indexes.multi.MultiIndex):
                #     df_col1 = df_xs.xs(col1, axis=1, level='variable')
                #     df_col2 = df_xs.xs(col2, axis=1, level='variable')
                # else:
                #     df_col1 = df_xs[col1]
                #     df_col2 = df_xs[col2]
                
                ax.fill_between(df_xs.index, df_xs[col1], df_xs[col2], 
                                color=fill_color[i], 
                                alpha=0.8,
                                zorder=fill_zorder[i], label=f'range{col1}')
        
        # Adjust the line widths in the plot
        for i, l in enumerate(ax.lines):
            if plot_setup['x'] == 't':
                plt.setp(l, linewidth=linewidth[i])
            plt.setp(l, linestyle=linestyle[i])
            
            if 'marker' in plot_setup.keys():
                plt.setp(l, marker=markers[i], markersize=markersize[i], markeredgewidth=markeredgewidth[i], markerfacecolor=color[i])
        
        if 'compare effective albedo' in plot_name.lower():
            if 'model 1' in folder.lower():
                if 1.0 in plot_setup['xs']['hw']:
                    model = 'model 1'
                if 0.0 in plot_setup['xs']['hw']:
                    model = 'model 0'
            if 'model 2' in folder.lower():
                model = 'model 2'
            if 'model 3' in folder.lower():
                model = 'model 3'
                
            if plot_setup['date'] == datetime.date(2023, 6, 15):
                aida_summer[model].plot(ax=ax, marker='x', linestyle='', label='Aida exp.')
                linewidth += [0]

                if model == 'model 1':
                    Visser_model1.plot(ax=ax)
                    Fortuniak_model1.plot(ax=ax)
                if model == 'model 2':
                    Visser_model2.plot(ax=ax)
                    Fortuniak_model2.plot(ax=ax)
                if model == 'model 3':
                    Visser_model3.plot(ax=ax)
                    linewidth += [1]

                    Schrijvers_model3.plot(ax=ax)
                    linewidth += [1]
                    pass
                    
                # df_visser[model].plot(ax=ax, marker='o', linestyle='', label='Visser sim.')
            if plot_setup['date'] == datetime.date(2022, 12, 3):
                aida_winter[model].plot(ax=ax, marker='x', linestyle='', label='Aida exp.')
                linewidth += [0]
        
        handles, labels = ax.get_legend_handles_labels()
        labels = plot_setup.get('labels', replace_keys_in_tuple_strings(labels, labels_dict))
        
        if plot_setup.get('print daily sum', False):   
            df_xs_daily = df_xs.mean() * 24 * 60 * 60 / 1e6
            labels = [f'{label}: {df_xs_daily[col]:.2f} MJ/day' for col, label in zip(df_xs.columns, labels)]
            
        
        # # Annotations - to be fixed
        # if plot_setup.get('annotations', False): 
        #     lines = ax.get_lines()
            
        #     coords = []
        #     line_coords = []
            
        #     for i in range(len(labels_renamed)):
        #         # Extract the line data from the DataFrame
        #         line_data = lines[i].get_data()
        #         x_coord = line_data[0][3]  # X-coordinate at the end of the line
        #         y_coord = max(line_data[1])  # Y-coordinate at the maximum value of the line
    
        #         coords.append((x_coord, y_coord))
        #         line_coords.append((x_coord, y_coord - (y_coord - min(line_data[1])) / 2))
            
        #     # Add annotations with lines
        #     for i, annotation in enumerate(labels_renamed):
        #         coord = coords[i]
        #         line_coord = line_coords[i]
        #         plt.annotate(annotation, coord, xytext=line_coord, arrowprops=dict(arrowstyle='-', color='black'), fontsize=10, ha='center')

        if 'order' in plot_setup:  
            handles, labels = [handles[idx] for idx in plot_setup['order']], [labels[idx] for idx in plot_setup['order']]
            
        if plot_setup.get('legend', True):   
            # Create the legend
            
            legend_title = plot_setup.get('legend_title', ax.get_legend().get_title().get_text())
            legend_title = legend_titles.get(legend_title, legend_title)
            
            labels = [', '.join(l) if isinstance(l, tuple) else l for l in labels]
            legend = ax.legend(handles, labels, title=legend_title, handlelength=3.0, **legend_kwargs)
            
            if plot_setup['x'] == 't':
                # Adjust the line widths in the legend
                for handle in legend.legendHandles:
                    handle.set_linewidth(linewidth[legend.legendHandles.index(handle)])
            
        plt.xlabel(plot_setup.get('xlabel', x_label.get(plot_setup['x'], plot_setup['x'])))
        if 'ylabel' in plot_setup:
            plt.ylabel(plot_setup['ylabel'])
        
        if 'xlim' in plot_setup.keys():
            plt.xlim(plot_setup['xlim'])
        
        if 'ylim' in plot_setup.keys():
            plt.ylim(plot_setup['ylim'])
            
        if not plot_setup.get('legend'):
            ax.get_legend().remove()
            
        if plot_setup.get('hline', True):
            hline = plot_setup.get('hline', True)
            if isinstance(plot_setup.get('hline', True), float):
                hline = plot_setup['hline']
            else:
                hline = 0
            plt.axhline(y=hline, color='black', linewidth=0.5, zorder=2)
        
        if plot_setup.get('secondary_xaxis', False):
            if plot_setup['x'] == 't':
                print('Can not create secondary_xaxis from datetime x-axis')
            else:
                x_ticks = ax.get_xticks()
                if plot_setup['x'] == 'svf':
                    x2_ticks = hw_to_svf(x_ticks)
                    
                if plot_setup['x'] == 'hw':
                    x2_ticks = svf_to_hw(x_ticks)
                    
                ax2 = plt.gca().secondary_xaxis(plot_setup['secondary_xaxis'])
                ax2.set_xticks(x2_ticks)
        
        if not plot_name.startswith('db') and SAVE:
            if not os.path.isdir(f'postprocessing plots/{folder}'):
                os.mkdir(f'postprocessing plots/{folder}')
            
            if add_to_name:
                plot_name += '\n ' + add_to_name
            
            plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(plot_name)}.png', bbox_inches="tight")
            
        plt.title(plot_name)
        plt.show()
        
import matplotlib.ticker as ticker
     

def plot_flux_per_timestep(df, plots_setup, locations, average_between_time=False, legend_kwargs={}, **kwargs):
    
    for plot_name, plot_setup in plots_setup.items():
        
        df_xs = filter_dataframe(df, plot_setup)
        
        if average_between_time:
            df_xs = df_xs.mean(0)
        
        if df_xs is None: 
            plot_name_0 = plot_name.split("\n")[0]
            print(f'Missing data for plot: "{plot_name_0}"')
            continue
    
        if plot_setup.get('legend', False):
            df_xs = df_xs.droplevel([l for l in df_xs.columns.names if l not in plot_setup['legend'] + ['location']], axis=1)
    
        for t in range(len(df_xs)):
            
            if 'show_minutes' in plot_setup.keys():
                if df.index.get_level_values('datetime')[t].minute not in plot_setup['show_minutes']:
                    continue
            
            figsize = kwargs.get('figsize', (5,3))
            
            fig, ax = plt.subplots(figsize=figsize)
            
            if len(plot_setup['legend']) > 1:
                raise ValueError("plot_setup['legend'] can have max length of 1.")
                
            level = plot_setup['legend'][0]
            keys = df_xs.columns.unique(level).values
            
            if 'color' in plot_setup.keys():
                if isinstance(plot_setup['color'], (list, tuple)):
                    color = plot_setup['color']
                else:
                    color = [line_properties[plot_setup['color']]['color'].get(col, 'tab:blue') for col in df_xs.columns.unique(plot_setup['color'])]
            else: 
                color = ['tab:blue'] * df_xs.shape[1]
            
            
            if 'linestyle' in plot_setup.keys():
                if isinstance(plot_setup['linestyle'], (list, tuple)):
                    linestyle = plot_setup['linestyle']
                else:
                    linestyle = [line_properties[plot_setup['linestyle']]['linestyle'].get(col, '-') for col in df_xs.columns.get_level_values(plot_setup['linestyle'])]
            else: 
                linestyle = ['-'] * df_xs.shape[1]
            
            for i, key in enumerate(keys):
                df_xs_key = df_xs.xs(key, axis=1, level=level, drop_level=False)
                data = [df_xs_key.xs(loc, axis=1, level='location').iloc[t].values[0] for loc in locations.values() if loc in df_xs_key.columns.get_level_values('location')]
                
                # for given regions, flip the order of the data 
                try:
                    data = [d[::-1] if f else d for d, f in zip(data, flip)]
                except TypeError:
                    continue
                    
                x_values = []
                y_values = []
                
                # Calculate x-values for each sublist
                num_sublists = len(data)
                x_step = len(data) / num_sublists  # Calculate the step size for x
                for j, sublist in enumerate(data):
                    if num_sublists > 1:
                        x_sublist = np.linspace(j * x_step + 0.001, (j+1) * x_step - 0.001, len(sublist), endpoint=True)
                    if num_sublists == 1:
                        x_sublist = np.linspace(j * x_step + 0.001, len(locations) * x_step - 0.001, len(sublist), endpoint=False)
                    x_values.extend(x_sublist)
                    y_values.extend(sublist)
                
                # if 'subtract' in plot_setup.keys():
                #     y_values -= plot_setup['subtract']
                
                # Create a line plot
                ax.plot(x_values, y_values, linestyle=linestyle[i], label=str(key), color=color[i])
                
                # except Exception as e:
                #     print(e)
                #     continue
                
            if 'data' not in locals():
                print('No data found.')
                return
                
            # Set x-axis labels
            offset=0.5
            ax.xaxis.set_minor_locator(ticker.FixedLocator(np.linspace(0+offset, len(data)+offset, num_sublists + 1)))
            ax.xaxis.set_minor_formatter(ticker.FixedFormatter(list(locations.keys())))
            ax.tick_params(which='minor', length=0)
            
            ax.xaxis.set_major_locator(ticker.FixedLocator(np.linspace(0, len(data), num_sublists + 1)))
            ax.xaxis.set_major_formatter(ticker.FixedFormatter(['','','','','']))
            
            # Add labels and title
            plt.xlabel(plot_setup.get('xlabel', x_label.get(plot_setup['x'], plot_setup['x'])))
            if 'ylabel' in plot_setup:
                plt.ylabel(plot_setup['ylabel'])
                
            if 'xlim' in plot_setup.keys():
                plt.xlim(plot_setup['xlim'])
            if 'ylim' in plot_setup:
                plt.ylim(plot_setup['ylim'])
                
            # Show the plot
            plt.grid(True)
            
            handles, labels = ax.get_legend_handles_labels()
            labels = plot_setup.get('labels', replace_keys_in_tuple_strings(labels, labels_dict))
        
            
            if plot_setup.get('legend', False) and not legend_kwargs is None:   
                # Create the legend
                
                legend_title = plot_setup.get('legend_title', '')
                legend_title = legend_titles.get(legend_title, legend_title)
                
                labels = [', '.join(l) if isinstance(l, tuple) else l for l in labels]
                legend = ax.legend(handles, labels, title=legend_title, handlelength=3.0, **legend_kwargs)
            
            # if plot_setup.get('legend', False) and not legend_kwargs is None:
            #     plt.legend(title=legend_titles.get(level, level), **legend_kwargs)
                
            if not plot_setup.get('legend'):
                ax.get_legend().remove()
                
            if plot_setup.get('hline', True):
                hline = plot_setup.get('hline', True)
                if isinstance(plot_setup.get('hline', True), float):
                    hline = plot_setup['hline']
                else:
                    hline = 0
                plt.axhline(y=hline, color='black', linewidth=0.5, zorder=2)
                
            if not plot_name.startswith('db') and SAVE:
                if not os.path.isdir(f'postprocessing plots/{folder}'):
                    os.mkdir(f'postprocessing plots/{folder}')
                plot_name_save = plot_name + str(df_xs.index.get_level_values("datetime")[t])
                plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(plot_name_save)}.png', bbox_inches="tight")
                
            plt.title(f'{plot_name} \n {df_xs.index.get_level_values("datetime")[t]}')
                
            plt.show()
            
            
def calculate_weights_inverse_density(time_stamps):
    kde = gaussian_kde(time_stamps)
    density_values = kde(time_stamps)
    inverse_density_values = 1 / density_values
    normalized_weights = inverse_density_values / np.sum(inverse_density_values)
    return normalized_weights


def get_table_from_sheet(data_sheet, name, date=None, index='Time', dropna=True):
    data = data_sheet[data_sheet.tables[name].ref]
    rows_list = [[cell.value for cell in row] for row in data]
    _df_load = pd.DataFrame(data=rows_list[1:], columns=rows_list[0])
    
    _df_load = _df_load.set_index(index)
    if dropna:
        _df_load = _df_load.dropna(axis=0)
     
    if date:
        _df_load.index = [datetime.datetime.combine(date, time_obj) for time_obj in _df_load.index] # add date
    
    # df_load = pd.DataFrame(data=_df_load.values, columns=_df_load.columns, index=_df_load.index.values)
    
    return _df_load


def reindex_dataframe(df, new_timestamps):
    add_timestamps = new_timestamps[~new_timestamps.isin(df.index)]
    df_new_timestamps = df.reindex(add_timestamps)
    df_concatenated = pd.concat([df, df_new_timestamps]).sort_index().interpolate(method='time')
    df_concatenated = df_concatenated.groupby(level=0).sum()
    
    df_result = df_concatenated.reindex(new_timestamps)
    
    return df_result

def hw_to_svf(hw):
    
    return np.sqrt(hw**2 + 1) - hw

def svf_to_hw(svf):
    return (1-svf**2)/(2*svf)
    
# %% Load data, set folder

# folder = 'CUSTOM - East-West Canyon 12m apart - 100m'
# folder = 'CUSTOM - North-South Canyon 12m apart - 100m'
# folder = 'CUSTOM - North-South Canyon 24m apart - 200m - FIXED_TEMP=False'
# folder = 'CUSTOM - East-West Canyon 24m apart - 200m - FIXED_TEMP=False'
# folder = 'CUSTOM - Single chimney model'
# folder = 'CUSTOM - Model 1 +2 HD'
# folder = 'CUSTOM - Flat Roof'
# folder = 'REAL - Rotterdam Oude Westen - 240m'
# # folder = 'REAL - Almere Faunabuurt - 120m'
# folder = 'REAL - Rotterdam Airport Gras - 50m'
# # folder = 'REAL - Almere Centrum - 180m'
# # folder = 'REAL - Rotterdam Beukelsdijk - 100m'
# folder = 'REAL - Parking Lot - 50m'
# folder = 'REAL - Rotterdam Oude Westen Zuid - 200m'
# folder = 'REAL - Rotterdam Oude Westen Kogelvangerstraat - 180m'
# folder = 'REAL - Rotterdam Bloemhof - 120m'
# folder = 'REAL - Rotterdam Oude Westen Adrianaplein - 120m'
# folder = 'REAL - Rotterdam Pendrecht - 24/0m - no trees'
# folder = 'REAL - Rotterdam Oude Westen - 170m - better VF'
# folder = 'REAL - Rotterdam Oude Westen - 170m - no trees'
# folder = 'REAL - Rotterdam Bloemhof - 180m'
# folder = 'REAL - Rotterdam Oude Westen Adrianaplein - 120m'

# folder = 'CUSTOM - Model 2 new'
# folder = 'CUSTOM - Model 3 - Japan - 20m apart - 200m'
folder = 'CUSTOM - Model 2 - Delft - 20m apart - 200m'
# folder = 'CUSTOM - Model 2 - Delft - 20m apart - 200m'
# folder = 'CUSTOM - Model 3 - Japan - 20m apart - 100m'
# folder = 'CUSTOM - Model 3 - Japan - 20m apart - 200m'
# folder = 'CUSTOM - Model 2 - Japan - 20m apart - 100m'

# folder = 'CUSTOM - Single chimney model' # validation

# folder = 'REAL - LCZ 1 - EW - Amsterdam - Centrum - Enge Lombardsteeg - 60m'
# folder = 'REAL - LCZ 5 - NS - Rotterdam - 110-Morgen - Herastraat - 140m'
# folder = 'REAL - LCZ 5 - Rotterdam - 110-Morgen - Herastraat - 140m'
# folder = 'REAL - LCZ 3 - Rotterdam - Oud Mathenesse - Finsestraat - 150m'

# folder = 'REAL - Rotterdam Overschie Kleinpolder - 240m'
# folder = 'REAL - Rotterdam Oude Westen Kogelvangerstraat - 180m'
# folder = 'REAL - Rotterdam Pendrecht - 220m - no trees'
# folder = 'REAL - Parking Lot - 50m'
# folder = 'REAL - Rotterdam Bloemhof - 180m'
# folder = 'CUSTOM - Model 3 - Delft - 20m apart - 200m'
# folder = 'REAL - Rotterdam Nieuwe Westen Ruilstraat - 240m'
# folder = 'REAL - Rotterdam Homerusbuurt - 200m'
# folder = 'REAL - Rotterdam - LCZ 2 - Oude Westen - Adrianastraat-Bajnonetstraat - 100m'
# folder = 'REAL - Rotterdam (Overschie) - LCZ5 - Kleinpolder - Lemkensstraat - 160m'
# folder = 'REAL - LCZ 2 - Rotterdam - Nieuwe Westen - Gerrit vd Lindestraat - 160m'

if not os.path.isdir(f'postprocessing plots/{folder}'):
    os.mkdir(f'postprocessing plots/{folder}')


# %% Load data

filename = 'results_log.pkl'

df = pd.read_pickle(f'demos/{folder}/{filename}').sort_index(axis=1).sort_index(axis=0)
df = df.replace('', np.nan)
df = df.dropna(how='all', axis=1)

sim_name = 'Sim.'

try:
    if 'CUSTOM' in folder:

        if 'north' in folder.lower() or 'model 1' in folder.lower():
            dir_a = 'east'
            dir_b = 'west'
            slice_name = ' (y-slice)'
            
        elif 'east' in folder.lower() or 'model 2' in folder.lower():
            dir_a = 'north'
            dir_b = 'south'
            slice_name = ' (x-slice)'
            
        elif 'well' in folder.lower() or 'single' in folder.lower():
            dir_a = 'east'
            dir_b = 'west'
            dir_c = 'south'
            dir_d = 'north'
            slice_name = ''
            
        else:
            raise ValueError('Could not detect if canyon is N-S or E-W')
            
        cols_walls = [f'{dir_a}-facing wall{slice_name}',
                      f'{dir_b}-facing wall{slice_name}']
        
        if 'well' in folder.lower():
            cols_walls.append(f'{dir_c}-facing wall{slice_name}')
            cols_walls.append(f'{dir_d}-facing wall{slice_name}')
            
        cols_total = cols_walls + [f'canyon street{slice_name}']
            
        df_sum_walls = pd.concat({'walls': (df.xs(cols_walls[0], axis=1, level='location') + df.xs(cols_walls[1], axis=1, level='location')).xs('results sum', 1, 'category', drop_level=False)}, names=['location'], axis=1).reorder_levels(order=df.columns.names, axis=1)
        df_sum_total = pd.concat({'total': (df.xs(cols_walls[0], axis=1, level='location') + df.xs(cols_walls[1], axis=1, level='location') + df.xs(f'canyon street{slice_name}', axis=1, level='location')).xs('results sum', 1, 'category', drop_level=False)}, names=['location'], axis=1).reorder_levels(order=df.columns.names, axis=1)
        df = pd.concat([df, df_sum_walls, df_sum_total], axis=1)
        
        # add columns with numbers of elements per surface 
        
        df_nr_els_cols_walls = pd.concat({'walls': (df.xs(cols_walls[0], axis=1, level='location').xs('nr els',1,'category', drop_level=False)) + df.xs(cols_walls[1], axis=1, level='location').xs('nr els',1,'category', drop_level=False)}, names=['location'], axis=1).reorder_levels(order=df.columns.names, axis=1)
        df_nr_els_total = pd.concat({'total': (df.xs(cols_walls[0], axis=1, level='location').xs('nr els',1,'category', drop_level=False)) + df.xs(cols_walls[1], axis=1, level='location').xs('nr els',1,'category', drop_level=False) + df.xs('nr els',1,'category', drop_level=False).xs(f'canyon street{slice_name}', axis=1, level='location')}, names=['location'], axis=1).reorder_levels(order=df.columns.names, axis=1)
        
        # df_nr_els_cols_walls = pd.concat({'walls': df.loc[:, ('nr els', '', slice(None), '', '', cols_walls, slice(None), slice(None))].groupby(['category', 'emissivity sky', 'hw', 'albedo', 'emissivity', 'skin thickness', 'vol. heat cap.', 'wind speed', 'variable', 'res'], axis=1).sum()}, names=['location'], axis=1).reorder_levels(order=df.columns.names, axis=1)
        # df_nr_els_total = pd.concat({'total': df.loc[:, ('nr els', '', slice(None), '', '', cols_total, slice(None), slice(None))].groupby(['category', 'emissivity sky', 'hw', 'albedo', 'emissivity', 'skin thickness', 'vol. heat cap.', 'wind speed', 'variable', 'res'], axis=1).sum()}, names=['location'], axis=1).reorder_levels(order=df.columns.names, axis=1)
        df = pd.concat([df, df_nr_els_cols_walls, df_nr_els_total], axis=1)
    else:
        raise ValueError

except (ValueError, KeyError):
    pass
    slice_name = ''

locs_lst = df['results sum'].columns.get_level_values('location')
hws_lst = df['results sum'].columns.get_level_values('hw')
res_lst = df['results sum'].columns.get_level_values('res')

try:
    nr_els_dict = dict(df['nr els'].iloc[0, 0:].droplevel(('albedo', 'emissivity', 'skin thickness', 'vol. heat cap.', 'wind speed', 'variable', 'emissivity sky')))
except:
    nr_els_dict = dict(df['nr els'].iloc[0, 0:].droplevel(('albedo', 'emissivity', 'skin thickness', 'density', 'spec. heat cap.', 'wind speed', 'variable', 'emissivity sky')))
# for hw, res in zip(hws_lst, res_lst):
#     nr_els_dict[(hw, 'bottom', res)] = nr_els_dict[(hw, 'sensors', res)]
#     nr_els_dict[(hw, 'top', res)] = nr_els_dict[(hw, 'sensors', res)]

nr_els_surf_lst = [nr_els_dict.get((hw, loc, res)) for loc, hw, res in zip(locs_lst, hws_lst, res_lst)]
nr_els_ground_lst = [nr_els_dict.get((hw, f'canyon street{slice_name}', res)) for hw, res in zip(hws_lst, res_lst)]

df = pd.concat([df, 
                pd.concat({'results, per m2 surface': df['results sum'] / nr_els_surf_lst}, names=['category'], axis=1), 
                pd.concat({'results, per m2 ground': df['results sum'] / nr_els_ground_lst}, names=['category'], axis=1)], axis=1)


# df = df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_dir']",1,'variable') * (1-df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_dir']",1,'variable').columns.get_level_values('albedo'))

# try:
#     df = pd.concat([df, 
#                 pd.concat({"q_lw['in']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_lw['in_from_els']",1,'variable') + df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_lw['in_from_sky']",1,'variable') }, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1),
#                 ], axis=1)

# except (KeyError, TypeError):
#     pass

# try:
#     df = pd.concat([df, 
#                     pd.concat({"q_sw['abs_dir']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_dir']",1,'variable') * (1-df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_dir']",1,'variable').columns.get_level_values('albedo'))}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1),
#                     pd.concat({"q_sw['abs_diff']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_diff']",1,'variable') * (1-df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_diff']",1,'variable').columns.get_level_values('albedo'))}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1),
#                     pd.concat({"q_sw['abs_from_els']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_from_els']",1,'variable') * (1-df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_from_els']",1,'variable').columns.get_level_values('albedo'))}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1),
#                     pd.concat({"q_sw['abs_from_sky']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_from_sky']",1,'variable') * (1-df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_from_sky']",1,'variable').columns.get_level_values('albedo'))}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1),
#                     pd.concat({"q_lw['abs']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_lw['in']",1,'variable') * (0.93)}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1),
#                     pd.concat({"q_lw['abs_from_els']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_lw['in_from_els']",1,'variable') * (0.93)}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1),
#                     pd.concat({"q_lw['abs_from_sky']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_lw['in_from_sky']",1,'variable') * (0.93)}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1),
#                     ], axis=1)
    
# except (KeyError, TypeError):
#     pass


# concat_operations = [
#     {"name": "q_sw['abs_dir']", "in_key": "q_sw['in_dir']"},
#     {"name": "q_sw['abs_diff']", "in_key": "q_sw['in_diff']"},
#     {"name": "q_sw['abs_from_els']", "in_key": "q_sw['in_from_els']"},
#     {"name": "q_sw['abs_from_sky']", "in_key": "q_sw['in_from_sky']"},
#     {"name": "q_lw['abs']"},
#     {"name": "q_lw['abs_from_els']"},
#     {"name": "q_lw['abs_from_sky']"}
# ]

# dfs_to_concat = []
# for operation in concat_operations:
#     try:
#         if 'in_key' in operation:
#             new_df = pd.concat({operation["name"]: df.xs('results, per m2 surface',1,'category',drop_level=False).xs(operation["in_key"],1,'variable') * (1-df.xs('results, per m2 surface',1,'category',drop_level=False).xs(operation["in_key"],1,'variable').columns.get_level_values('albedo'))}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1)
#         else:
#             new_df = pd.concat({operation["name"]: df.xs('results, per m2 surface',1,'category',drop_level=False).xs(operation["name"].split("[")[1].split("]")[0],1,'variable') * (0.93)}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1)
#         dfs_to_concat.append(new_df)
#     except (KeyError, TypeError):
#         pass

# try:
#     df = pd.concat([df] + dfs_to_concat, axis=1)
# except (KeyError, TypeError):
#     pass


try:
    df_concat_1 = pd.concat({"q_sw['abs_dir']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_dir']",1,'variable') * (1-df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_dir']",1,'variable').columns.get_level_values('albedo'))}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1)
except (KeyError, TypeError):
    df_concat_1 = pd.DataFrame()

try:
    df_concat_2 = pd.concat({"q_sw['abs_diff']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_diff']",1,'variable') * (1-df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_diff']",1,'variable').columns.get_level_values('albedo'))}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1)
except (KeyError, TypeError):
    df_concat_2 = pd.DataFrame()

try:
    df_concat_3 = pd.concat({"q_sw['abs_from_els']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_from_els']",1,'variable') * (1-df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_from_els']",1,'variable').columns.get_level_values('albedo'))}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1)
except (KeyError, TypeError):
    df_concat_3 = pd.DataFrame()

try:
    df_concat_4 = pd.concat({"q_sw['abs_from_sky']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_from_sky']",1,'variable') * (1-df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_sw['in_from_sky']",1,'variable').columns.get_level_values('albedo'))}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1)
except (KeyError, TypeError):
    df_concat_4 = pd.DataFrame()

try:
    df_concat_5 = pd.concat({"q_lw['abs']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_lw['in']",1,'variable') * (0.93)}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1)
except (KeyError, TypeError):
    df_concat_5 = pd.DataFrame()

try:
    df_concat_6 = pd.concat({"q_lw['abs_from_els']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_lw['in_from_els']",1,'variable') * (0.93)}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1)
except (KeyError, TypeError):
    df_concat_6 = pd.DataFrame()

try:
    df_concat_7 = pd.concat({"q_lw['abs_from_sky']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_lw['in_from_sky']",1,'variable') * (0.93)}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1)
except (KeyError, TypeError):
    df_concat_7 = pd.DataFrame()
    
    
try:
    df_concat_8 = pd.concat({"q_lw['in']": df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_lw['in_from_sky']",1,'variable') + df.xs('results, per m2 surface',1,'category',drop_level=False).xs("q_lw['in_from_els']",1,'variable')}, names=['variable'], axis=1).reorder_levels(df.columns.names, axis=1)
except (KeyError, TypeError):
    df_concat_8 = pd.DataFrame()
    
try:
    df = pd.concat([df, df_concat_1, df_concat_2, df_concat_3, df_concat_4, df_concat_5, df_concat_6, df_concat_7, df_concat_8], axis=1)
except (KeyError, TypeError):
    pass

df = df.loc[:, ~df.columns.duplicated()] # remove duplicates that could arise due to the added 'abs' columns

# print('Drop emm sky other than 0.8')
# df=df.xs(0.8, axis=1, level='emissivity sky', drop_level=False) #!!!

freq = pd.infer_freq(df.index.get_level_values('datetime'))

if 'model' in folder.lower():
    if 'chimney' in folder.lower() or '2023' in folder.lower():
        # date = datetime.date(2023, 6, 15)
        date = datetime.date(2023, 9, 7)
    elif 'japan' in folder.lower():
        date = datetime.date(2023, 6, 15) 
    else:
        date = datetime.date(2019, 7, 23)
    # date = datetime.date(2023, 9, 7)

else:
    date = datetime.date(2019, 7, 23)

if not 'japan' in folder.lower():
    bsrn_file = f'data/CAB_radiation_{date.strftime("%Y")}-{date.strftime("%m")}.tab'
    with open(bsrn_file, 'r') as file:
        skiprows = next(i for i, line in enumerate(file, 1) if '*/' in line)
    
    
    bsrn_df = pd.read_csv(bsrn_file, sep='\t', skiprows=skiprows, index_col='Date/Time', parse_dates=['Date/Time'])
    bsrn_df.index = bsrn_df.index.tz_localize('UTC').tz_convert('Europe/Amsterdam').tz_localize(None)
    
    bsrn_df = bsrn_df.resample(freq).mean().interpolate() 
    bsrn_df = bsrn_df.loc[df.index[0][0]: df.index[-1][0]]
    bsrn_df['DIR [W/m**2]']=bsrn_df['DIR [W/m**2]'].where(bsrn_df['DIR [W/m**2]']>0,0)
    bsrn_df['DIF [W/m**2]']=bsrn_df['DIF [W/m**2]'].where(bsrn_df['DIF [W/m**2]']>0,0)
    
    bsrn_df['T ground [°C]'] = (bsrn_df['LWU [W/m**2]']/0.98/(5.67*10**(-8)))**(1/4)-273.15
    
    K = (bsrn_df['DIR [W/m**2]']*(0.06*np.cos(np.deg2rad(df.index.get_level_values('zenith')))+0.22*np.sin(np.deg2rad(df.index.get_level_values('zenith'))))+0.5*(bsrn_df['DIF [W/m**2]'] + bsrn_df ['SWU [W/m**2]'])).where(df.index.get_level_values('zenith')<90, 0)
    L = 0.5*(bsrn_df['LWD [W/m**2]'] + bsrn_df ['LWU [W/m**2]'])
    
    S_str = 0.97*L + 0.7*K
    bsrn_df['MRT'] = (S_str/(0.97*5.67*10**(-8)))**(1/4)-273.15


if 'japan' in folder.lower():
    print('Drop albedo other than 0.4, (disabled)')

hws = sorted(df.xs('results sum',1,'category').columns.unique('hw'))
hws = [0, 0.5, 1.0, 2, 4]
ress = sorted(df['results sum'].columns.unique('res'))
    
if 'real' in folder.lower():
    level = 'skin thickness'
    
elif 'custom' in folder.lower():
    level = 'hw' #!!! klopt niet helemaal want hw zit al in de kwargs

keys = df.columns.unique(level)


# Load min mean max df EW and NS

# %%% Loading T vs SVF and plotting

try:
    # list of file names
    file_names = ['CUSTOM - Model 1 - Delft - 20m apart - 200m', 'CUSTOM - Model 2 - Delft - 20m apart - 200m']
    
    # Construct a dictionary of DataFrames using a dictionary comprehension
    dfs_dict = {f: pd.read_pickle(os.path.join('demos', f, 'df_svf.pkl')) for f in file_names}
    
    # Concatenate all DataFrames in the dictionary along the columns axis with keys parameter
    loaded_dfs = pd.concat(dfs_dict, axis=1, names=['folder']).interpolate()  
except:
    print('Could not load min mean max dfs.')
    
time_names = ['All_day', 'Evening', 'Around_Noon', 'Before_Sunrise']

# Temperature_Results=load_workbook('demos/Temperature Results.xlsx')['Data']
# dfs_dict = {t: get_table_from_sheet(Temperature_Results, name=t, index=['svf', 'name']) for t in time_names}
# T_dfs = pd.concat(dfs_dict, axis=1, names=['time_interval'])
# new_time_names = {'All_day': '24-H (00:00-24:00)',
#                   'Evening':'Evening (18:00-20:00)',
#                   'Around_Noon':'Around Noon (12:00 - 16:00)',
#                   'Before_Sunrise': 'Before sunrise (04:00)'}

df_results = pd.read_csv('demos/df_real.csv', header=[0, 1], index_col=0)
# svf_dict = {f: pd.read_csv(f'demos/{f}/svf_log.csv', index_col=(0,1)).loc[(1.0,1.0)]['street mean'] for f in df_results.index.get_level_values('folder')}

svf_dict = {f: pd.read_csv(f'demos/{f}/svf_log.csv', index_col=(0,1)).loc[(1.0, 0.5)]['street mean'] if (1.0, 0.5) in pd.read_csv(f'demos/{f}/svf_log.csv', index_col=(0,1)).index else pd.read_csv(f'demos/{f}/svf_log.csv', index_col=(0,1)).loc[(1.0,1.0)]['street mean'] for f in df_results.index.get_level_values('folder')}

df_results.set_index(pd.Index([svf_dict[i] for i in df_results.index], name='svf'), append=True, inplace=True)

df_results.set_index(pd.Index([folder.split(' - ')[1].replace('LCZ ','') for folder in df_results.index.get_level_values('folder')], name='LCZ'), append=True, inplace=True)
df_results.set_index(pd.Index([folder.split(' - ')[2] for folder in df_results.index.get_level_values('folder')], name='Orientation'), append=True, inplace=True)



# %%%% Plot idealised street T, min mean max


# T_dfs.rename(columns=new_time_names, level='time_interval', inplace=True)

for time_name, time_interval in time_dict.items():
    for stat in ['mean', 'min', 'max']:
        
    
    
        loaded_dfs.xs(stat, 1, 'statistics').xs(time_name, 1, 'time_interval').plot(xlim=(1.05, -0.05))
        plt.title(f'{time_name} \n {stat}')
        plt.show()
   

# %%%% Plot idealised vs simulated ave street T    
   
color_dict = {'NS': colors[0], 'EW': colors[1], 'NS-EW': colors[2], }
marker_dict = {'NS': '$||$', 'EW': '$=$', 'NS-EW': '$+$', }


plot_min_max = 'errorbar'
idealised_canyons_style = 'fill_between'
idealised_canyons_style = False


# legend_lst = ['Ideal. NS canyon', 'Ideal. EW canyon', 'Real NS geom.', 'Real EW geom.', 'Real Crossroad geom.']


for plot_min_max in ['errorbar', False]:
    # for legend in [False, 'inside', 'on side']:        
    for legend in ['on side', False]:        
    
        for time_name, time_interval in time_dict.items():
            for stat in ['mean', 'min', 'max'] if not plot_min_max else ['mean']:
                for orientation in df_results.index.unique('Orientation'):
                    fig, ax = plt.subplots(figsize=(5,3))
                    ax.set_xlim((1.05, -0.05))
                    if plot_min_max == 'errorbar':
                        # NS
                        if idealised_canyons_style == 'fill_between': # fill_between
                            if orientation == 'NS':
                                loaded_dfs_xs = loaded_dfs.xs(time_name, 1, 'time_interval').xs(0.2, 1, 'albedo').xs('CUSTOM - Model 1 - Delft - 20m apart - 200m',1,'folder')
                                ax.fill_between(loaded_dfs_xs.index, loaded_dfs_xs['min'], loaded_dfs_xs['max'], label='Simulated (range)', color=colors[0], alpha=0.2)
                                  
                            if orientation == 'EW':
                                # EW 
                                loaded_dfs_xs = loaded_dfs.xs(time_name, 1, 'time_interval').xs(0.2, 1, 'albedo').xs('CUSTOM - Model 2 - Delft - 20m apart - 200m',1,'folder')
                                ax.fill_between(loaded_dfs_xs.index, loaded_dfs_xs['min'], loaded_dfs_xs['max'], label='Simulated (range)', color=colors[1], alpha=0.2)
                        if idealised_canyons_style == 'lines':
                            
                            loaded_dfs_xs = loaded_dfs.xs(time_name, 1, 'time_interval').xs(0.2, 1, 'albedo').xs('CUSTOM - Model 1 - Delft - 20m apart - 200m',1,'folder')
                            loaded_dfs_xs.loc[:, ('min')].plot(ax=ax, label='Simulated (range)', color=colors[0], lw=0.8, ls='dotted', alpha=0.5)
                            loaded_dfs_xs.loc[:, ('max')].plot(ax=ax, label='Simulated (range)', color=colors[0], lw=0.8, ls='dashed', alpha=0.5)
                              
                            # EW 
                            loaded_dfs_xs = loaded_dfs.xs(time_name, 1, 'time_interval').xs(0.2, 1, 'albedo').xs('CUSTOM - Model 2 - Delft - 20m apart - 200m',1,'folder')
                            loaded_dfs_xs.loc[:, ('min')].plot(ax=ax, label='Simulated (range)', color=colors[1], lw=0.8, ls='dotted', alpha=0.5)
                            loaded_dfs_xs.loc[:, ('max')].plot(ax=ax, label='Simulated (range)', color=colors[1], lw=0.8, ls='dashed', alpha=0.5)
                        
                            
                    else:
                        loaded_dfs.xs(stat, 1, 'statistics').xs(time_name, 1, 'time_interval').xs(0.2, 1, 'albedo').plot(ax=ax)
                        
                    
                        
                    # T_dfs_xs=T_dfs.droplevel('name', 0).xs(time_name,1,'time_interval')
                    
                    colors_markers = [color_dict[o] for o in df_results.index.get_level_values('Orientation')]
                    df_results_xs = df_results.xs(time_name, 1, 'time_name').xs(orientation, 0, 'Orientation').droplevel(('folder','LCZ'))
                    
                    if plot_min_max == 'errorbar':
                        # for o in df_results.index.unique('Orientation'):
                        df_results_xs_orientation = df_results_xs
                            # df_results_xs_orientation = df_results_xs.xs(o, axis=0, level='Orientation')
                        df_results_xs_orientation.plot(
                            ax=ax, 
                            y='T_mean', 
                            yerr=[df_results_xs_orientation['T_mean'] - df_results_xs_orientation['T_min'], df_results_xs_orientation['T_max'] - df_results_xs_orientation['T_mean']], 
                            kind='line', 
                            label='Simulated (mean)', 
                            ls='', 
                            marker='_', 
                            color=color_dict[orientation])
                        ax.plot(df_results_xs_orientation.index, df_results_xs_orientation['T_max'], marker='_', markersize=5, linestyle='none', color=color_dict[orientation])  # Bottom markers
                        ax.plot(df_results_xs_orientation.index, df_results_xs_orientation['T_min'], marker='_', markersize=5, linestyle='none', color=color_dict[orientation])  # Top markers
                    
                    elif plot_min_max == 'fill_between':
                        ax.fill_between(df_results_xs.index, df_results_xs['T_min'], df_results_xs['T_max'], label='Simulated (range)', color=colors[3], alpha=0.6)
                    else:
                        for o in df_results.index.unique('Orientation'):
                            df_results_xs_orientation = df_results_xs.xs(o, axis=0, level='Orientation')
                            ax.scatter(x=df_results_xs_orientation.index, y=df_results_xs_orientation[f'T_{stat}'], marker=marker_dict[orientation], color=color_dict[orientation])
                            
                    ax.set_xlabel('Street SVF  [-]')
                    # ax.set_ylabel(y_label_deg_C)
                    ax.set_ylabel(r'Street temperature [$^\circ$C]')
                    
                    title = f'{time_name} \n {stat} surface temperature \n {orientation}'
                    
                    handles, _ = ax.get_legend_handles_labels()
                    # handles = handles
                    
                    if plot_min_max == 'errorbar':
                        if orientation == 'NS':
                            legend_lst = ['Ideal. NS canyon (range)', 'Real NS geom. (range)']
                        if orientation == 'EW':
                            legend_lst = ['Ideal. EW canyon (range)', 'Real EW geom. (range)']
                        
                    else:
                        legend_lst = ['Ideal. NS canyon', 'Ideal. EW canyon', 'Real NS geom.', 'Real EW geom.', 'Real Crossroad geom.']
    
                    
                    if legend == 'on side':
                        if plot_min_max == 'errorbar':
                            plt.legend(handles=handles, labels=legend_lst, bbox_to_anchor=(1.04, 0.5), loc="center left", borderaxespad=0)
                        else:
                            plt.legend(labels=legend_lst, bbox_to_anchor=(1.04, 0.5), loc="center left", borderaxespad=0)
                        title += f'\n legend {legend}'
                        
                    elif legend == 'inside':
                        plt.legend(legend_lst)
                        title += f'\n legend {legend}'
                        
                    else:
                        ax.get_legend().remove()
                        title += '\n no legend'
                        
                    if plot_min_max:
                        title += f'\n {plot_min_max}'
                    
                    # ax.set_xlim=(1.05, -0.05)
                    if not 'sunrise' in time_name:
                        if stat == 'mean' and not '24-H' in time_name:
                            
                            if not plot_min_max:
                                ax.set_ylim(28, 65)
                            else:
                                pass
                                # ax.set_ylim(Non, 65)
                        else:
                            ax.set_ylim(12, 73)
                    
                    ax.set_ylim(27, 73)
                    
                    plt.savefig(f'postprocessing plots/Chapter 6/{transform_to_filename(title)}', bbox_inches="tight")
                    
                    plt.title(title)
                    plt.show()
    
#%%% check if external forcings match

equals = []
for var in df['external forcings'].columns.unique('variable'):
    _df_var = df.xs(var, axis=1, level='variable').dropna()
    
    equal = _df_var.eq(_df_var.iloc[:, 0], axis=0).all(1).all()
    equals.append(equal)
    
    if equals:
        print(f'Same {var} used for all calculations')
    else:
        print(f'\n\n Different {var} used for some calculations \n\n')
     
if all(equals):
    col0 = {key: value for key, value in zip(df.dropna(axis=1).columns.names, df.dropna(axis=1).columns[-1])}
    df_rads = df.droplevel(('azimuth', 'zenith')).xs(('external forcings', col0['emissivity sky'], col0['hw'], col0['albedo'], col0['emissivity'], col0['res'], ''), axis=1, level=('category', 'emissivity sky', 'hw', 'albedo', 'emissivity', 'res', 'location')).droplevel(['skin thickness', 'vol. heat cap.', 'wind speed'], axis=1)
    df_rads = df_rads.iloc[:, :len(df_rads.columns.unique())]
    df_lw_in = df.droplevel(('azimuth', 'zenith')).xs("q_lw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category').xs(f'floating top{slice_name}', axis=1, level='location').iloc[:, 0]
    
    # fig, ax = plt.subplots()
    df_rads.loc[date.strftime("%Y-%m-%d")].plot()
    # df_lw_in.loc[date.strftime("%Y-%m-%d")].plot(label=r'$L^{\downarrow}_{\mathrm{sky}}$')
    
    plt.xlabel('Local time [H]')
    plt.ylabel(r'Radiation flux [W m$^{-2}$]')
    plt.legend()
    
    # title = f'Radiations at {date}'
    # plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(title)}.png', bbox_inches="tight")
    # plt.title(title)
    plt.show()
    
    # df_lw_in = df.droplevel(('azimuth', 'zenith')).xs("q_lw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category').xs(f'floating top{slice_name}', axis=1, level='location').iloc[:, 0]
    # df_rads['LW_sky']
    
    
# plot radiations

df_rads2 = df.droplevel(('azimuth', 'zenith')).xs('external forcings', axis=1, level='category').dropna(axis=1).droplevel([0, 1, 2, 3, 4, 5, 6, 7, 9], axis=1)
df_rads2 = df_rads2.iloc[:, :len(df_rads2.columns.unique())].loc[date.strftime("%Y-%m-%d")]

fig, ax = plt.subplots()

df_rads2['Dir_hor'].plot(ax=ax, label=r'$K^{\downarrow}_{\mathrm{dir,hor}}$')
df_rads2['Dir_perp'].plot(ax=ax, label=r'$K^{\downarrow}_{\mathrm{dir,\perp}}$')
df_rads2['Diff_hor'].plot(ax=ax, label=r'$K^{\downarrow}_{\mathrm{diff}}$')
df_rads2['G_hor'].plot(ax=ax, label=r'$K^{\downarrow}_{\mathrm{sky}}$')
df_rads2['LW_sky'].plot(ax=ax, label=r'$L^{\downarrow}_{\mathrm{sky}}$')
ax.set_ylim(None, 1050)

# ax.legend([ax.get_lines()[0], ax.right_ax.get_lines()[0]], ['A','B','C'], )
# lines = ax.get_lines() + ax.right_ax.get_lines()
# ax.legend(lines, [l.get_label() for l in lines], bbox_to_anchor=(1.4, 0.75))
ax.legend()

ax.set_ylabel(r'Radiation flux [W m$^{-2}$]')
ax.set_xlabel(r'Datetime')

title = f'Radiations at {date}'
plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(title)}.png', bbox_inches="tight")
plt.title(title)
plt.show()

fig, ax = plt.subplots()
title = f'T_air and windspeed at {date}'

(df_rads2['T_air']-273.15).plot(ax=ax, label=r'$T_{\mathrm{air}}$')
# bsrn_df['T ground [°C]'].loc[date.strftime("%Y-%m-%d")].plot(ax=ax, label=r'$T_{\mathrm{ground}}$')
df_rads2['wind_speed'].plot(ax=ax, label=r'$v$', secondary_y=True, ls='dashed')
ax.right_ax.set_ylim((0, int(df_rads2['wind_speed'].max()+1)))

ax.set_xlabel(r'Datetime')
ax.set_ylabel(r'Temperature [$^\circ$C]')
ax.right_ax.set_ylabel(r'Wind speed [m s$^{-1}$]')

handles,labels = [],[]
for ax in fig.axes:
    for h,l in zip(*ax.get_legend_handles_labels()):
        handles.append(h)
        labels.append(l)

plt.legend(handles,labels)

plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(title)}.png', bbox_inches="tight")
plt.title(title)
plt.show()



# %% Print values for table | 1 moment
# if 'custom' in folder.lower() or True:
print_label = False
    
if True:
    
    # for hw in [1.0]:
    for hw in [0.0, 0.5, 1.0, 4.0]:
        # for alb in [0.0, 0.2, 0.4, 0.6]:
        for alb in [0.2]:
            print_max_min = False
            print(f'\n\n H/W={hw}, albedo={alb} \n')
            # for time in ['04:00', '09:00', '14:00', '19:00']:
            for time in ['04:00', '14:00', '19:00']:
                print(time)
                # df_xs = df.xs(0.02, 1, 'skin thickness').loc[str(date)+f' {time}', (slice(None), slice(None), [0, 0.5, 1, 2, 4], [0.4])]
                try:
                    df_xs = df.loc[str(date)+f' {time}'].xs(1.0,1,'res').xs(hw,1,'hw').xs(alb,1,'albedo')
                    print('res=1')
                except:
                    df_xs = df.loc[str(date)+f' {time}'].xs(2.0,1,'res').xs(hw,1,'hw').xs(alb,1,'albedo')
                    print('res=2')
                print_alb = str((df_xs.xs('floating bottom'+slice_name, axis=1, level='location').xs("q_sw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category') / df_xs.xs('floating top'+slice_name, axis=1, level='location').xs("q_sw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category')).values[0].round(3)).strip('[]')
                print_Kstar = str((df_xs.xs('floating top'+slice_name, axis=1, level='location').xs("q_sw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category') - df_xs.xs('floating bottom'+slice_name, axis=1, level='location').xs("q_sw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category')).values[0].round(0)).strip('[]')
                print_Lstar = str((df_xs.xs('floating top'+slice_name, axis=1, level='location').xs("q_lw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category') - df_xs.xs('floating bottom'+slice_name, axis=1, level='location').xs("q_lw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category')).values[0].round(0)).strip('[]')
                
                
                if 'custom' in folder.lower():
                    print_T_mean = str((df_xs.xs(f'canyon street{slice_name}' , axis=1, level='location')).xs("T", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category').values[0].round(1)-273).strip('[]')
                else:
                    print_T_mean = str((df_xs.xs('ground' , axis=1, level='location').xs("T", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category').values[0].round(1)-273)).strip('[]')
                
                try:
                    print_T_mr = (str((df_xs.xs('street'+slice_name, axis=1, level='location')).xs("T_mr", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(lambda x: sum(x)/len(x)).values[0].round(1)-273).strip('[]'))
                    print_T_max = str((df_xs.xs(f'canyon street{slice_name}' , axis=1, level='location')).xs("T", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(lambda x: max(x)).values[0].round(1)-273).strip('[]')
                    print_T_min = str((df_xs.xs(f'canyon street{slice_name}' , axis=1, level='location')).xs("T", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(lambda x: min(x)).values[0].round(1)-273).strip('[]')
                    print_max_min = True
                        
                    # print(print_T_mr)
                except:
                    pass
                if print_label: 
                    print_alb = 'alb=' + print_alb
                    print_T = f'T_mean='+print_T_mean
                    print_T_mr = 'T_mr_mean='+print_T_mr
                    
                        
                    
                
                
                
                if print_max_min:
                    print(f'alb='+print_alb)
                    print(f'T_mean='+print_T_mean)
                    print(f'T_min='+print_T_min)
                    print(f'T_max='+print_T_max)
                    print(f'T_mr_mean='+print_T_mr)
                    
                else:  
                    print(print_alb, print_T, print_T_mr)
                
                # print(f'T_mr_min '+str((df_xs.xs('street'+slice_name, axis=1, level='location')).xs("T_mr", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(min).values[0]-273).strip('[]'))
                # print(f'T_mr_max '+str((df_xs.xs('street'+slice_name, axis=1, level='location')).xs("T_mr", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(max).values[0]-273).strip('[]'))
                # print(f'T_mr mean {time} = '+str((df.xs('street', axis=1, level='location')).xs("T_mr", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(lambda x: sum(x)/len(x)).values[0]-273))
                print('\n\n')

# %% Print values for table | time interval
# if 'custom' in folder.lower() or True:
print_label = True


try: 
    df_results = pd.read_csv('demos/df_real.csv', header=[0, 1], index_col=0)

except:
    df_results = pd.DataFrame(index=[],
                              columns=pd.MultiIndex.from_tuples([], names=['time_name', 'variable']))

# Assign 'folder' as the name of the index
df_results.index.name = 'folder'
if True:
    
    # for hw in [1.0]:
    for hw in [1.0]:
        # for alb in [0.0, 0.2, 0.4, 0.6]:
        # for alb in [0.0, 0.2, 0.4, 0.6, 1.0]:
        for alb in [0.2]:
            print_max_min = False
            print(f'\n\n H/W={hw}, albedo={alb} \n')
            # for time in ['04:00', '09:00', '14:00', '19:00']:
            for time_name, time_interval in time_dict.items():
                print(time_name)
                # df_xs = df.xs(0.02, 1, 'skin thickness').loc[str(date)+f' {time}', (slice(None), slice(None), [0, 0.5, 1, 2, 4], [0.4])]
                df_xs = df.loc[str(date)].droplevel(('azimuth', 'zenith'))
                df_xs = df_xs.between_time(time_interval[0], time_interval[1])
                df_xs = df_xs.xs(hw,1,'hw')
                try:
                    df_xs = df_xs.xs(alb,1,'albedo')
                except KeyError:
                    pass
                
                try:
                    df_xs = df_xs.xs(0.5, 1, 'res')
                except:
                    try:
                        df_xs = df_xs.xs(1.0,1,'res')
                        # print('res=1')
                    except:
                        df_xs = df_xs.xs(2.0,1,'res')
                        # print('res=2')
                alb = str((df_xs.xs('floating bottom'+slice_name, axis=1, level='location').xs("q_sw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category') / df_xs.xs('floating top'+slice_name, axis=1, level='location').xs("q_sw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category')).mean().values[0].round(3)).strip('[]')
                Kstar = str((df_xs.xs('floating top'+slice_name, axis=1, level='location').xs("q_sw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category') - df_xs.xs('floating bottom'+slice_name, axis=1, level='location').xs("q_sw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category')).mean().values[0].round(0)).strip('[]')
                Lstar = str((df_xs.xs('floating top'+slice_name, axis=1, level='location').xs("q_lw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category') - df_xs.xs('floating bottom'+slice_name, axis=1, level='location').xs("q_lw_sensors['in']", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category')).mean().values[0].round(0)).strip('[]')
                
                
                if 'custom' in folder.lower():
                    T_mean = ((df_xs.xs(f'canyon street{slice_name}' , axis=1, level='location')).xs("T", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category').mean().values[0]-273.15).round(1)
                else:
                    T_mean = ((df_xs.xs('ground cropped' , axis=1, level='location').xs("T", axis=1, level='variable').xs('results, per m2 surface', axis=1, level='category').mean().values[0]-273.15).round(1))
                
                try:
                    T_mr = ((df_xs.xs('street'+slice_name, axis=1, level='location')).xs("T_mr", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(lambda x: sum(x)/len(x)).mean().values[0]-273.15).round(1)
                except:
                    pass
                
                try:
                    
                    T_max = ((df_xs.xs(f'canyon street{slice_name}' , axis=1, level='location')).xs("T", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(lambda x: max(x)).max().values[0]-273.15).round(1)
                    T_min = ((df_xs.xs(f'canyon street{slice_name}' , axis=1, level='location')).xs("T", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(lambda x: min(x)).min().values[0]-273.15).round(1)
                    max_min = True
                except:
                    try:
                        T_max = ((df_xs.xs('ground cropped' , axis=1, level='location')).xs("T", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(lambda x: max(x)).max().values[0]-273.15).round(1)
                        T_min = ((df_xs.xs('ground cropped' , axis=1, level='location')).xs("T", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(lambda x: min(x)).min().values[0]-273.15).round(1) 
                        max_min = True
                    except:
                        pass

                    
                df_results.loc[folder, (time_name, 'Albedo')] = alb
                df_results.loc[folder, (time_name,  'T_mean')] = T_mean
                df_results.loc[folder, (time_name,  'T_min')] = T_min
                df_results.loc[folder, (time_name,  'T_max')] = T_max
                df_results.loc[folder, (time_name,  'T_mr_mean')] = T_mr

                # if label: 
                #     alb = 'alb=' + alb
                #     T = f'T_mean='+T_mean
                #     T_mr = 'T_mr_mean='+T_mr
                
                if max_min:
                    print(f'alb='+str(alb))
                    print(f'T_mean='+str(T_mean))
                    print(f'T_min='+str(T_min))
                    print(f'T_max='+str(T_max))
                    print(f'T_mr_mean='+str(T_mr))
                    
                else:  
                    print(alb, T_mean, T_mr)
                
                # print(f'T_mr_min '+str((df_xs.xs('street'+slice_name, axis=1, level='location')).xs("T_mr", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(min).values[0]-273).strip('[]'))
                # print(f'T_mr_max '+str((df_xs.xs('street'+slice_name, axis=1, level='location')).xs("T_mr", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(max).values[0]-273).strip('[]'))
                # print(f'T_mr mean {time} = '+str((df.xs('street', axis=1, level='location')).xs("T_mr", axis=1, level='variable').xs('results raw', axis=1, level='category').applymap(lambda x: sum(x)/len(x)).values[0]-273))
                print('\n\n')

df_results.to_csv('demos/df_real.csv')

# =============================================================================
 # %% Load data from google sheet
# =============================================================================

if 'single' in folder.lower():
    sheet_id = "1ocbQ9Un-9l7tsemIWxXdqsfMOpX4SYfiGEq84rOh5DI"
    sheet_name = "sheet_actual"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={sheet_name}"
    dateparse = lambda dates: [datetime.strptime(d, '%d/%m/%Y %H:%M:%S') for d in dates]
    df_measurements = pd.read_csv(url, index_col=0, parse_dates=True, dayfirst=True)
    # df_measurements.index = aida_winter.index.map(lambda t: t.replace(year=2022, month=12, day=3))
    _dates = ['2023-09-06', '2023-09-07', '2023-09-08', '2023-09-09']
    for i, _date in enumerate(_dates[:-1]):
        df_measurements.plot(linestyle='', marker='.', xlim=(_dates[i], _dates[i+1]), ylim=(15, 65))
    
if 'japan' in folder.lower():
    sheet_id = "1IlwzN0Q2G6i5l_R6vjBmjSLTCAqlPN-uPVuUJ-gemAQ"
    sheet_name = "winter"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={sheet_name}"
    aida_winter = pd.read_csv(url, index_col=0, parse_dates=True)
    aida_winter.index = aida_winter.index.map(lambda t: t.replace(year=2022, month=12, day=3))
    
    sheet_id = "1IlwzN0Q2G6i5l_R6vjBmjSLTCAqlPN-uPVuUJ-gemAQ"
    sheet_name = "zomer"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={sheet_name}"
    aida_summer = pd.read_csv(url, index_col=0, parse_dates=True)
    aida_summer.index = aida_summer.index.map(lambda t: t.replace(year=2023, month=6, day=15))
    
    
    data_sheet = load_workbook('data/Compare simulations with Standing Method.xlsx')['Raw data']
    
    # for model in ('model 1', 'model 2', 'model 3'):
        
    Visser_model1 = get_table_from_sheet(data_sheet=data_sheet, name='Visser_model1', date=date)
    Visser_model2 = get_table_from_sheet(data_sheet=data_sheet, name='Visser_model2', date=date)
    Visser_model3 = get_table_from_sheet(data_sheet=data_sheet, name='Visser_model3', date=date)
    Fortuniak_model1 = get_table_from_sheet(data_sheet=data_sheet, name='Fortuniak_model1', date=date)
    Fortuniak_model2 = get_table_from_sheet(data_sheet=data_sheet, name='Fortuniak_model2', date=date)
    Schrijvers_model3 = get_table_from_sheet(data_sheet=data_sheet, name='Schrijvers_model3', date=date)
    
    for hw in df.columns.unique('hw'):
        df_alb0400 = (df.xs("q_sw_sensors['in']", axis=1, level='variable').xs(f'floating bottom{slice_name}', axis=1, level='location').xs(0.400, axis=1, level='albedo').xs('results, per m2 surface', axis=1, level='category') / df.xs("q_sw_sensors['in']", axis=1, level='variable').xs(f'floating top{slice_name}', axis=1, level='location').xs(0.400, axis=1, level='albedo').xs('results, per m2 surface', axis=1, level='category')).xs(hw, 1, 'hw').loc[str(date)].dropna(axis=1, how='all')
        df_alb0400 = df_alb0400.replace([np.inf, -np.inf], np.nan).iloc[:, 0].droplevel(('azimuth', 'zenith'))
        
        df_alb0405 = (df.xs("q_sw_sensors['in']", axis=1, level='variable').xs(f'floating bottom{slice_name}', axis=1, level='location').xs(0.405, axis=1, level='albedo').xs('results, per m2 surface', axis=1, level='category') / df.xs("q_sw_sensors['in']", axis=1, level='variable').xs(f'floating top{slice_name}', axis=1, level='location').xs(0.405, axis=1, level='albedo').xs('results, per m2 surface', axis=1, level='category')).xs(hw, 1, 'hw').loc[str(date)].dropna(axis=1, how='all')
        df_alb0405 = df_alb0405.replace([np.inf, -np.inf], np.nan).iloc[:, 0].droplevel(('azimuth', 'zenith'))
        
        fig,ax= plt.subplots()
        
    
        
        if 'model 0' in folder.lower():
            model = 'model 0'
        if 'model 1' in folder.lower():
            model = 'model 1'
            
            pd.Series(data=Visser_model1['Visser sim.'].values, index=Visser_model1.index.values).plot(ax=ax, ls='--', label='Visser')
            pd.Series(data=Fortuniak_model1['Fortuniak sim.'].values, index=Fortuniak_model1.index.values).plot(ax=ax, ls='--', label='Fortuniak')
            
            df_to_compare = pd.DataFrame(index=aida_summer['model 1'].dropna(axis=0).index)
            df_to_compare['Visser'] = reindex_dataframe(df=Visser_model1, new_timestamps=aida_summer['model 1'].dropna(axis=0).index)
            df_to_compare['Fortuniak'] = reindex_dataframe(df=Fortuniak_model1, new_timestamps=aida_summer['model 1'].dropna(axis=0).index)
            
        if 'model 2' in folder.lower():
            model = 'model 2'
            
            pd.Series(data=Visser_model2['Visser sim.'].values, index=Visser_model2.index.values).plot(ax=ax, ls='--', label='Visser')
            pd.Series(data=Fortuniak_model2['Fortuniak sim.'].values, index=Fortuniak_model2.index.values).plot(ax=ax, ls='--', label='Fortuniak')
            
            df_to_compare = pd.DataFrame(index=aida_summer['model 2'].dropna(axis=0).index)
            df_to_compare['Visser'] = reindex_dataframe(df=Visser_model2, new_timestamps=aida_summer['model 2'].dropna(axis=0).index)
            df_to_compare['Fortuniak'] = reindex_dataframe(df=Fortuniak_model2, new_timestamps=aida_summer['model 2'].dropna(axis=0).index)
            
        if 'model 3' in folder.lower():
            model = 'model 3'
            
            pd.Series(data=Visser_model3['Visser sim.'].values, index=Visser_model3.index.values).plot(ax=ax, ls='--', label='Visser')
            pd.Series(data=Schrijvers_model3['Schrijvers sim.'].values, index=Schrijvers_model3.index.values).plot(ax=ax, ls='--', label='Schrijvers', c=colors[5])
            
            df_to_compare = pd.DataFrame(index=aida_summer['model 3'].dropna(axis=0).index)
            df_to_compare['Visser'] = reindex_dataframe(df=Visser_model3, new_timestamps=aida_summer['model 3'].dropna(axis=0).index)
            df_to_compare['Schrijvers'] = reindex_dataframe(df=Schrijvers_model3, new_timestamps=aida_summer['model 3'].dropna(axis=0).index)
        
        
        
        
        df_alb0400.plot(ax=ax, label=r'Current sim. ($\alpha_\mathrm{roof}=0.4$)', color=colors[2])
        df_alb0405.plot(ax=ax, label=r'Current sim. ($\alpha_\mathrm{roof}=0.405$)', color=colors[3])
        
        
        aida_summer[model].dropna(axis=0).plot(ax=ax, label='Aida exp.', marker='.', ls='', color=colors[4])
        
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%H:%M')) 
        # Change the tick interval
        plt.gca().xaxis.set_major_locator(mdates.HourLocator(interval=1)) 
        
        plt.xlim(f'{date} 07:00', f'{date} 17:00')
        plt.ylim(0.2, 0.5)
        
        plt.xlabel('Local time [H]')
        plt.ylabel(r'$\alpha_\mathrm{urban}$ [-]')
    
        
        plt.legend()
        title = 'effective albedo' + 'hw=' + str(hw) 
        
        plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(title)}.png', bbox_inches="tight")
        plt.title(title)
        plt.show()    
        
        reindex = index=aida_summer[model].dropna(axis=0).index
    
        df_to_compare['Current sim. (alb 0.4)'] = reindex_dataframe(df=df_alb0400, new_timestamps=reindex)
        df_to_compare['Current sim. (alb 0.405)'] = reindex_dataframe(df=df_alb0405, new_timestamps=reindex)
        
        
        df_difference = df_to_compare.sub(aida_summer[model].dropna(axis=0), axis=0)
        df_MAE = df_difference.abs().mean(0)
        df_RMSE = ((df_difference**2).mean(0))**(1/2)
    

# %% Plot measurements

# %%% Plot brick wall temps and bitumen
if 'single' in folder.lower():
    day_to_plot = 6
    
    df_measurements_hour6 = df_measurements.loc[:, '2023-09-06'].copy()
    df_measurements_hour7 = df_measurements.loc[:, '2023-09-07'].copy()
    df_measurements_hour8 = df_measurements.loc[:, '2023-09-08'].copy()
    
    df_measurements_hour6.index=df_measurements_hour6.index.map(lambda t: t.replace(year=2023, month=9, day=day_to_plot))
    df_measurements_hour7.index=df_measurements_hour7.index.map(lambda t: t.replace(year=2023, month=9, day=day_to_plot))
    df_measurements_hour8.index=df_measurements_hour8.index.map(lambda t: t.replace(year=2023, month=9, day=day_to_plot))
    
    colors=plt.rcParams['axes.prop_cycle'].by_key()['color']
    
    xlim = ('2023/09/0'+str(day_to_plot), '2023/09/0'+str(day_to_plot+1))
    # ylim = (12, 75)
    
    df_xs = df.loc['2023-09-0'+str(day_to_plot), :].xs('T', axis=1, level='variable').xs('results raw', axis=1, level='category')
    # df_xs = df_xs.xs(1.25, axis=1, level='wind speed')
    # df_xs = df_xs.xs('custom', axis=1, level='albedo')
    # df_xs = df_xs.xs('custom', axis=1, level='albedo')
    # keep_cols = ['skin thickness']
    keep_cols = ['skin thickness']
    levels_to_drop = [l for l in df_xs.columns.names if (df_xs.columns.get_level_values(l).nunique() == 1 and not l in keep_cols)]
    df_xs = df_xs.droplevel(levels_to_drop, axis=1).sort_index(axis=1)
    
    # df_xs = df_xs.xs(0.02, axis=1, level='skin thickness')
    # df_xs = df_xs.xs('custom', axis=1, level='albedo')
    # df_xs = df_xs.xs(0.10, axis=1, level='albedo')
    
    import matplotlib.dates as md
    # for orient in ['N', 'Z', 'E', 'W', 'bitumen zon', 'bitumen shaduw']:
        
    legend = False
    for orient in ['N', 'Z', 'E', 'W', 'bitumen zon', 'bitumen shaduw', 'bitumen shaduw']:
        fig, ax = plt.subplots()
        df_measurements_hour6.loc[:, orient].plot(ax=ax, xlim=xlim, linestyle='', marker='.', label='6 Sept.', color=colors[0], zorder=2)
        df_measurements_hour7.loc[:, orient].plot(ax=ax, xlim=xlim, linestyle='', marker='.', label='7 Sept.', color=colors[1], zorder=2)
        df_measurements_hour8.loc[:, orient].plot(ax=ax, xlim=xlim, linestyle='', marker='.', label='8 Sept.', color=colors[2], zorder=2)
        
        if orient == 'E':
            (df_xs.droplevel(('azimuth', 'zenith')).xs('east-facing wall', axis=1, level='location').applymap(lambda x: select_nth_element(x, 277))-273).plot(ax=ax, color=colors[3:], linewidth=1.5, zorder=1)
        if orient == 'W':
            (df_xs.droplevel(('azimuth', 'zenith')).xs('west-facing wall', axis=1, level='location').applymap(lambda x: select_nth_element(x, 267))-273).plot(ax=ax, color=colors[3:], linewidth=1.5, zorder=1)
        if orient == 'N':
            (df_xs.droplevel(('azimuth', 'zenith')).xs('north-facing wall', axis=1, level='location').applymap(lambda x: select_nth_element(x, 507))-273).plot(ax=ax, color=colors[3:], linewidth=1.5, zorder=1)
        if orient == 'Z':
            (df_xs.droplevel(('azimuth', 'zenith')).xs('south-facing wall', axis=1, level='location').applymap(lambda x: select_nth_element(x, 47))-273).plot(ax=ax, color=colors[3:], linewidth=1.5, zorder=1)
        if orient == 'bitumen zon':
            (df_xs.droplevel(('azimuth', 'zenith')).xs('upwards-facing roof', axis=1, level='location').applymap(max)-273).plot(ax=ax, color=colors[3:], linewidth=1.5, zorder=1)
        if orient == 'bitumen shaduw':
            (df_xs.droplevel(('azimuth', 'zenith')).xs('upwards-facing roof', axis=1, level='location').applymap(min)-273).plot(ax=ax, color=colors[3:], linewidth=1.5, zorder=1)

        lines, labels = ax.get_legend_handles_labels()
        
        for i, label in enumerate(labels):
            if i < 3:
                labels[i] = rf'Exp: {label}'
            else:
                # skin = 'skin'
                # labels[i] = r'Sim: $\Delta_{skin}$='+label
                labels[i] = f'{sim_name}: ' + str(''.join([rf'{variable_dict.get(var, var)} = {label} m' for var, name in zip(keep_cols, label)]))
                
        # labels[-1] = 'Simulation'
        # if orient.beginswith('bitumen'):
        #     labels[-2] = 'Simulation'
        # print(labels)
        # Update the legend
        # order = [3,4,5,6,0,1,2]
        # lines, labels = [lines[idx] for idx in order], [labels[idx] for idx in order]
        
        
        
        if legend:
            ax.legend(lines, labels, mode='expand')
            ax.axis('off')
            lines = plt.gca().get_lines()
            for line in lines:
                line.set_visible(False)
            orient = 'legend'
            plt.tight_layout()
        else:
            ax.get_legend().remove()
            
            if orient == 'bitumen shaduw':
                legend = True
        
        plt.ylim(9, 75)
        # plt.xlim(ylim)
        #ax.xaxis.set_major_locator(md.MinuteLocator(byminute = [0]))
        ax.xaxis.set_major_formatter(md.DateFormatter('%H:%M'))
        plt.xlabel('Local time [H]')
        plt.ylabel(y_label_deg_C)
        
        
        # figlegend.show()
        # figlegend.savefig('legend.png')    
        
        
        
        plot_name = f'measurement_and_simulation_{orient}'
        
        plt.tight_layout()
        plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(plot_name)}.png', bbox_inches="tight")
        plt.title(orient)
        plt.show()
    

# %%% Plot bitumen temps

if 'single' in folder.lower():
    markers = ['*', 'x']
    import matplotlib.dates as md
    ax = df_measurements_hour6.loc[:, 'bitumen zon'].plot(xlim=('2023/09/07', '2023/09/08'), linestyle='', label='Sunny: 6 sept.', marker='*', color=colors[0])
    df_measurements_hour7.loc[:, 'bitumen zon'].plot(ax=ax, xlim=('2023/09/07', '2023/09/08'), linestyle='', label='Sunny: 7 sept.', marker='*', color=colors[1])
    df_measurements_hour8.loc[:, 'bitumen zon'].plot(ax=ax, xlim=('2023/09/07', '2023/09/08'), linestyle='', label='Sunny: 8 sept.', marker='*', color=colors[2])
    plt.gca().set_prop_cycle(None)
    df_measurements_hour6.loc[:, 'bitumen shaduw'].plot(ax=ax, xlim=('2023/09/07', '2023/09/08'), linestyle='', label='Shaded: 6 sept.', marker='x', color=colors[0])
    df_measurements_hour7.loc[:, 'bitumen shaduw'].plot(ax=ax, xlim=('2023/09/07', '2023/09/08'), linestyle='', label='Shaded: 7 sept.', marker='x', color=colors[1])
    df_measurements_hour8.loc[:, 'bitumen shaduw'].plot(ax=ax, xlim=('2023/09/07', '2023/09/08'), linestyle='', label='Shaded: 8 sept.', marker='x', color=colors[2])
    plt.title('Bitumen temperature')
    # plt.ylim(18, 65)
    #ax.xaxis.set_major_locator(md.MinuteLocator(byminute = [0]))
    ax.xaxis.set_major_formatter(md.DateFormatter('%H:%M'))
    plt.legend()
    
    plt.show()


# %%% RMSE between measurement and sim
if 'chimney' in folder.lower():

    from sklearn.metrics import mean_squared_error
    df_xs = df.loc['2023-09-07', :].xs('T', axis=1, level='variable').xs('results raw', axis=1, level='category')
    
    df_xs = df_xs.sort_index()
    df_xs = df_xs.loc[:, ~df_xs.applymap(lambda x: all(pd.isna(x))).all(axis=0)]
    # df_xs = df_xs.xs(0.15, axis=1, level='albedo', drop_level=False)
    # df_xs = df_xs.xs('custom', axis=1, level='albedo', drop_level=False)
    # df_xs = df_xs.xs(2.0, axis=1, level='wind speed', drop_level=False)
    # df_xs = df_xs.xs(0.02, axis=1, level='skin thickness', drop_level=False)
        
    x_var = 'wind speed'
    x_var = 'albedo'
    # x_var = 'skin thickness'
    
    rmse_values = pd.DataFrame(
        data=[], 
        # index=df_xs.columns.unique('skin thickness'), 
        # index=df_xs.columns.unique(x_var).astype(float), 
        index=df_xs.columns.unique(x_var),
        columns=pd.MultiIndex.from_product([list(df_xs.columns.unique('albedo')), ['N', 'Z', 'E', 'W', 'bitumen zon', 'bitumen shaduw']], names=['albedo', 'skin thickness'],)
        )
    
    rmse_w_values = pd.DataFrame(
        data=[], 
        # index=df_xs.columns.unique('skin thickness'), 
        index=df_xs.columns.unique(x_var),
        # index=df_xs.columns.unique(x_var).astype(float), 
        columns=pd.MultiIndex.from_product([list(df_xs.columns.unique('albedo')), ['N', 'Z', 'E', 'W', 'bitumen zon', 'bitumen shaduw']], names=['albedo', 'skin thickness'],)
        )
    
    for orient in ['N', 'Z', 'E', 'W', 'bitumen zon', 'bitumen shaduw']:
        if orient == 'E':
            simulation_df = (df_xs.droplevel(('azimuth', 'zenith')).xs('east-facing wall', axis=1, level='location').applymap(lambda x: select_nth_element(x, 277))-273)
        if orient == 'W':
            simulation_df = (df_xs.droplevel(('azimuth', 'zenith')).xs('west-facing wall', axis=1, level='location').applymap(lambda x: select_nth_element(x, 267))-273)
        if orient == 'N':
            simulation_df = (df_xs.droplevel(('azimuth', 'zenith')).xs('north-facing wall', axis=1, level='location').applymap(lambda x: select_nth_element(x, 507))-273)
        if orient == 'Z':
            simulation_df = (df_xs.droplevel(('azimuth', 'zenith')).xs('south-facing wall', axis=1, level='location').applymap(lambda x: select_nth_element(x, 47))-273)
        if orient == 'bitumen zon':
            simulation_df = (df_xs.droplevel(('azimuth', 'zenith')).xs('upwards-facing roof', axis=1, level='location').applymap(max)-273)
        if orient == 'bitumen shaduw':
            simulation_df = (df_xs.droplevel(('azimuth', 'zenith')).xs('upwards-facing roof', axis=1, level='location').applymap(min)-273)
        
        
        for col in simulation_df.columns:
            rmse_values_day = []
            rmse_w_values_day = []
            # for day, df_measurements_hour_day_filtered in enumerate([df_measurements_hour6.resample('15T').mean().dropna(), df_measurements_hour7.resample('15T').mean().dropna(), df_measurements_hour8.resample('15T').mean().dropna()]):
            for day, df_measurements_hour_day_filtered in enumerate([df_measurements_hour6.resample('15T').mean().dropna(), df_measurements_hour7.resample('15T').mean().dropna(), df_measurements_hour8.resample('15T').mean().dropna()]):
                
                weights = calculate_weights_inverse_density(df_measurements_hour_day_filtered.index.to_series().view(int) // 10**9)
                
                # Calculate RMSE for the current column
                simulation_df_sliced = simulation_df.loc[df_measurements_hour_day_filtered.index]
                rmse = np.sqrt(mean_squared_error(df_measurements_hour_day_filtered[orient], simulation_df_sliced[col]))
                rmse_values_day.append(rmse)
                
                rmse_w = np.sqrt(mean_squared_error(df_measurements_hour_day_filtered[orient], simulation_df_sliced[col], sample_weight=weights))
                # print('weights', weights)
                rmse_w_values_day.append(rmse_w)
                
                
            print('',col[simulation_df.columns.names.index(x_var)], (col[2], orient))
            rmse_values.loc[col[simulation_df.columns.names.index(x_var)], (col[2], orient)] = np.sqrt(np.mean(np.square(rmse_values_day)))
            rmse_w_values.loc[col[simulation_df.columns.names.index(x_var)], (col[2], orient)] = np.sqrt(np.mean(np.square(rmse_w_values_day)))
            
    rmse_values = rmse_values.sort_index()
    # rmse_values = rmse_values.sort_index(1)
    rmse_w_values = rmse_w_values.sort_index()
    # rmse_w_values = rmse_w_values.sort_index(1)
            
    # rmse_values.loc[slice(None), (slice(None), ['N', 'E', 'Z', 'W'])].plot(ylim=(0, 20), linestyle='', marker='.', title='unweighed')
    rmse_w_values.loc[slice(None), (slice(None), ['N', 'E', 'Z', 'W'])].plot(ylim=(0, 20), linestyle='', marker='.', title='weighed')
    rmse_w_total_walls = np.sqrt(np.mean(np.square(rmse_w_values.loc[slice(None), (slice(None), ['N', 'E', 'Z', 'W'])]), axis=1))
    plt.legend()
    # plt.xlim(0, 5)
    plt.show()
    # rmse_values.loc[slice(None), (slice(None), ['bitumen zon', 'bitumen shaduw'])].plot(ylim=(0, 20), linestyle='', marker='.', title='unweighed')
    rmse_w_values.loc[slice(None), (slice(None), ['bitumen zon', 'bitumen shaduw'])].plot(ylim=(0, 20), linestyle='', marker='.', title='weighed')
    rmse_w_total_roofs = np.sqrt(np.mean(np.square(rmse_w_values.loc[slice(None), (slice(None), ['bitumen zon', 'bitumen shaduw'])]), axis=1))
    plt.legend()
    # plt.xlim(0, 5)

    
    plt.show()
    
# %% PLOT 1 POINT IN CANYON

df_xs = (df.xs('T_mr',1,'variable').xs(0.2,1,'albedo').xs('results raw',1,'category').applymap(lambda x: x[0]).droplevel(('azimuth', 'zenith'))-273.15).sort_index(1).loc[str(date)]
df_xs = df_xs.loc[:, (slice(None), [0.0, 0.5, 1.0, 1.5, 2.0, 4.0])]
df_xs.plot(ylim=(13,120),ylabel='Mean Radiant Temperature', xlabel='Local time [H]', title='Mean radiant temperature at North-side in EW canyon with Albedo=0.2')
plt.legend(df_xs.columns.get_level_values('hw'))
plt.show()

# %%

df_xs = (df.xs('T',1,'variable').xs(f'canyon street{slice_name}',1,'location').xs(0.4,1,'albedo').xs('results raw',1,'category').applymap(lambda x: x[0]).droplevel(('azimuth', 'zenith'))-273.15).sort_index(1).loc[str(date)]
df_xs = df_xs.loc[:, (slice(None), [0.0, 0.5, 1.0, 1.5, 2.0, 4.0])]
df_xs.plot(ylim=(13,120),ylabel='Street surface temperature', xlabel='Local time [H]', title='Street surface temprature at North-side in EW canyon with Albedo=0.1')
plt.legend(df_xs.columns.get_level_values('hw'))
    



# %% TOTAL CANYON SCALE 

# %%% L*, K* en Q* vs t for h/w=1

# dates = (datetime.date(2022, 12, 3), datetime.date(2023, 6, 17))
dates = (datetime.date(2019, 7, 23), )

if 'real' in folder.lower():
    albedos = slice(None)
    ress = slice(None)
elif 'custom' in folder.lower():
    albedos = [0.2]
    albedos = slice(None)
    ress = [1.0, 2.0]
    

for date in dates:
    plots_setup = {
        f'SW & LW Radiation through floating sensors above environment \n {date}': 
            {'xs': {
                'category': 'results, per m2 surface',
                'emissivity sky': slice(None),
                'hw': [1.0],
                'albedo': albedos,
                'emissivity': slice(None),
                'skin thickness': slice(None),
                # 'skin thickness': [0.05],
                ##'density': slice(None),
                'vol. heat cap.': slice(None),
                'wind speed': slice(None),
                'location': [f'floating top{slice_name} - floating bottom{slice_name}'],
                # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                # 'variable': ["q_sw_sensors['in']"],
                'variable': ["q_sw_sensors['in']", "q_lw_sensors['in']", "q_sw_sensors['in'] + q_lw_sensors['in']"],
                # 'res': ress,
                'res': [2.0],
                },
                
            'linestyle': 'variable',
            # 'print daily sum': True,
            # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
            'annotations': True,
            'x': 't',
            'ylabel': y_label1,
            # 'xlim': xlim_hw,
            'ylim': ylim_W,
            # 'y_scale': 1e-6,
            'date': date,
            'droplevel': ('zenith', 'azimuth'),
            'legend': ['variable'],
            },
        }
        
    plot_lines(df, plots_setup, line_properties) 
    
    
# %%% Oke Lup, Ldown, Kup, Ldown en Q* vs t for h/w=1

# dates = (datetime.date(2022, 12, 3), datetime.date(2023, 6, 17))

# ress = [1.0, 2.0]

# albedos = ['custom']

plots_setup = {}

for key in keys:
    for date in [date]:
        plots_setup.update({
            # f'Oke: SW & LW Radiation through floating sensors above environment H/W=1 \n {date}, {level}={key}': 
                # {'xs': {
                    
                #     'category': 'results, per m2 surface',
                #     'emissivity sky': slice(None),
                #     'hw': [1.0],
                #     'albedo': albedos,
                #     'emissivity': slice(None),
                #     'skin thickness': slice(None),
                #     ##'density': slice(None),
                #     'vol. heat cap.': slice(None),
                #     'wind speed': slice(None),
                #     'location': [f'floating bottom{slice_name}', f'floating top{slice_name}'],
                #     # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                #     # 'variable': ["q_sw_sensors['in']"],
                #     # 'variable': ["q_sw_sensors['in']", "q_lw_sensors['in']", "q_sw_sensors['in'] + q_lw_sensors['in']"],
                #     'variable': ["q_lw_sensors['in']", "q_sw_sensors['in']"],
                #     'res': ress,
                #     keys.name: [key],
                #     },
                    
                # 'linestyle': ['dashed', 'dashed', 'solid', 'solid'],
                # 'color': [colors[0], colors[1], colors[0], colors[1]],
                # 'order': [3, 1, 2, 0],
                # # 'print daily sum': True,
                # 'fill_between': [[(f'floating bottom{slice_name}', "q_lw_sensors['in']"), (f'floating top{slice_name}', "q_lw_sensors['in']")],
                #                  [(f'floating bottom{slice_name}', "q_sw_sensors['in']"), (f'floating top{slice_name}', "q_sw_sensors['in']")]],
                # 'fill_color': ['#63a3c9', '#b8405a'],
                # 'fill_zorder': [2, 1],
                # 'annotations': True,
                # 'x': 't',
                # 'ylabel': y_label1,
                # # 'xlim': xlim_hw,
                # # 'ylim': ylim_W,
                # # 'y_scale': 1e-6,
                # 'date': date,
                # 'droplevel': ('zenith', 'azimuth'),
                # 'labels': [r'$L^{\uparrow}$', r'$K^{\uparrow}$', r'$L^{\downarrow}$', r'$K^{\downarrow}$'],
                # 'legend': ['variable', 'location'],
                # 'legend_title': 'variable',
                # },
                
            f'Oke: SW & LW Radiation through floating sensors above environment {level}={key} \n {date}':
                {'xs': {
                    
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': [0.0],
                    'albedo': [0.2],
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    ##'density': slice(None),
                    'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'floating bottom{slice_name}', f'floating top{slice_name}'],
                    # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                    # 'variable': ["q_sw_sensors['in']"],
                    # 'variable': ["q_sw_sensors['in']", "q_lw_sensors['in']", "q_sw_sensors['in'] + q_lw_sensors['in']"],
                    'variable': ["q_sw_sensors['in']", "q_lw_sensors['in']"],
                    'res': [2.0],
                    keys.name: [key],
                    },
                    
                'linestyle': ['dashed', 'dashed', 'solid', 'solid'],
                'color': [colors[0], colors[1], colors[0], colors[1]],
                'linewidth': [1.5, 1.5, 1.5, 1.5, 0, 0],
                'order': [3, 1, 5, 2, 0, 4],
                # 'print daily sum': True,
                'fill_between': [[(f'floating bottom{slice_name}', "q_lw_sensors['in']"), (f'floating top{slice_name}', "q_lw_sensors['in']")],
                                 [(f'floating bottom{slice_name}', "q_sw_sensors['in']"), (f'floating top{slice_name}', "q_sw_sensors['in']")]],
                'fill_color': ['#63a3c9', '#b8405a'],
                'fill_zorder': [2, 1],
                'annotations': True,
                'x': 't',
                'ylabel': y_label1,
                # 'xlim': xlim_hw,
                # 'ylim': ylim_W,
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'labels': [r'$L^{\uparrow}$', r'$K^{\uparrow}$', r'$L^{\downarrow}$', r'$K^{\downarrow}$', r'$-L^\ast$', r'$K^\ast$'],
                'legend': ['variable', 'location'],
                # 'legend': False,
                'legend_title': 'variable',
                },
            })
        
    

# plot_lines(df, plots_setup, line_properties) 
plot_lines(df, plots_setup, line_properties, legend_kwargs={'bbox_to_anchor':(1.04, 0.5), 'loc':"center left", 'borderaxespad':0}, add_name='lgnd side') 
    
    
# %%% K_up vs t

plots_setup = {}

keys = df.columns.unique('albedo')

for key in keys:
    plots_setup.update({
        f'SW out radiation through floating sensors above environment \n {date}, \n {keys.name}={key}': 
            {'xs': {
                'category': 'results, per m2 surface',
                'emissivity sky': slice(None),
                # 'hw': [1.0, 0.0, 0.5, 2.0, 3.0, 4.0],
                'hw': hws,
                'albedo': [0.2],
                'emissivity': slice(None),
                'skin thickness': slice(None),
                # 'skin thickness': [0.05],
                ##'density': slice(None),
                'vol. heat cap.': slice(None),
                'wind speed': slice(None),
                'location': [f'floating bottom{slice_name}'],
                # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                # 'variable': ["q_sw_sensors['in']"],
                'variable': ["q_sw_sensors['in']"],
                'res': [2.0],
                keys.name: [key],
                # 'res': [2.0],
                },
                
            # 'linestyle': 'variable',
            # 'print daily sum': True,
            # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
            'color': 'hw',
            'annotations': True,
            'x': 't',
            'ylabel': r'K$_{\uparrow,\mathrm{urban}}$ [W m$^{-2}$]',
            # 'xlim': xlim_hw,
            # 'ylim': ylim_W,
            # 'y_scale': 1e-6,
            'date': date,
            'droplevel': ('zenith', 'azimuth'),
            'legend': ['hw'],
            },
        })
    
plot_lines(df, plots_setup, line_properties) 


# %%% Eff alb / trapping enh.

# %%%% Effective alb vs t for hws

ress = [1.0, 2.0]

plots_setup = {}


keys = df.columns.unique('hw')

for key in keys:
    plots_setup.update({
        f'Eff albedo between time, all albedo, \n {keys.name}={key}': 
            {'xs': {
                'category': 'results, per m2 surface',
                'emissivity sky': slice(None),
                'hw': [1.0],
                'albedo': slice(None),
                'emissivity': slice(None),
                'skin thickness': slice(None),
                #'density': slice(None),
                'vol. heat cap.': slice(None),
                'wind speed': slice(None),
                'location': [f'floating bottom{slice_name} / floating top{slice_name}'],
                # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                # 'variable': ["q_sw_sensors['in']"],
                'variable': ["q_sw_sensors['in']"],
                'res': ress,
                keys.name: [key],
                },
                
            'color': 'hw',
            'linestyle': 'albedo',
            'annotations': True,
            'between_time': ('07:00', '21:00'),
    
            'x': 't',
            'ylabel': 'Canyon Effective albedo [-]',
            'ylim': (-0.01, None),
            'date': date,
            'droplevel': ('zenith', 'azimuth'),
            'legend': ['albedo'],
            },
            
        # f'Eff albedo between time, all hw, alb=0.4, \n {keys.name}={key}': 
        #     {'xs': {
        #         'category': 'results, per m2 surface',
        #         'emissivity sky': slice(None),
        #         'hw': hws,
        #         'albedo': [0.2],
        #         'emissivity': slice(None),
        #         'skin thickness': slice(None),
        #         #'density': slice(None),
        #         'vol. heat cap.': slice(None),
        #         'wind speed': slice(None),
        #         'location': [f'floating bottom{slice_name} / floating top{slice_name}'],
        #         # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
        #         # 'variable': ["q_sw_sensors['in']"],
        #         'variable': ["q_sw_sensors['in']"],
        #         'res': ress,
        #         keys.name: [key],
        #         },
                
        #     'color': 'hw',
        #     'annotations': True,
        #     'between_time': ('07:00', '21:00'),
    
        #     'x': 't',
        #     'ylabel': 'Canyon Effective albedo [-]',
        #     'ylim': (-0.01, None),
        #     'date': date,
        #     'droplevel': ('zenith', 'azimuth'),
        #     'legend': ['hw'],
        #     },
            
        f'Eff albedo, \n {keys.name}={key}': 
            {'xs': {
                'category': 'results, per m2 surface',
                'emissivity sky': slice(None),
                'hw': hws,
                'albedo': slice(None),
                'emissivity': slice(None),
                'skin thickness': slice(None),
                #'density': slice(None),
                'vol. heat cap.': slice(None),
                'wind speed': slice(None),
                'location': [f'floating bottom{slice_name} / floating top{slice_name}'],
                # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                # 'variable': ["q_sw_sensors['in']"],
                'variable': ["q_sw_sensors['in']"],
                # 'res': ress,
                'res': [2.0],
                keys.name: [key],
                },
                
            'color': 'hw',
            'annotations': True,
            # 'between_time': ('07:00', '17:00'),
    
            'x': 't',
            'ylabel': 'Canyon Effective albedo [-]',
            'ylim': (-0.01, 1),
            'date': date,
            'droplevel': ('zenith', 'azimuth'),
            'legend': ['hw'],
            },
        })
        
keys = df.columns.unique('albedo')
        
for key in keys:
    plots_setup.update({
        f'Eff albedo between time, all hw, alb=0.4, \n {keys.name}={key}': 
            {'xs': {
                'category': 'results, per m2 surface',
                'emissivity sky': slice(None),
                'hw': hws,
                'albedo': [0.4],
                'emissivity': slice(None),
                'skin thickness': slice(None),
                #'density': slice(None),
                'vol. heat cap.': slice(None),
                'wind speed': slice(None),
                'location': [f'floating bottom{slice_name} / floating top{slice_name}'],
                # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                # 'variable': ["q_sw_sensors['in']"],
                'variable': ["q_sw_sensors['in']"],
                # 'res': ress,
                'res': [2.0],
                keys.name: [key],
                },
                
            'color': 'hw',
            'annotations': True,
            'between_time': ('07:00', '21:00'),
    
            'x': 't',
            'ylabel': 'Canyon Effective albedo [-]',
            'ylim': (-0.01, None),
            'date': date,
            'droplevel': ('zenith', 'azimuth'),
            'legend': ['hw'],
            },
        })
        

plot_lines(df, plots_setup, line_properties)

# %%%% Effective alb vs hws for different albedos

plots_setup = {
    f'Mean eff urban albedo': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': [0.0, 0.5, 1.5, 2.0, 4.0, 8.0],
            'albedo': [0.1, .2, .4, .6],
            'emissivity': slice(None),
            'skin thickness': slice(None),            
            # 'skin thickness': [0.016, 0.017, 0.018, 0.019],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'floating bottom{slice_name} / floating top{slice_name}'],
            # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            # 'variable': ["q_sw_sensors['in']"],
            'variable': ["q_sw_sensors['in']"],
            'res': [2.0],
            },
            
        # 'color': 'hw',
        'color': 'albedo',
        'annotations': True,
        'between_time': ('07:00', '21:00'),
        'droplevel_col': ('skin thickness', 'vol. heat cap.'),

        'x': 'hw',
        'ylabel': 'Daily Canyon Effective albedo [-]',
        'ylim': (-0.01, None),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
        
    
        
    }

plot_lines(df, plots_setup, line_properties)

# %%%% Trapping enhancement factor

plots_setup = {
    f'Shortwave trapping enhancement for h/w=1, different albedos': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': slice(None),
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'floating top{slice_name} / floating bottom{slice_name}'],
            # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            # 'variable': ["q_sw_sensors['in']"],
            'variable': ["q_sw_sensors['in']"],
            'res': [2.0],
            },
            
        'color': 'albedo',
        'multiply': 'albedo',
        'annotations': True,
        'between_time': ('07:00', '21:00'),
    
        'x': 't',
        'ylabel': 'SW trapping enh. factor [-]',
        # 'ylim': (-0.01, 0.5),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    f'Shortwave trapping enhancement for albedo=0.2, different hws': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [.0, .5, 1, 2, 3, 4],
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'floating top{slice_name} / floating bottom{slice_name}'],
            # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            # 'variable': ["q_sw_sensors['in']"],
            'variable': ["q_sw_sensors['in']"],
            'res': [2.0],
            },
            
        'color': 'hw',
        'multiply': 'albedo',
        'annotations': True,
        'between_time': ('07:00', '21:00'),
    
        # 'linewidth':'hw',
        'x': 't',
        'ylabel': 'SW trapping enh. factor [-]',
        'ylim': (0.9, 2.1),
        # 'ylim': (-0.01, 0.5),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'] if 'Model 2' in folder.lower() else False,

        },
        
    f'Shortwave trapping enhancement for albedo=0.6, different hws': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': hws,
            'albedo': [0.6],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'floating top{slice_name} / floating bottom{slice_name}'],
            # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            # 'variable': ["q_sw_sensors['in']"],
            'variable': ["q_sw_sensors['in']"],
            'res': [2.0],
            },
            
        'color': 'hw',
        'multiply': 'albedo',
        'annotations': True,
        'between_time': ('07:00', '21:00'),
    
        # 'linewidth':'hw',
        'x': 't',
        'ylabel': 'SW trapping enh. factor [-]',
        'ylim': (0.9, 2.1),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    f'Shortwave trapping enhancement vs hw, different albedos': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': [0.1, 0.2, 0.4, 0.6],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'floating top{slice_name} / floating bottom{slice_name}'],
            # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            # 'variable': ["q_sw_sensors['in']"],
            'variable': ["q_sw_sensors['in']"],
            'res': [2.0],
            },
            
        # 'linestyle': ['--','--','--','--','--'],
        'color': 'albedo',
        'multiply': 'albedo',
        'annotations': True,
        'between_time': ('07:00', '21:00'),
        'droplevel_col': ('skin thickness', 'vol. heat cap.'),
        'interpolate': True,
        'x': 'hw',
        'xlim': (-0.1, 4.1),
        ''
        'ylabel': 'Mean SW trapping enh. factor [-]',
        'ylim': (None, 1.8),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
    }

plot_lines(df, plots_setup, line_properties, legend_kwargs={'bbox_to_anchor':(1.04, 0.5), 'loc':"center left", 'borderaxespad':0})
    
# %%% Eff albedo vs time, compare with aida
        
ress = [1.0, 2.0]
dates = [date]

if 'japan' in folder.lower():

    for date in dates:
        plots_setup = {
            f'Compare Effective albedo \n {date}': 
                {'xs': {
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': [1.0],
                    'albedo': [0.400],
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    'density': slice(None),
                    'spec. heat cap.': slice(None),
                    # 'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'floating bottom{slice_name} / floating top{slice_name}'],
                    # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                    # 'variable': ["q_sw_sensors['in']"],
                    'variable': ["q_sw_sensors['in']"],
                    'res': ress,
                    # 'res': [2.0],
                    },
                    
                # 'linestyle': 'variable',
                # 'print daily sum': True,
                # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
                # 'between_time': ('07:00', '17:00'),
                'annotations': True,
                'x': 't',
                'ylabel': y_label1,
                'ylim': (-0.01, 0.5),
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'legend': ['variable'],
                # 'labels': ['Simulation', 'Aida exp.', 'Visser sim.', 'Schrijvers sim.'],
                },
                
            f'Compare Effective albedo (between 07:00-17:00) \n {date}': 
                {'xs': {
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': [1.0],
                    'albedo': [0.400],
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    #'density': slice(None),
                    'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'floating bottom{slice_name} / floating top{slice_name}'],
                    # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                    # 'variable': ["q_sw_sensors['in']"],
                    'variable': ["q_sw_sensors['in']"],
                    # 'res': ress,
                    'res': [2.0],
                    },
                    
                # 'linestyle': ['-', ''],
                # 'print daily sum': True,
                # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
                'between_time': ('07:00', '17:00'),
                'annotations': True,
                'x': 't',
                'ylabel': r'$\alpha_\mathregular{urban}$ [-]',
                'xlim': (f'{date} 07:00', f'{date} 17:00'),
                'ylim': (-0.01, 0.5),
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'legend': ['variable'],
                'legend_title': None,
                # 'labels': ['Simulation', 'Aida exp.', 'Visser sim.', 'Schrijvers sim.'],
                },
                
            f'Compare Effective albedo H/W=0 (between 07:00-17:00) \n {date}': 
                {'xs': {
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': [0.0],
                    'albedo': slice(None),
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    #'density': slice(None),
                    'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'floating bottom{slice_name} / floating top{slice_name}'],
                    # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                    # 'variable': ["q_sw_sensors['in']"],
                    'variable': ["q_sw_sensors['in']"],
                    # 'res': ress,
                    'res': [2.0],
                    },
                    
                'linestyle': ['-', ''],
                # 'print daily sum': True,
                # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
                'between_time': ('07:00', '17:00'),
                'annotations': True,
                'x': 't',
                'ylabel': r'$\alpha_\mathregular{urban}$ [-]',
                # 'xlim': xlim_hw,
                'ylim': (-0.01, 0.5),
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'legend': ['variable'],
                'legend_title': None,
                'labels': ['Simulation', 'Aida exp.'],
                },
            }
            
        plot_lines(df, plots_setup, line_properties)
        
    
 # %%% L up vs t, 
    
plots_setup = {
    f'LW out radiation through floating sensors above environment, alb=0.2': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            # 'hw': [1.0, 0.0, 0.5, 2.0, 3.0, 4.0],
            'hw': hws,
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            # 'skin thickness': [0.05],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'floating bottom{slice_name}'],
            # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            # 'variable': ["q_sw_sensors['in']"],
            'variable': ["q_lw_sensors['in']"],
            'res': [2.0],
            # 'res': [2.0],
            },
            
        # 'linestyle': 'variable',
        # 'print daily sum': True,
        # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
        'color': 'hw',
        'annotations': True,
        'x': 't',
        'ylabel': r'L$_{\uparrow,\mathrm{urban}}$ [W m$^{-2}$]',
        'hline': False,
        # 'xlim': xlim_hw,
        'ylim': (340, 700),
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    f'LW out radiation through floating sensors above environment, hw=1.0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            # 'hw': [1.0, 0.0, 0.5, 2.0, 3.0, 4.0],
            'hw': [1.0],
            'albedo': slice(None),
            'emissivity': slice(None),
            'skin thickness': slice(None),
            # 'skin thickness': [0.05],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'floating bottom{slice_name}'],
            # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            # 'variable': ["q_sw_sensors['in']"],
            'variable': ["q_lw_sensors['in']"],
            'res': [2.0],
            # 'res': [2.0],
            },
            
        # 'linestyle': 'variable',
        # 'print daily sum': True,
        # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
        'color': 'albedo',
        'annotations': True,
        'x': 't',
        'ylabel': r'L$_{\uparrow,\mathrm{urban}}$ [W m$^{-2}$]',
        'hline': False,
        # 'xlim': xlim_hw,
        'ylim': (340, 700),
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
    }

plot_lines(df, plots_setup, line_properties)


# %%% Q star vs t, 
    
plots_setup = {
    f'Q net radiation through floating sensors above environment, alb=0.2': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            # 'hw': [1.0, 0.0, 0.5, 2.0, 3.0, 4.0],
            'hw': hws,
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            # 'skin thickness': [0.05],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'floating top{slice_name} - floating bottom{slice_name}'],
            # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            # 'variable': ["q_sw_sensors['in']"],
            'variable': ["q_lw_sensors['in'] + q_sw_sensors['in']"],
            'res': [2.0],
            # 'res': [2.0],
            },
            
        # 'linestyle': 'variable',
        # 'print daily sum': True,
        # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
        'color': 'hw',
        'annotations': True,
        'x': 't',
        'ylabel': r'Q$^\ast$ [W m$^{-2}$]',
        # 'hline': False,
        # 'xlim': xlim_hw,
        'ylim': (-90, 670),
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    f'Q net radiation through floating sensors above environment, hw=1.0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            # 'hw': [1.0, 0.0, 0.5, 2.0, 3.0, 4.0],
            'hw': [1.0],
            'albedo': slice(None),
            'emissivity': slice(None),
            'skin thickness': slice(None),
            # 'skin thickness': [0.05],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'floating top{slice_name} - floating bottom{slice_name}'],
            # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            # 'variable': ["q_sw_sensors['in']"],
            'variable': ["q_lw_sensors['in'] + q_sw_sensors['in']"],
            'res': [2.0],
            # 'res': [2.0],
            },
            
        # 'linestyle': 'variable',
        # 'print daily sum': True,
        # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
        'color': 'albedo',
        'annotations': True,
        'x': 't',
        'ylabel': r'Q$^\ast$ [W m$^{-2}$]',
        # 'hline': False,
        # 'xlim': xlim_hw,
        'ylim': (-90, 670),
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
    }

plot_lines(df, plots_setup, line_properties)


# %% URBAN SURFACE

# %%% Flux per t over all sections

if 'model 1' in folder.lower():
    locations = {'W Roof': f'west roof{slice_name}', 
                 'W Wall': f'east-facing wall{slice_name}', 
                 'Street': f'canyon street{slice_name}', 
                 'E Wall': f'west-facing wall{slice_name}', 
                 'E Roof': f'east roof{slice_name}'}
    
    flip = [False, True, False, False, False]
    
if 'model 2' in folder.lower():
    locations = {'S Roof': f'south roof{slice_name}', 
                 'S Wall': f'north-facing wall{slice_name}', 
                 'Street': f'canyon street{slice_name}', 
                 'N Wall': f'south-facing wall{slice_name}', 
                 'N Roof': f'north roof{slice_name}'}
    
    flip = [True, True, True, False, True]
    

# between_time = ('04:00', '04:01')
between_time = ('13:50', '13:51')
# between_time = ('18:40', '18:41')

ress = [2.0]
alb = 0.2

plots_setup = {
    f'Fluxplot, Q net, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_lw['net'] + q_sw['abs']"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$Q^\ast$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-200, 600),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
    
    
    f'T-plot, Surf Temp, alb={alb}, res=2,': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["T"],
            'res': [2.0],
            },
            
        'subtract': 273,
        'color': 'hw',
        'between_time': between_time,
        # 'between_time': ('00:00', '23:59'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': y_label_deg_C,
        'hline': False,
        'xlim': (0, 5),
        'ylim': (12, 75),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, LW out rad, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_lw['out_rad']"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$L^\uparrow_\mathrm{emm}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    
    f'Fluxplot, LW in from sky, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_lw['in_from_sky']"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$L^\downarrow_\mathrm{sky}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, LW in from env, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_lw['in_from_els']"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$L^{\downarrow}_\mathrm{env}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, LW in, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_lw['in_from_els'] + q_lw['in_from_sky']"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$L^{\downarrow}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, SW in dir, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in_dir']"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'K$^{\downarrow}_\mathrm{dir}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, SW in diff, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in_diff']"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'K$^{\downarrow}_\mathrm{diff}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, SW in, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in']"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'K$^{\downarrow}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, SW abs, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in'] - q_sw['out_els_refl']"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        # 'between_time': ('14:00', '14:30'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$K^{\ast}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, SW abs 2, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in']"],
            'res': ress,
            },
            
        'multiply': (1-alb),
        'color': 'hw',
        'between_time': between_time,
        # 'between_time': ('14:00', '14:30'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$K^{\ast}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, SW in from els, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in_from_els']"],
            'res': ress,
            },
            
        'color': 'hw',
        # 'multiply': (1-alb),
        'between_time': between_time,
        # 'between_time': ('14:00', '14:30'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$K^{\downarrow}_\mathrm{env}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
        
    f'Fluxplot, SW abs from els, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in_from_els']"],
            'res': ress,
            },
            
        'color': 'hw',
        'multiply': (1-alb),
        # 'multiply': (1-alb),
        'between_time': between_time,
        # 'between_time': ('14:00', '14:30'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$K^{*}_\mathrm{env}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, SW in from sky, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in_dir'] + q_sw['in_diff']"],
            'res': ress,
            },
            
        'color': 'hw',
        # 'multiply': (1-alb),
        'between_time': between_time,
        # 'between_time': ('14:00', '14:30'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$K^{\downarrow}_\mathrm{sky}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, SW abs from sky, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in_dir'] + q_sw['in_diff']"],
            'res': ress,
            },
            
        'color': 'hw',
        'multiply': (1-alb),
        'between_time': between_time,
        # 'between_time': ('14:00', '14:30'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$K^{*}_\mathrm{sky}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
        
    f'Fluxplot, LW net, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_lw['net']"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        # 'between_time': ('14:00', '14:30'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$L^{\ast}$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-350, 50),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'Fluxplot, Q conv in, alb={alb}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_convection"],
            'res': ress,
            },
            
        'color': 'hw',
        'between_time': between_time,
        # 'between_time': ('14:00', '14:30'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'$Q_H$ [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-300, 70),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    # f'Fluxplot, LW out, alb={alb}':  # niet zo boeiend, is dezelfde patroon als T
    #     {'xs': {
    #         'category': 'results raw',
    #         'emissivity sky': slice(None),
    #         'hw': hws,
    #         # 'hw': slice(None),
    #         'albedo': [alb],
    #         'emissivity': slice(None),
    #         'skin thickness': slice(None),
    #         #'density': slice(None),
    #         'vol. heat cap.': slice(None),
    #         'wind speed': slice(None),
    #         'location': slice(None),
    #         'variable': ["q_lw['out_rad']"],
    #         'res': ress,
    #         },
            
    #     'color': 'hw',
    #     'between_time': between_time,
    #     'x': 't',
    #     'xlabel': 'Scaled distance [-]',
    #     'ylabel': r'$L^{\uparrow}$ [W m$^{-2}$]',
    #     'hline': False,
    #     'xlim': (0, 5),
    #     'ylim': (-10, 1050),
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
        
    #     'show_minutes': [],
    #     },
        
    # f'T-plot, Surf Temp, alb={alb}, res=2,': 
    #     {'xs': {
    #         'category': 'results raw',
    #         'emissivity sky': slice(None),
    #         'hw': hws,
    #         # 'hw': slice(None),
    #         'albedo': [alb],
    #         'emissivity': slice(None),
    #         'skin thickness': slice(None),
    #         #'density': slice(None),
    #         'vol. heat cap.': slice(None),
    #         'wind speed': slice(None),
    #         'location': slice(None),
    #         'variable': ["T"],
    #         'res': [2.0],
    #         },
            
    #     'subtract': 273,
    #     'color': 'hw',
    #     'between_time': between_time,
    #     # 'between_time': ('00:00', '23:59'),
    #     'x': 't',
    #     'xlabel': 'Scaled distance [-]',
    #     'ylabel': y_label_deg_C,
    #     'hline': False,
    #     'xlim': (0, 5),
    #     'ylim': (12, 75),
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
        
    #     'show_minutes': [],
    #     },
        
    # f'T-plot, Surf Temp, alb={0.6}, res=2,': 
    #     {'xs': {
    #         'category': 'results raw',
    #         'emissivity sky': slice(None),
    #         'hw': hws,
    #         # 'hw': slice(None),
    #         'albedo': [0.6],
    #         'emissivity': slice(None),
    #         'skin thickness': slice(None),
    #         #'density': slice(None),
    #         'vol. heat cap.': slice(None),
    #         'wind speed': slice(None),
    #         'location': slice(None),
    #         'variable': ["T"],
    #         'res': [2.0],
    #         },
            
    #     'subtract': 273,
    #     'color': 'hw',
    #     'between_time': between_time,
    #     # 'between_time': ('00:00', '23:59'),
    #     'x': 't',
    #     'xlabel': 'Scaled distance [-]',
    #     'ylabel': y_label_deg_C,
    #     'hline': False,
    #     'xlim': (0, 5),
    #     'ylim': (12, 75),
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
        
    #     'show_minutes': [],
    #     },
        
    # f'T-plot, Surf Temp, at night alb={alb}, res=2,': 
    #     {'xs': {
    #         'category': 'results raw',
    #         'emissivity sky': slice(None),
    #         'hw': hws,
    #         # 'hw': slice(None),
    #         'albedo': [alb],
    #         'emissivity': slice(None),
    #         'skin thickness': slice(None),
    #         #'density': slice(None),
    #         'vol. heat cap.': slice(None),
    #         'wind speed': slice(None),
    #         'location': slice(None),
    #         'variable': ["T"],
    #         'res': [2.0],
    #         },
            
    #     'subtract': 273,
    #     'color': 'hw',
    #     # 'between_time': between_time,
    #     'between_time': ('04:00', '04:10'),
    #     'x': 't',
    #     'xlabel': 'Scaled distance [-]',
    #     'ylabel': y_label_deg_C,
    #     'hline': False,
    #     'xlim': (0, 5),
    #     'ylim': (12, 22),
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
        
    #     'show_minutes': [],
    #     },
        
    # f'T-plot, Surf Temp, alb={alb}, res=1,': 
    #     {'xs': {
    #         'category': 'results raw',
    #         'emissivity sky': slice(None),
    #         'hw': hws,
    #         # 'hw': slice(None),
    #         'albedo': [alb],
    #         'emissivity': slice(None),
    #         'skin thickness': slice(None),
    #         #'density': slice(None),
    #         'vol. heat cap.': slice(None),
    #         'wind speed': slice(None),
    #         'location': slice(None),
    #         'variable': ["T"],
    #         'res': [1.0],
    #         },
            
    #     'subtract': 273,
    #     'color': 'hw',
    #     'between_time': between_time,
    #     # 'between_time': ('00:00', '23:59'),
    #     'x': 't',
    #     'xlabel': 'Scaled distance [-]',
    #     'ylabel': y_label_deg_C,
    #     'hline': False,
    #     'xlim': (0, 5),
    #     'ylim': (12, 75),
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
        
    #     'show_minutes': [],
    #     },
    }


plot_flux_per_timestep(
    df=df, 
    plots_setup=plots_setup, 
    locations=locations, 
    # average_between_time=True,
    flip=flip, 
    # legend_kwargs={'bbox_to_anchor':(1.04, 0.5), 'loc':"center left", 'borderaxespad':0}
    legend_kwargs=None,
    )

#%%% Flux per t per h/w

plots_setup = {
    f'Variables-plot incoming splitted, H/W=1.0, alb={0.2}, res=2,': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [1.0],
            # 'hw': slice(None),
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['in_from_els']", "q_lw['in_from_sky']", "q_sw['in_from_els']",  "q_sw['in_dir'] + q_sw['in_diff']"],
            # 'variable': ["q_sw['in_from_els']"],
            'res': [2.0],
            },
            
        # 'subtract': 273,
        'color': 'variable',
        'linestyle': 'variable',
        'between_time': between_time,
        # 'between_time': ('00:00', '23:59'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'Radiation flux [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        'legend_title': '',
        
        # 'show_minutes': [],
        },
        
    f'Variables-plot incoming splitted, H/W=1.0, alb={0.6}, res=2,': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [1.0],
            # 'hw': slice(None),
            'albedo': [0.6],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['in_from_els']", "q_lw['in_from_sky']", "q_sw['in_from_els']",  "q_sw['in_dir'] + q_sw['in_diff']"],
            # 'variable': ["q_sw['in_from_els']"],
            'res': [2.0],
            },
            
        # 'subtract': 273,
        'color': 'variable',
        'linestyle': 'variable',
        'between_time': between_time,
        # 'between_time': ('00:00', '23:59'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'Radiation flux [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        'legend_title': '',
        
        # 'show_minutes': [],
        },
        
    f'Variables-plot incoming splitted, H/W=4.0, alb={0.2}, res=2,': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [4.0],
            # 'hw': slice(None),
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['in_from_els']", "q_lw['in_from_sky']", "q_sw['in_from_els']",  "q_sw['in_dir'] + q_sw['in_diff']"],
            # 'variable': ["q_sw['in_from_els']"],
            'res': [2.0],
            },
            
        # 'subtract': 273,
        'color': 'variable',
        'linestyle': 'variable',
        'between_time': between_time,
        # 'between_time': ('00:00', '23:59'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'Radiation flux [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        'legend_title':'',
        
        # 'show_minutes': [],
        },
        
    f'Variables-plot incoming splitted, H/W=0.0, alb={0.2}, res=2,': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [0.0],
            # 'hw': slice(None),
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["q_sw['in']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['in_from_els']", "q_lw['in_from_sky']", "q_sw['in_from_els']",  "q_sw['in_dir'] + q_sw['in_diff']"],
            # 'variable': ["q_sw['in_from_els']"],
            'res': [2.0],
            },
            
        # 'subtract': 273,
        'color': 'variable',
        'linestyle': 'variable',
        'between_time': between_time,
        # 'between_time': ('00:00', '23:59'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': r'Radiation flux [W m$^{-2}$]',
        'hline': False,
        'xlim': (0, 5),
        'ylim': (-10, 1050),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        'legend_title': '',
        
        # 'show_minutes': [],
        },
        
    
    }


plot_flux_per_timestep(
    df=df, 
    plots_setup=plots_setup, 
    locations=locations, 
    # average_between_time=True,
    flip=flip, 
    # legend_kwargs={'bbox_to_anchor':(1.04, 0.5), 'loc':"center left", 'borderaxespad':0}
    legend_kwargs=None,
    )


# %%% T per t over all sections

plots_setup = {
    f'T-plot, Surf Temp, alb={0.2}, res=2,': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["T"],
            'res': [2.0],
            },
            
        'subtract': 273,
        'color': 'hw',
        'between_time': between_time,
        # 'between_time': ('00:00', '23:59'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': y_label_deg_C,
        'hline': False,
        'xlim': (0, 5),
        'ylim': (12, 75),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'T-plot, Surf Temp, alb={0.6}, res=2,': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [0.6],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["T"],
            'res': [2.0],
            },
            
        'subtract': 273,
        'color': 'hw',
        'between_time': between_time,
        # 'between_time': ('00:00', '23:59'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': y_label_deg_C,
        'hline': False,
        'xlim': (0, 5),
        'ylim': (12, 75),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'T-plot, Surf Temp, at night alb={alb}, res=2,': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["T"],
            'res': [2.0],
            },
            
        'subtract': 273,
        'color': 'hw',
        # 'between_time': between_time,
        'between_time': ('04:00', '04:10'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': y_label_deg_C,
        'hline': False,
        'xlim': (0, 5),
        'ylim': (12, 22),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
        
    f'T-plot, Surf Temp, alb={0.2}, res=1,': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': hws,
            # 'hw': slice(None),
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': slice(None),
            'variable': ["T"],
            'res': [1.0],
            },
            
        'subtract': 273,
        'color': 'hw',
        'between_time': between_time,
        # 'between_time': ('00:00', '23:59'),
        'x': 't',
        'xlabel': 'Scaled distance [-]',
        'ylabel': y_label_deg_C,
        'hline': False,
        'xlim': (0, 5),
        'ylim': (12, 75),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        
        # 'show_minutes': [],
        },
    }


plot_flux_per_timestep(
    df=df, 
    plots_setup=plots_setup, 
    locations=locations, 
    flip=flip, 
    # legend_kwargs={'bbox_to_anchor':(1.04, 0.5), 'loc':"center left", 'borderaxespad':0}
    legend_kwargs=None,
    )




# %% STREET LEVEL

# %%% RADIATIONS AT STREET FOR DIFFERENT HWS VS TIME


plots_setup = {

    f'Absorbed SW Radiation from els at canyon street \n alb={alb}': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0, .5, 1, 2, 4],
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['in_from_els']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
        'multiply': (1-alb),
        'color': 'hw',
        'x': 't',
        'ylabel': y_label1,
        'ylim': (-20,650),
        #'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    f'Absorbed SW Radiation at canyon street \n alb={alb}': 
    
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0, .5, 1, 2, 4],
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['in']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
        'multiply': (1-alb),
        'color': 'hw',
        'x': 't',
        'ylabel': y_label1,
        'ylim': (-20,650),
        #'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    
        
    f'LW Radiation abs from env at canyon street \n alb={alb}': 
    
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0, .5, 1, 2, 4],
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_els']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
        'color': 'hw',
        'multiply':0.93,
        'x': 't',
        'ylabel': y_label1,
        'ylim': (-20,650),
        #'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    f'LW Radiation abs from sky at canyon street \n alb={alb}': 
    
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0, .5, 1, 2, 4],
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_sky']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
        'color': 'hw',
        'multiply':0.93,
        'x': 't',
        'ylabel': y_label1,
        'ylim': (-20,650),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    f'LW Radiation abs at canyon street \n alb={alb}': 
    
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0, .5, 1, 2, 4],
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_sky'] + q_lw['in_from_els']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
        'color': 'hw',
        'multiply':0.93,
        'x': 't',
        'ylabel': y_label1,
        'ylim': (-20,650),
        #'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    }

plot_lines(df, plots_setup, line_properties)



# %%% RADIATIONS VS H/W (or SVF) FOR PARTS OF THE DAY

albs = [0.2]

x_vars = ['hw', 'svf']
x_vars = ['hw']

for x_var in x_vars:
    for alb in albs:
        xlim = [0, 4.25] if x_var == 'hw' else [1.05, -0.05]
        for time_name, time_interval in time_dict.items():
            plots_setup = {
                
                f'SW and LW Radiation on street vs {x_var} \n {time_name} \n Albedo={alb}': 
                    {'xs': {
                        'category': 'results, per m2 surface',
                        'emissivity sky': slice(None),
                        'hw': slice(None),
                        'albedo': [alb],
                        'emissivity': slice(None),
                        'skin thickness': slice(None),
                        #'density': slice(None),
                        'vol. heat cap.': slice(None),
                        'wind speed': slice(None),
                        'location': [f'canyon street{slice_name}'],
                        'variable': [
                            "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs']",
                            "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs_from_sky']",
                            "q_sw['abs_from_els'] + q_lw['abs_from_els']",
                            # 'q_convection'
                            ],
                        'res': 2.0,
                        },
                        
                    # 'marker': 'variable',
                    'between_time': time_interval,
                    'linestyle': 'variable',
                    'color': 'variable',
                    'x': x_var,
                    'ylabel': y_label1,
                    # 'xlim': xlim_hw,
                    'xlim': xlim,
                    'ylim': (-20, 1100) if not 'no convection' in folder else (-20, 1300),
                    # 'y_unit': 'MJ/m2/d',
                    # 'ylim': ylim_MJ_normalised,
                    # 'y_scale': 1e-6,
                    'date': date,
                    'droplevel': ('zenith', 'azimuth'),
                    'legend': ['variable'],
                    },
                    
                f'SW and LW Radiation and convection on street vs {x_var} \n {time_name} \n Albedo={alb}': 
                    {'xs': {
                        'category': 'results, per m2 surface',
                        'emissivity sky': slice(None),
                        'hw': slice(None),
                        'albedo': [alb],
                        'emissivity': slice(None),
                        'skin thickness': slice(None),
                        #'density': slice(None),
                        'vol. heat cap.': slice(None),
                        'wind speed': slice(None),
                        'location': [f'canyon street{slice_name}'],
                        'variable': [
                            "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['net']",
                            "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs']",
                            "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs_from_sky']",
                            "q_sw['abs_from_els'] + q_lw['abs_from_els']",
                            'q_convection'
                            ],
                        'res': 2.0,
                        },
                        
                    # 'marker': 'variable',
                        
                    'between_time': time_interval,
                    'linestyle': 'variable',
                    'color': 'variable',
                    'x': x_var,
                    'ylabel': y_label1,
                    # 'xlim': xlim_hw,
                    'xlim': xlim,
                    'ylim': (-400, 1100) if not 'no convection' in folder else (-20, 1300),
                    # 'y_unit': 'MJ/m2/d',
                    # 'ylim': ylim_MJ_normalised,
                    # 'y_scale': 1e-6,
                    'date': date,
                    'droplevel': ('zenith', 'azimuth'),
                    'legend': ['variable'],
                    },
                    
                f'SW Radiation on street vs {x_var} \n {time_name} \n Albedo={alb}': 
                    {'xs': {
                        'category': 'results, per m2 surface',
                        'emissivity sky': slice(None),
                        'hw': slice(None),
                        'albedo': [alb],
                        'emissivity': slice(None),
                        'skin thickness': slice(None),
                        #'density': slice(None),
                        'vol. heat cap.': slice(None),
                        'wind speed': slice(None),
                        'location': [f'canyon street{slice_name}'],
                        'variable': [
                            "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff']", 
                            "q_sw['abs_dir'] + q_sw['abs_diff']", 
                            "q_sw['abs_from_els']",
                            ],
                        'res': 2.0,
                        },
                        
                    # 'marker': 'variable',
                        
                    'between_time': time_interval,
                    'linestyle': 'variable',
                    'color': 'variable',
                    'x': x_var,
                    'ylabel': y_label1,
                    # 'xlim': xlim_hw,
                    'xlim': xlim,
                    # 'y_unit': 'MJ/m2/d',
                    'ylim': (-20, 1100) if not 'no convection' in folder else (-20, 1300),
                    # 'y_scale': 1e-6,
                    'date': date,
                    'droplevel': ('zenith', 'azimuth'),
                    'legend': ['variable'],
                    },
                    
                f'SW Radiation on street, split vs {x_var} \n {time_name} \n Albedo={alb}': 
                    {'xs': {
                        'category': 'results, per m2 surface',
                        'emissivity sky': slice(None),
                        'hw': slice(None),
                        'albedo': [alb],
                        'emissivity': slice(None),
                        'skin thickness': slice(None),
                        #'density': slice(None),
                        'vol. heat cap.': slice(None),
                        'wind speed': slice(None),
                        'location': [f'canyon street{slice_name}'],
                        'variable': [
                            "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff']", 
                            "q_sw['abs_dir'] + q_sw['abs_diff']", 
                            "q_sw['abs_dir']", 
                            "q_sw['abs_diff']", 
                            "q_sw['abs_from_els']",
                            ],
                        'res': 2.0,
                        },
                        
                    # 'marker': 'variable',
                        
                    'between_time': time_interval,
                    'linestyle': 'variable',
                    'color': 'variable',
                    'x': x_var,
                    'ylabel': y_label1,
                    # 'xlim': xlim_hw,
                    'xlim': xlim,
                    # 'y_unit': 'MJ/m2/d',
                    'ylim': (-20, 1100) if not 'no convection' in folder else (-20, 1300),
                    # 'y_scale': 1e-6,
                    'date': date,
                    'droplevel': ('zenith', 'azimuth'),
                    'legend': ['variable'],
                    },
                    
                    
                f'LW Radiation on street vs {x_var} \n {time_name} \n Albedo={alb}': 
                    {'xs': {
                        'category': 'results, per m2 surface',
                        'emissivity sky': slice(None),
                        'hw': slice(None),
                        'albedo': [alb],
                        'emissivity': slice(None),
                        'skin thickness': slice(None),
                        #'density': slice(None),
                        'vol. heat cap.': slice(None),
                        'wind speed': slice(None),
                        'location': [f'canyon street{slice_name}'],
                        'variable': [
                            "q_lw['abs_from_sky'] + q_lw['abs_from_els']", 
                            "q_lw['abs_from_sky']", 
                            "q_lw['abs_from_els']", 
                            ],
                        'res': 2.0,
                        },
                        
                    # 'marker': 'variable',
                        
                    'between_time': time_interval,
                    'linestyle': 'variable',
                    'color': 'variable',
                    'x': x_var,
                    'ylabel': y_label1,
                    # 'xlim': xlim_hw,
                    'xlim': xlim,
                    'ylim': (-20, 1100) if not 'no convection' in folder else (-20, 1300),
                    # 'y_unit': 'MJ/m2/d',
                    # 'ylim': ylim_MJ_normalised,
                    # 'y_scale': 1e-6,
                    'date': date,
                    'droplevel': ('zenith', 'azimuth'),
                    'legend': ['variable'],
                    },
                    
                f'LW Radiation with net on street vs {x_var} \n {time_name} \n Albedo={alb}': 
                    {'xs': {
                        'category': 'results, per m2 surface',
                        'emissivity sky': slice(None),
                        'hw': slice(None),
                        'albedo': [alb],
                        'emissivity': slice(None),
                        'skin thickness': slice(None),
                        #'density': slice(None),
                        'vol. heat cap.': slice(None),
                        'wind speed': slice(None),
                        'location': [f'canyon street{slice_name}'],
                        'variable': [
                            "q_lw['abs_from_sky'] + q_lw['abs_from_els']", 
                            "q_lw['abs_from_sky']", 
                            "q_lw['abs_from_els']", 
                            "q_lw['net']", 
                            ],
                        'res': 2.0,
                        },
                        
                    # 'marker': 'variable',
                        
                    'between_time': time_interval,
                    'linestyle': 'variable',
                    'color': 'variable',
                    'x': x_var,
                    'ylabel': y_label1,
                    # 'xlim': xlim_hw,
                    'xlim': xlim,
                    'ylim': (-400, 1100) if not 'no convection' in folder else (-20, 1300),
                    # 'y_unit': 'MJ/m2/d',
                    # 'ylim': ylim_MJ_normalised,
                    # 'y_scale': 1e-6,
                    'date': date,
                    'droplevel': ('zenith', 'azimuth'),
                    'legend': ['variable'],
                    },
                    
                }
        
            plot_lines(df, plots_setup, line_properties, sort=False)
   
    
# %%% RADIATIONS VS H/W (or SVF) FOR PARTS OF THE DAY multiple hws

albs = [0.2, 0.6]

x_vars = ['hw', 'svf']
x_vars = ['hw']

for x_var in x_vars:
    xlim = [0, 4.25] if x_var == 'hw' else [1.05, -0.05]
    for time_name, time_interval in time_dict.items():
        plots_setup = {
            
            f'SW and LW Radiation on street vs {x_var} \n {time_name} \n Albedo={albs}': 
                {'xs': {
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': slice(None),
                    'albedo': albs,
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    #'density': slice(None),
                    'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'canyon street{slice_name}'],
                    'variable': [
                        "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs']",
                        "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs_from_sky']",
                        "q_sw['abs_from_els'] + q_lw['abs_from_els']",
                        # 'q_convection'
                        ],
                    'res': 2.0,
                    },
                    
                # 'marker': 'variable',
                    
                'between_time': time_interval,
                'linestyle': 'variable',
                'color': 'albedo',
                'x': x_var,
                'ylabel': y_label1,
                # 'xlim': xlim_hw,
                'xlim': xlim,
                # 'ylim': (-20, 1100) if not 'no convection' in folder else (-20, 1300),
                'ylim': (-520, 1050) if not 'no convection' in folder else (-20, 1300),
                # 'y_unit': 'MJ/m2/d',
                # 'ylim': ylim_MJ_normalised,
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'legend': ['albedo', 'variable'],
                },
                
            f'SW and LW Radiation and convection on street vs {x_var} \n {time_name} \n Albedo={albs}': 
                {'xs': {
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': slice(None),
                    'albedo': albs,
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    #'density': slice(None),
                    'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'canyon street{slice_name}'],
                    'variable': [
                        "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['net']",
                        "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs']",
                        "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs_from_sky']",
                        "q_sw['abs_from_els'] + q_lw['abs_from_els']",
                        'q_convection'
                        ],
                    'res': 2.0,
                    },
                    
                # 'marker': 'variable',
                    
                'between_time': time_interval,
                'linestyle': 'variable',
                'color': 'albedo',
                'x': x_var,
                'ylabel': y_label1,
                # 'xlim': xlim_hw,
                'xlim': xlim,
                # 'ylim': (-400, 1100) if not 'no convection' in folder else (-20, 1300),
                'ylim': (-520, 1050) if not 'no convection' in folder else (-20, 1300),
                # 'y_unit': 'MJ/m2/d',
                # 'ylim': ylim_MJ_normalised,
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'legend': ['albedo', 'variable'],
                },
                
            f'Q net Radiation and convection on street vs {x_var} \n {time_name} \n Albedo={albs}': 
                {'xs': {
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': slice(None),
                    'albedo': albs,
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    #'density': slice(None),
                    'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'canyon street{slice_name}'],
                    'variable': [
                        "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['net']",
                        'q_convection'
                        ],
                    'res': 2.0,
                    },
                    
                # 'marker': 'variable',
                    
                'between_time': time_interval,
                'linestyle': 'variable',
                'color': 'albedo',
                'x': x_var,
                'ylabel': y_label1,
                # 'xlim': xlim_hw,
                'xlim': xlim,
                # 'ylim': (-400, 1100) if not 'no convection' in folder else (-20, 1300),
                'ylim': (-520, 1050) if not 'no convection' in folder else (-20, 1300),
                # 'y_unit': 'MJ/m2/d',
                # 'ylim': ylim_MJ_normalised,
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'legend': ['albedo', 'variable'],
                },
                
            f'SW Radiation on street vs {x_var} \n {time_name} \n Albedo={albs}': 
                {'xs': {
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': slice(None),
                    'albedo': albs,
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    #'density': slice(None),
                    'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'canyon street{slice_name}'],
                    'variable': [
                        "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff']", 
                        "q_sw['abs_dir'] + q_sw['abs_diff']", 
                        "q_sw['abs_from_els']",
                        ],
                    'res': 2.0,
                    },
                    
                # 'marker': 'variable',
                    
                'between_time': time_interval,
                'linestyle': 'variable',
                'color': 'albedo',
                'x': x_var,
                'ylabel': y_label1,
                # 'xlim': xlim_hw,
                'xlim': xlim,
                # 'y_unit': 'MJ/m2/d',
                # 'ylim': (-20, 1100) if not 'no convection' in folder else (-20, 1300),                
                'ylim': (-520, 1050) if not 'no convection' in folder else (-20, 1300),
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'legend': ['albedo', 'variable'],
                },
                
            f'SW Radiation on street, split vs {x_var} \n {time_name} \n Albedo={albs}': 
                {'xs': {
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': slice(None),
                    'albedo': albs,
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    #'density': slice(None),
                    'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'canyon street{slice_name}'],
                    'variable': [
                        "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff']", 
                        "q_sw['abs_dir'] + q_sw['abs_diff']", 
                        "q_sw['abs_dir']", 
                        "q_sw['abs_diff']", 
                        "q_sw['abs_from_els']",
                        ],
                    'res': 2.0,
                    },
                    
                # 'marker': 'variable',
                    
                'between_time': time_interval,
                'linestyle': 'variable',
                'color': 'albedo',
                'x': x_var,
                'ylabel': y_label1,
                # 'xlim': xlim_hw,
                'xlim': xlim,
                # 'y_unit': 'MJ/m2/d',
                # 'ylim': (-20, 1100) if not 'no convection' in folder else (-20, 1300),
                'ylim': (-520, 1050) if not 'no convection' in folder else (-20, 1300),
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'legend': ['albedo', 'variable'],
                },
                
                
            f'LW Radiation on street vs {x_var} \n {time_name} \n Albedo={albs}': 
                {'xs': {
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': slice(None),
                    'albedo': albs,
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    #'density': slice(None),
                    'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'canyon street{slice_name}'],
                    'variable': [
                        "q_lw['abs_from_sky'] + q_lw['abs_from_els']", 
                        "q_lw['abs_from_sky']", 
                        "q_lw['abs_from_els']", 
                        ],
                    'res': 2.0,
                    },
                    
                # 'marker': 'variable',
                    
                'between_time': time_interval,
                'linestyle': 'variable',
                'color': 'albedo',
                'x': x_var,
                'ylabel': y_label1,
                # 'xlim': xlim_hw,
                'xlim': xlim,
                # 'ylim': (-20, 1100) if not 'no convection' in folder else (-20, 1300),
                'ylim': (-520, 1050) if not 'no convection' in folder else (-20, 1300),
                # 'y_unit': 'MJ/m2/d',
                # 'ylim': ylim_MJ_normalised,
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'legend': ['albedo', 'variable'],
                },
                
            f'LW Radiation with net and out on street vs {x_var} \n {time_name} \n Albedo={albs}': 
                {'xs': {
                    'category': 'results, per m2 surface',
                    'emissivity sky': slice(None),
                    'hw': slice(None),
                    'albedo': albs,
                    'emissivity': slice(None),
                    'skin thickness': slice(None),
                    #'density': slice(None),
                    'vol. heat cap.': slice(None),
                    'wind speed': slice(None),
                    'location': [f'canyon street{slice_name}'],
                    'variable': [
                        "q_lw['abs_from_sky'] + q_lw['abs_from_els']", 
                        "q_lw['abs_from_sky']", 
                        "q_lw['abs_from_els']", 
                        "q_lw['net'] - q_lw['in_from_els'] - q_lw['in_from_sky']", 
                        # "q_lw['net']", 
                        ],
                    'res': 2.0,
                    },
                    
                # 'marker': 'variable',
                    
                'between_time': time_interval,
                'linestyle': 'variable',
                'color': 'albedo',
                'x': x_var,
                'ylabel': y_label1,
                # 'xlim': xlim_hw,
                'xlim': xlim,
                'ylim': (-520, 1050) if not 'no convection' in folder else (-20, 1300),
                # 'y_unit': 'MJ/m2/d',
                # 'ylim': ylim_MJ_normalised,
                # 'y_scale': 1e-6,
                'date': date,
                'droplevel': ('zenith', 'azimuth'),
                'legend': ['variable', 'albedo'],
                },
                
            }
    
        plot_lines(df, plots_setup, line_properties, sort=False, legend_kwargs={'bbox_to_anchor':(1.04, 0.5), 'loc':"center left", 'borderaxespad':0})
    


# %%% T MEAN MIN MAX vs ...


# %%%% T vs t, min/mean/max

T_var = 'T'
T_loc = f'canyon street{slice_name}'
# lims = (289, 323)


# T_var = 'T_mr'
# hw = 0.0 if 'gras' in folder.lower() else 'street'
# T_loc = f'street{slice_name}'

lims = (10, 73)

albs = [0.2]

df_xs = df.loc[df.index.get_level_values('datetime').date == date].xs(T_var, axis=1, level='variable').xs(T_loc, axis=1, level='location').xs('results raw', axis=1, level='category')

# df_xs = df_xs.xs(0.2, axis=1, level='skin thickness')
df_xs.index = df_xs.index.droplevel(('zenith', 'azimuth'))

# df_xs



# hws = [0, 0.5, 1, 2, 4, 0.75, 1.5]
# hws = [1.0]


separate = False

hws = [0, 0.5, 1, 4]
hws = [0, 0.5, 1, 4]
hws = [0.0]
# hws=[]
# hws = [0.5, 0.8, .9, 1, 1.1, 1.2]
# hws = [4, 1, .5, 0]


for alb in albs:
    df_xs_alb=df_xs.xs(alb, axis=1, level='albedo')
    
    df_mean = df_xs_alb.applymap(lambda lst: sum(lst) / len(lst) if lst is not None else None)
    df_min = df_xs_alb.applymap(lambda lst: min(lst) if lst is not None else None)
    df_max = df_xs_alb.applymap(lambda lst: max(lst) if lst is not None else None)

    df_mean -= 273
    df_min -= 273
    df_max -= 273
    
    for i, hw in enumerate(hws):
        
        try:
            if T_var == 'T_mr':
                title = f'Mean Radiant Temperature perceived by a person on the street'
            elif T_var == 'T':
                title=f'Street Surface Temperature'
            
            title +=  f'\n albedo={alb}'
            if separate:
                title += f', H/W={hw}'
            else:
                title += f', H/W={hws}'
            
            
            if i == 0 or separate:
                fig, ax = plt.subplots(figsize=(5,3))
                
            # label_mean = 
            label = '' if separate else f'H/W={hw}, '
                
            color = line_properties['hw']['color'].get(hw, 'tab:grey') 
            df_min.xs(hw, axis=1, level='hw').iloc[:, 0].plot(alpha=0.6, linewidth=1.0, linestyle='--', color=color, ax=ax, label=label+r'$T_\mathrm{min}$')
            df_max.xs(hw, axis=1, level='hw').iloc[:, 0].plot(alpha=0.8, linewidth=1, linestyle='-', color=color, ax=ax, label=label+r'$T_\mathrm{max}$')
            # ax.fill_between(df_min.xs(hw, axis=1, level='hw').iloc[:, 0].index, 
            #                 df_min.xs(hw, axis=1, level='hw').iloc[:, 0], 
            #                 df_max.xs(hw, axis=1, level='hw').iloc[:, 0], 
            #                 color=color, 
            #                 alpha=0.3, 
            #                 # label=label+'min ... max',
            #                 )
            df_mean.xs(hw, axis=1, level='hw').iloc[:, 0].plot(ax=ax, label=label+r'$T_\mathrm{mean}$', linewidth=2.0, color=color)
    
            
            if 'model 2' in folder.lower() or True:
                plt.legend(**{'bbox_to_anchor':(1.04, 0.5), 'loc':"center left", 'borderaxespad':0})
            if 'no convection' in folder:
                plt.ylim(0, 120)
            else:
                plt.ylim(lims)
            # plt.title(f'H/W={hw}')
            plt.ylabel(r'Temperature [$^\circ$C]')
            plt.xlabel(x_label['t'])
            
            # title=f'Street Surface Temperature \n albedo=0.25, H/W={hw}'
            
            
            if SAVE:
                if not os.path.isdir(f'postprocessing plots/{folder}'):
                    os.mkdir(f'postprocessing plots/{folder}')
                
                if separate or (not separate and i == (len(hws)-1)):
                    print('saving')
                    plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(title)}.png', bbox_inches="tight")
            
            
            
            if separate:
                if TITLE:
                    plt.title(title)
                plt.show()
            else:
                if i == (len(hw)-1):
                    if TITLE:
                        plt.title(title)
                    print('show')
                    plt.show()
        except:
            pass
    
    
# %%%% T vs H/W or SVF, min/mean/maxdf.

df_xs = df.loc[df.index.get_level_values('datetime').date == date].xs('T', axis=1, level='variable').xs(f'canyon street{slice_name}', axis=1, level='location').xs('results raw', axis=1, level='category')
df_xs.index = df_xs.index.droplevel(('zenith', 'azimuth'))

# df_mean = df_xs.applymap(lambda lst: sum(lst) / len(lst))
# df_min = df_xs.applymap(lambda lst: min(lst))
# df_max = df_xs.applymap(lambda lst: max(lst))

legend = True

df_xs
albs=[0.2, 0.6]

df_svf = pd.DataFrame([])

for x_var in ['hw']:
# for x_var in ['svf', 'hw']:
    for time_name, time_interval in time_dict.items():
        fig, ax = plt.subplots(figsize=(5,3))
        for alb in albs:
            
            df_xs_alb = df_xs.xs(alb, axis=1, level='albedo')
            df_xs_alb.columns = df_xs_alb.columns.remove_unused_levels()
            # if x_axis == 'x_var':
            if x_var == 'svf':
                # df_xs_alb.columns = df_xs_alb.columns.set_levels(df_xs_alb.columns.get_level_values('hw').map(hw_to_svf), level='hw', verify_integrity=False).rename(names='svf', level='hw')
                df_xs_alb.columns = df_xs_alb.columns.set_levels(list(df_xs_alb.columns.get_level_values('hw').map(hw_to_svf)), level='hw').rename(names='svf', level='hw')
                # df_xs_alb = df_xs_alb.droplevel('hw',axis=1).append([)], keys=[df_xs_alb.columns.get_level_values('hw').map(hw_to_svf)], names=['svf'], axis=1)
            
            df_mean = df_xs_alb.applymap(lambda lst: sum(lst) / len(lst) if lst is not None else None)
            df_min = df_xs_alb.applymap(lambda lst: min(lst) if lst is not None else None)
            df_max = df_xs_alb.applymap(lambda lst: max(lst) if lst is not None else None)
        
            df_mean -= 273
            df_min -= 273
            df_max -= 273
        
            df_mean2 = df_mean.between_time(time_interval[0], time_interval[1]).groupby(x_var, axis=1).mean().mean(0)
            df_min2 = df_min.between_time(time_interval[0], time_interval[1]).groupby(x_var, axis=1).min().min(0)
            df_max2 = df_max.between_time(time_interval[0], time_interval[1]).groupby(x_var, axis=1).max().max(0)
            
            if x_var == 'svf':
                df_minmeanmax = pd.concat([df_mean2, df_min2, df_max2], axis=1, keys=['mean', 'min', 'max'], names=['statistics'])
                df_minmeanmax = pd.concat([df_minmeanmax], axis=1, keys=[alb], names=['albedo'])
                df_svf = pd.concat([df_svf, pd.concat({time_name: df_minmeanmax}, names=['time_interval'], axis=1)], axis=1)
            
            color = line_properties['albedo']['color'].get(alb, 'tab:grey') 
            label = rf'{alb}, '
        
            
            df_mean2.plot(ax=ax, label=label+r'$T_\mathrm{mean}$', linewidth=2.5, color=color)
            # df_max2.plot(alpha=0.8, linewidth=1.2, linestyle='-', color=color, ax=ax, label=label+r'$T_\mathrm{max}$')
            # df_min2.plot(alpha=0.8, linewidth=1.2, linestyle=':', color=color, ax=ax, label=label+r'$T_\mathrm{min}$')
            fill_between = True
            if fill_between:
                # ax.fill_between(df_min2.index, df_min2, df_max2, color=color, alpha=0.3, label=label+r'$T_\mathrm{min ... max}$')
                ax.fill_between(df_min2.index, df_min2, df_max2, color=color, alpha=0.3, label=label+r'$T$ (range)')
            # plt.legend(title='albedo,variable', bbox_to_anchor=(1.04, 0.5), loc="center left",borderaxespad=0)
            
            if not 'no convection' in folder:
                if 'sunrise' in time_name:
                    plt.ylim(12.5, 19.5)  
                elif '24-H' in time_name:
                    plt.ylim(10, 73)    
                else:
                    plt.ylim(27, 73)    
            else:
                if 'sunrise' in time_name:
                    plt.ylim(0, 15)  
                else:
                    plt.ylim(0, 120)  
            plt.ylabel(r'Steet temperature [$^\circ$C]')
            
            
            if x_var == 'svf':
                plt.xlim(1.05, -0.05)
                plt.xlabel('Street SVF [-]')
            else:
                plt.xlim(0, 4.25)
                plt.xlabel('H/W ratio [-]')
        
            alb_title = ' and '.join([str(alb).replace('.', '') for alb in albs])
            legend_title = "no legend" if not legend else ""
        
            title=f'{time_name} \n min, mean and max Street Surface Temperature albs {alb_title} {legend_title} {"with fill" if fill_between else ""} vs {x_var}'
        
            
        handles, labels = ax.get_legend_handles_labels()
        
        if fill_between:
            indices_to_rearrange = [i // 2 if i % 2 == 0 else len(handles) // 2 + i // 2 for i in range(len(handles))]
            handles = [handles[i] for i in indices_to_rearrange]
            labels = [labels[i] for i in indices_to_rearrange]
        if legend:
            ax.legend(handles=handles, labels=labels, title='albedo,variable', bbox_to_anchor=(1.04, 0.5), loc="center left",borderaxespad=0)
        else:
            pass
        
        if SAVE:
            if not os.path.isdir(f'postprocessing plots/{folder}'):
                os.mkdir(f'postprocessing plots/{folder}')
            
            plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(title)}.png', bbox_inches="tight")       
            
        if TITLE:
            plt.title(title)
            
        plt.show()
    
df_svf.to_pickle(f'demos/{folder}/df_svf.pkl')

 # %%%% T vs t, various wind speed
plots_setup = {
    
    'Surface temperature on street \n H/W=0.0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        # 'linewidth': [0.75, 1, 1.75, 2.5, 3.25],
        'linewidth': 'wind speed',

        'x': 't',
        'ylabel': y_label_T,
        # 'xlim': xlim_hw,
        'ylim': ylim_T,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['wind speed'],
        },
    
    'Surface temperature on street \n H/W=1.0, various wind speed': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        # 'linewidth': [0.75, 1, 1.75, 2.5, 3.25],
        'linewidth': 'wind speed',

        'x': 't',
        'ylabel': y_label_T,
        # 'xlim': xlim_hw,
        'ylim': ylim_T,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['wind speed'],
        },
    }
    
plot_lines(df, plots_setup, line_properties)

# %%%% T vs t, various skin thickness
plots_setup = {
    'Surface temperature on street \n H/W=0.0, various skin thickness': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.05, 0.1, 0.15, 0.01, 0.25],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': 0,
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linewidth': 'skin thickness',
        'x': 't',
        'ylabel': y_label_T,
        # 'xlim': xlim_hw,
        'ylim': ylim_T,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['skin thickness'],
        },
    
    
    'Surface temperature on street \n H/W=1.0, various skin thickness': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.05, 0.1, 0.15, 0.01, 0.25],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': 0,
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linewidth': 'skin thickness',
        'x': 't',
        'ylabel': y_label_T,
        # 'xlim': xlim_hw,
        'ylim': ylim_T,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['skin thickness'],
        },
    }
    
plot_lines(df, plots_setup, line_properties)


# %%%% T vs t, various albedo

plots_setup = {

    'Surface temperature on street \n H/W=0.0, various albedo': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0],
            'albedo': slice(None),
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'linestyle': 'albedo',
            
        'x': 't',
        'ylabel': y_label_T,
        # 'xlim': xlim_hw,
        'ylim': ylim_T,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    'Surface temperature on street \n H/W=1.0, various albedo': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': slice(None),
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'linestyle': 'albedo',
            
        'x': 't',
        'ylabel': y_label_T,
        # 'xlim': xlim_hw,
        'ylim': ylim_T,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    # 'Surface temperature on street \n H/W=2.0, various albedo': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [2.0],
    #         'albedo': slice(None),
    #         'emissivity': 1.0, 
    #         'skin thickness': [0.1],
    #         #'density': slice(None),
    #         'vol. heat cap.': slice(None),
    #         'wind speed': [0],
    #         # 'location': [f'canyon street{slice_name}', 'walls'],
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["T"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     # 'marker': 'variable',
            
    #     'x': 't',
    #     'ylabel': y_label_T,
    #     'ylim': ylim_T,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
    'Surface temperature on street \n H/W=4.0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [4.0],
            'albedo': slice(None),
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'linestyle': 'albedo',        
        
        'x': 't',
        'ylabel': y_label_T,
        # 'xlim': xlim_hw,
        'ylim': ylim_T,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    }

plot_lines(df, plots_setup, line_properties)


# %%%% T vs t, various hw


plots_setup = {

    # 'Surface temperature on street \n  various H/W, albedo=0.0': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': hws,
    #         'albedo': [0.2],
    #         'emissivity': slice(None),
    #         'skin thickness': slice(None),
    #         #'density': slice(None),
    #         'vol. heat cap.': slice(None),
    #         'wind speed': [0],
    #         # 'location': [f'canyon street{slice_name}', 'walls'],
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["T"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     # 'marker': 'variable',
    #     # 'color': colors[9],
    #     'linewidth': 'hw',
    #     'x': 't',
    #     # 'subtract': 273,
    #     'ylabel': y_label_deg_C,
    #     # 'xlim': xlim_hw,
    #     'ylim': ylim_T,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
    'Surface temperature on street \n  various H/W, albedo=0.1': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': hws,
            'albedo': [0.1],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        # 'marker': 'variable',
            
        # 'color': colors[9],
        'subtract': 273,
        'color': 'hw',
        'x': 't',
        'ylabel': y_label_T,
        # 'xlim': xlim_hw,
        'ylim': (None, 69),
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    'Surface temperature on street \n  various H/W, albedo=0.2': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': hws,
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        # 'marker': 'variable',
        # 'color': colors[9],
        'color': 'hw',
        'x': 't',
        'subtract': 273,
        'ylabel': y_label_T,
        'ylim': (None, 69),
        # 'ylim': ylim_T,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    'Surface temperature on street \n various H/W, albedo=0.6': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': hws,
            'albedo': [0.6],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        # 'marker': 'variable',
        'subtract': 273,
        'ylabel': y_label_T,
        'ylim': (None, 69),
        # 'color': colors[9],
        'color': 'hw',
        'x': 't',
        
        # 'xlim': xlim_hw,
        # 'ylim': ylim_T,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    }

plot_lines(df, plots_setup, line_properties)




# %%%% Delta T vs t (alb fixed) (HW!=0 - HW=0)


plots_setup = {

    'Delta Surface Temperature on canyon street \n (H/W>0 - H/W=0) \n Albedo=0.0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': ['0.5 - 0.0', '1.0 - 0.0', '2.0 - 0.0', '4.0 - 0.0'],
            'albedo': [0.0],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',        
        'linewidth': 'hw',
            
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': [-25, 4],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
        
    'Delta Surface Temperature on canyon street \n (H/W>0 - H/W=0) \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': ['0.5 - 0.0', '1.0 - 0.0', '2.0 - 0.0', '4.0 - 0.0'],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        'linewidth': 'hw',
            
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': [-25, 4],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    'Delta Surface Temperature on canyon street \n (H/W>0 - H/W=0) \n Albedo=0.5': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': ['0.5 - 0.0', '1.0 - 0.0', '2.0 - 0.0', '4.0 - 0.0'],
            'albedo': [0.5],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            # 'location': [f'canyon street{slice_name}', 'walls'],
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        'linewidth': 'hw',
        
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': [-25, 4],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    # f'Albedo=0.25 \n LW Radiation \n divided by the surface area': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         # 'location': [f'canyon street{slice_name}', 'walls'],
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["T"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     # 'marker': 'variable',
            
    #     'x': 't',
    #     'ylabel': y_label1,
    #     # 'xlim': xlim_hw,
    #     'ylim': ylim_T,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
    # f'Albedo=0.5 \n LW Radiation \n divided by the surface area': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
    #         'albedo': [0.5],
    #         'emissivity': 1.0, 
    #         # 'location': [f'canyon street{slice_name}', 'walls'],
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["T"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     # 'marker': 'variable',
            
    #     'x': 't',
    #     'ylabel': y_label1,
    #     'ylim': ylim_T,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
    # f'Albedo=1.0 \n LW Radiation \n divided by the surface area': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
    #         'albedo': [1.0],
    #         'emissivity': 1.0, 
    #         # 'location': [f'canyon street{slice_name}', 'walls'],
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["T"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     # 'marker': 'variable',
            
    #     'x': 't',
    #     'ylabel': y_label1,
    #     # 'xlim': xlim_hw,
    #     'ylim': ylim_T,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
    }

plot_lines(df, plots_setup, line_properties)


# %%% HEATMAP

# %%%% Heatmap T


albedo = 0.2

plots_setup = {}
# hws = [0.0, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 8.0]

for hw in hws:

    plots_setup.update({
        
        f'Heatmap T, on street \n H/W={float(hw)}, Albedo={float(albedo)}': 
            {'xs': {
                'category': 'results raw',
                'emissivity sky': slice(None),
                'hw': [hw],
                'albedo': [albedo],
                'emissivity': slice(None),
                'skin thickness': slice(None),
                #'density': slice(None),
                'vol. heat cap.': slice(None),
                'wind speed': slice(None),
                'location': [f'canyon street{slice_name}'],
                'variable': ["T"],
                # 'res': [1.0, 2.0],1
                'res': 2.0,
                },
                
            'slice': slice(6,10) if hw==0 else slice(None),
            'x': 't',
            'subtract': 273,
            # 'ylim': ylim_T,        
            'ylim': (12, 70),
            'title': f'T-heatmap for H/W={hw}',
            'cmap': 'coolwarm',
            'colorbar_label': 'Temperature [K]',
            'date': date,
            'droplevel': ('zenith', 'azimuth'),
            'legend': ['location'],
            
            'print daily sum': True,
            'linestyle': 'location',
            },
            
        })
    
    
        
plot_heatmap(df, plots_setup)    

    
# %%%% Heatmap fluxes
      
hw = 1.0
# emissivity = 0.85
    
plots_setup = {
        
    f'Heatmap Incoming SW Radiation reflected from els on street \n H/W={float(hw)}, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['in_from_els']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        'ylim': [0, 750],        
        'colorbar_label': colorbar_label_W,

        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
    f'Heatmap Incoming SW Radiation from sky on street \n H/W={float(hw)}, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['in_diff'] + q_sw['in_dir']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        'ylim': [0, 750],        
        'colorbar_label': colorbar_label_W,

        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
    f'Heatmap Incoming Direct SW Radiation from sky on street \n H/W={float(hw)}, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['in_dir']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        'ylim': [0, 750],        
        'colorbar_label': colorbar_label_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
    f'Heatmap Incoming Diffuse SW Radiation from sky on street \n H/W={float(hw)}, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['in_diff']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        'ylim': [0, None],        
        'colorbar_label': colorbar_label_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
    f'Heatmap Incoming LW Radiation from els on street \n H/W={float(hw)}, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_els']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        # 'ylim': [0, 750],        
        'colorbar_label': colorbar_label_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
    f'Heatmap Incoming LW Radiation from sky on street \n H/W={float(hw)}, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_sky']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        # 'ylim': [0, 600],        
        'colorbar_label': colorbar_label_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
    f'Heatmap Incoming LW Radiation on street \n H/W={float(hw)}, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_sky'] + q_lw['in_from_els']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        # 'ylim': [0, 600],        
        'colorbar_label': colorbar_label_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
        
    f'Heatmap Radiated LW Radiation on street \n H/W={float(hw)}, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['out_rad']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        # 'ylim': [0, 600],        
        'colorbar_label': colorbar_label_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
    
        
    }
    
plot_heatmap(df, plots_setup)    

    
# %%%% Heatmap bottom sensors

hw = 0.5
albedo = 0.2

plots_setup = {
    f'Heatmap Incoming SW Radiation at bottom of sensors \n H/W={float(hw)}, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'bottom{slice_name}'],
            'variable': ["q_sw_sensors['in']"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        # 'ylim': [0, 750],        
        'colorbar_label': colorbar_label_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
    }
    
plot_heatmap(df, plots_setup)    


# %%%% Heatmap Delta T

hw = 1.0
albedo = 0.2

plots_setup = {
    
    # f'Heatmap Delta T, on street \n H/W=0.0 - 0.0, Albedo={float(albedo)}': 
    #     {'xs': {
    #         'category': 'results raw',
    #         'emissivity sky': slice(None),
    #         'hw': ['0.0 - 0.0'],
    #         'albedo': [albedo],
    #         'emissivity': slice(None),
    #         'skin thickness': slice(None),
    #         #'density': slice(None),
    #         'vol. heat cap.': slice(None),
    #         'wind speed': slice(None),
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["T"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 2.0,
    #         },
            
    #     'x': 't',
    #     'ylim': [-20, 20],
    #     'cmap': 'coolwarm',
    #     'colorbar_label': 'Temperature [K]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'droplevel_col': ('skin thickness', 'vol. heat cap.'),
    #     'legend': ['location'],
    #     'title': r'$\Delta$T: H/W 0.5 - 0.0',
        
    #     'print daily sum': True,
    #     'linestyle': 'location',
    #     },
        
    f'Heatmap Delta T, on street \n H/W=0.5 - 0.0, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': ['0.5 - 0.0'],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        'ylim': [-20, 20],
        'cmap': 'coolwarm',
        'colorbar_label': 'Temperature [K]',
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'droplevel_col': ('skin thickness', 'vol. heat cap.'),
        'legend': ['location'],
        'title': r'$\Delta$T: H/W 0.5 - 0.0',
        
        'print daily sum': True,
        'linestyle': 'location',
        },
    
    f'Heatmap Delta T, on street \n H/W=1.0 - 0.0, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': ['1.0 - 0.0'],
            'albedo': [albedo],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        'ylim': [-20, 20],
        'cmap': 'coolwarm',
        'colorbar_label': 'Temperature [K]',
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'droplevel_col': ('skin thickness', 'vol. heat cap.'),
        'legend': ['location'],
        'title': r'$\Delta$T: H/W 1.0 - 0.0',
        
        'print daily sum': True,
        'linestyle': 'location',
        },
        
    f'Heatmap Delta T, on street \n H/W=2.0 - 0.0, Albedo={float(albedo)}': 
        {'xs': {
            'category': 'results raw',
            'emissivity sky': slice(None),
            'hw': ['2.0 - 0.0'],
            'albedo': [albedo],
            'emissivity': slice(None), 
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': ["T"],
            # 'res': [1.0, 2.0],1
            'res': 2.0,
            },
            
        'x': 't',
        'ylim': [-20, 20],
        'cmap': 'coolwarm',
        'colorbar_label': 'Temperature [K]',
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'droplevel_col': ('skin thickness', 'vol. heat cap.'),
        'legend': ['location'],
        'title': r'$\Delta$T: H/W 2.0 - 0.0',
        
        'print daily sum': True,
        'linestyle': 'location',
        },
        
    # f'Heatmap Delta T, on street \n H/W=2.0 - 0.0, Albedo={float(albedo)}': 
    #     {'xs': {
    #         'category': 'results raw',
    #         'emissivity sky': slice(None),
    #         'hw': ['2.0 - 0.0'],
    #         'albedo': [albedo],
    #         'emissivity': emissivities, 
    #         'skin thickness': slice(None),
    #         #'density': slice(None),
    #         'vol. heat cap.': slice(None),
    #         'wind speed': slice(None),
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["T"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 't',
    #     'ylim': [-20, 20],
    #     'cmap': 'coolwarm',
    #     'colorbar_label': 'Temperature [K]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'droplevel_col': ('skin thickness', 'vol. heat cap.'),
    #     'legend': ['location'],
    #     'title': r'$\Delta$T: H/W 2.0 - 0.0',
        
    #     'print daily sum': True,
    #     'linestyle': 'location',
    #     },
        
        
    }
    

plot_heatmap(df, plots_setup)    


#%%%% Heatmap MRT

albedo = 0.2

plots_setup = {}
# hws = [0.0, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 8.0]
hws = [0.0, 1.0, 2.0]

for hw in hws:

    plots_setup.update({
        
        f'Heatmap MRT for pedestrians \n H/W={float(hw)} Albedo={float(albedo)}': 
            {'xs': {
                'category': 'results raw',
                'emissivity sky': slice(None),
                'hw': [hw],
                'albedo': [albedo],
                'emissivity': slice(None),
                'skin thickness': slice(None),
                #'density': slice(None),
                'vol. heat cap.': slice(None),
                'wind speed': slice(None),
                'location': [f'street{slice_name}'],
                'variable': ["T_mr"],
                # 'res': [1.0, 2.0],1
                'res': 2.0,
                },
                
            'x': 't',
            'subtract': 273,
            # 'ylim': ylim_T,        
            'ylim': (12, 72),
            # 'title': f'MRT-heatmap for H/W={hw}',
            'cmap': 'coolwarm',
            'colorbar_label': 'Temperature [K]',
            'date': date,
            'droplevel': ('zenith', 'azimuth'),
            'legend': ['location'],
            
            'print daily sum': True,
            'linestyle': 'location',
            },
            
        })
    
plot_heatmap(df, plots_setup)   


#%%%% Heatmap Delta MRT

albedo = 0.2

plots_setup = {}
# hws = [0.0, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 8.0]
hws = [1.0, 2.0]

for hw in hws:

    plots_setup.update({
        
        f'Heatmap Delta MRT for pedestrians \n H/W={float(hw)} Albedo={float(albedo)}': 
            {'xs': {
                'category': 'results raw',
                'emissivity sky': slice(None),
                'hw': [f'{hw} - 0.0'],
                'albedo': [albedo],
                'emissivity': slice(None),
                'skin thickness': slice(None),
                #'density': slice(None),
                'vol. heat cap.': slice(None),
                'wind speed': slice(None),
                'location': [f'street{slice_name}'],
                'variable': ["T_mr"],
                # 'res': [1.0, 2.0],1
                'res': 2.0,
                },
                
            'x': 't',
            # 'subtract': 273,
            # 'ylim': ylim_T,        
            'ylim': (-20, 20),
            # 'title': f'MRT-heatmap for H/W={hw}',
            'cmap': 'coolwarm',
            'colorbar_label': 'Temperature [K]',
            'date': date,
            'droplevel': ('zenith', 'azimuth'),
            'droplevel_col': ('skin thickness', 'vol. heat cap.'),
            'legend': ['location'],
            
            'print daily sum': True,
            'linestyle': 'location',
            },
            
        })
    
plot_heatmap(df, plots_setup)   


# %% calc MRT df


T_var = 'T'
T_loc = f'canyon street{slice_name}'
lims = (289, 323)


T_var = 'T_mr'
# hw = 0.0 if 'gras' in folder.lower() else 'street'
T_loc = f'street{slice_name}'

lims = (15, 73)
lims = (0, 75)

df_xs = df.loc[df.index.get_level_values('datetime').date == date].xs(T_var, axis=1, level='variable').xs(T_loc, axis=1, level='location').xs('results raw', axis=1, level='category').xs(0, axis=1, level='wind speed')
df_xs = df_xs.dropna(how='all', axis=1)

# df_xs = df_xs.xs(0.2, axis=1, level='skin thickness')
df_xs.index = df_xs.index.droplevel(('zenith', 'azimuth'))

df_mean = df_xs.applymap(lambda lst: sum(lst) / len(lst))
df_min = df_xs.applymap(lambda lst: min(lst))
df_max = df_xs.applymap(lambda lst: max(lst))

df_xs = pd.concat({'mean': df_mean,
                   'min': df_min,
                   'max': df_max}, names=['min_max_mean'], axis=1)

# keep_cols = ['hw', 'category', 'albedo']
keep_cols = ['hw']
keep_cols = [level]
levels_to_drop = [l for l in df_xs.columns.names if (df_xs.columns.get_level_values(l).nunique() == 1 and not l in keep_cols)]
# levels_to_drop = [l for l in df_xs.columns.names if (df_xs.columns.get_level_values(l).nunique() == 1)]
# levels_to_drop += ['skin thickness', 'vol. heat cap.']
levels_to_drop += ['vol. heat cap.']
levels_to_drop = list(set(levels_to_drop)) # get unique
df_xs = df_xs.droplevel(levels_to_drop, axis=1)

# levels = [l for l in df_xs.columns.names if (df_xs.columns.get_level_values(l).nunique() > 1 and not l == 'category')]
levels = [l for l in df_xs.columns.names if (not l == 'min_max_mean')]

df_xs -= 273
df_xs = df_xs.sort_index(axis=1)

# df_xs = df_xs.xs(2, 1, 'res')

if 'real' in folder.lower():
    hws = [1.0, 0.0]
    # albedos = [slice(None)]
    albedos = df.columns.unique('albedo')
elif 'custom' in folder.lower():
    hws = [0.0, 0.5, 1.0, 2.0]
    albedos = [0.0, 0.1, 0.2, 0.4, 0.6]

# %%% MRT, Min Max Mean vs t 

show_fill = True

if True:
    for vals in product(*[(df_xs.columns.unique(l)) for l in levels]):
        fig, ax = plt.subplots()
        try:
            if T_var == 'T_mr':
                title = f'Mean Radiant Temperature perceived by a person on the street'
            else:
                title = f'Street Surface Temperature'
                
            title += f'\n {[f"{level}={val}" for val, level in zip(vals, levels)]}'
            
            # color = line_properties['hw']['color'].get(hw, colors[0])
            color = colors[0]
            
            df_xs_xs = df_xs.xs(vals, axis=1, level=levels)
            
            df_xs_xs['mean'].plot(ax=ax, label=f'mean', linewidth=2, color=color)
            df_xs_xs['min'].plot(ax=ax, alpha=0.6, linewidth=0.8, linestyle=':', color=color, label=f'min')
            df_xs_xs['max'].plot(ax=ax, alpha=0.6, linewidth=0.8, linestyle='-', color=color, label=f'max')
            if show_fill and df_xs_xs.shape[1] == 3:
                ax.fill_between(
                    df_xs_xs.index, 
                    df_xs_xs['min'], 
                    df_xs_xs['max'], 
                    color=color, alpha=0.3)
                
            plt.legend()
                
            ax.set_ylim(lims)
            
            ax.set_ylabel('MRT [$^{\circ}$C]')
            ax.set_xlabel('Local time [H]')
            
            if SAVE:
                if not os.path.isdir(f'postprocessing plots/{folder}'):
                    os.mkdir(f'postprocessing plots/{folder}')
                
                plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(title)}.png', bbox_inches="tight")
            
            ax.set_title(title)
            
    
        except KeyError:
            pass
            plt.close()
        
        plt.show()
        
# all albedos
if False:
# if True:
    for hw in hws:
    
        fig, ax = plt.subplots()
        try:
            if 'REAL' in folder:
                title = f'Mean Radiant Temperature perceived by a person on the street'
            else:
                title=f'Street Surface Temperature \n H/W={hw}, various albedos'
            
            try:
                color = [line_properties['albedo']['color'].get(alb, colors[0]) for alb in albedos]
            except TypeError:
                color = colors
                
            
            df_xs.xs('mean', axis=1, level='category').xs(hw, axis=1, level='hw').plot(ax=ax, label=f'mean, {alb}', linewidth=2, color=color)
            # df_xs.xs('min', axis=1, level='category', drop_level=False).xs(hw, axis=1, level='hw').xs(alb, axis=1, level='albedo').plot(ax=ax, alpha=0.6, linewidth=0.8, linestyle=':', color=color, label=f'min, {alb}')
            # df_xs.xs('max', axis=1, level='category', drop_level=False).xs(hw, axis=1, level='hw').xs(alb, axis=1, level='albedo').plot(ax=ax, alpha=0.6, linewidth=0.8, linestyle='-', color=color, label=f'max, {alb}')
            # if show_fill:
            #     ax.fill_between(
            #         df_xs.index, 
            #         df_xs.xs('min', axis=1, level='category', drop_level=False).xs(hw, axis=1, level='hw').xs(alb, axis=1, level='albedo').iloc[:,0], 
            #         df_xs.xs('max', axis=1, level='category', drop_level=False).xs(hw, axis=1, level='hw').xs(alb, axis=1, level='albedo').iloc[:,0], 
            #         color=color, alpha=0.3)
            ax.set_ylim(lims)
            
            ax.set_ylabel('MRT [$\deg$ C]')
            ax.set_xlabel('Local time [H]')
            
            # title=f'Street Surface Temperature \n albedo=0.25, H/W={hw}'
            # if TITLE:
            #     ax.set_title(title)
            
            if SAVE:
                if not os.path.isdir(f'postprocessing plots/{folder}'):
                    os.mkdir(f'postprocessing plots/{folder}')
                
                plt.savefig(f'postprocessing plots/{folder}/{transform_to_filename(title)}.png', bbox_inches="tight")
            
            ax.set_title(title)
            
    
        except KeyError:
            pass
        
        plt.show()
    
    

# plots_setup = {}

# for key in keys:
#     plots_setup.update({
#         f'Ave MRT vs t, various albedos H/W=1, \n {level}={key}': 
#             {'xs': {
#                 'category': 'mean',
#                 'emissivity sky': slice(None),
#                 # 'hw': [1.0, 0.0, 0.5, 2.0, 3.0, 4.0],
#                 'hw': [1.0],
#                 'albedo': albedos,
#                 # 'emissivity': slice(None),
#                 # 'skin thickness': slice(None),
#                 # 'skin thickness': [0.05],
#                 #'density': slice(None),
#                 'vol. heat cap.': slice(None),
#                 'wind speed': slice(None),
#                 'location': [f'street{slice_name}'],
#                 # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
#                 # 'variable': ["q_sw_sensors['in']"],
#                 'variable': ['T_mr'],
#                 'res': slice(None),
#                 # 'res': slice(None),
#                 level: [key],
#                 },
                
#             # 'linestyle': 'variable',
#             # 'print daily sum': True,
#             # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
#             'color': 'albedo',
#             'annotations': True,
#             'x': 't',
#             'ylabel': r'MRT [$^\circ$C]',
#             'hline': False,
#             # 'xlim': xlim_hw,
#             'ylim': (0, 80),
#             # 'y_scale': 1e-6,
#             'date': date,
#             # 'droplevel': ('zenith', 'azimuth'),
#             'legend': ['albedo'],
#             },
            
#         })

# plot_lines(df_xs, plots_setup, line_properties)
    
    
# %%% ave MRT vs t

plots_setup = {}

keys = df_xs.columns.unique('min_max_mean')

for key in keys:
    plots_setup.update({
        f'Ave MRT vs t, various albedos {keys.name}={key}': 
            {'xs': {
                # 'min_max_mean': ['min', 'mean'],
                # 'emissivity sky': slice(None),
                'hw': [0.0, 0.5, 1.0, 2.0, 4.0],
                # 'hw': slice(None),
                # 'albedo': albedos,
                # # 'emissivity': slice(None),
                # # 'skin thickness': slice(None),
                # # 'skin thickness': [0.05],
                # #'density': slice(None),
                # 'vol. heat cap.': slice(None),
                # 'wind speed': slice(None),
                # 'location': [f'street{slice_name}'],
                # # 'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
                # # 'variable': ["q_sw_sensors['in']"],
                # 'variable': ['T_mr'],
                # 'res': slice(None),
                # 'res': slice(None),
                keys.name: [key],
                },
                
            # 'linestyle': 'variable',
            # 'print daily sum': True,
            # 'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
            'color': 'hw',
            'annotations': True,
            'x': 't',
            'ylabel': r'MRT [$^\circ$C]',
            'hline': False,
            # 'xlim': xlim_hw,
            'ylim': (0, 80),
            # 'y_scale': 1e-6,
            'date': date,
            # 'droplevel': ('zenith', 'azimuth'),
            'legend': ['hw'],
            },
            
        })

plot_lines(df_xs, plots_setup, line_properties)



#%% --- semi-old below
    
# %% overview all vs t (HW=0, alb=0.25)

if 'East-West' in folder:
    labels = ['Roof', 'Canyon street', 'North Wall', 'South Wall']
    location = [f'top{slice_name}', f'canyon street{slice_name}', 'north-facing wall (x-slice)', 'south-facing wall (x-slice)']
    linestyle = ['-', '-.', '--', ':', (0, (2, 1)), (0, (2, 3))]
elif 'North-South' in folder:
    labels = ['Roof', 'Canyon street', 'East Wall', 'West Wall']
    linestyle = ['-', '-.', '--', ':', (0, (2, 1)), (0, (2, 3))]
    location = ['top (y-slice)', f'canyon street{slice_name}', 'east-facing wall (y-slice)', 'west-facing wall (y-slice)']
    
    
elif 'Single Square Building' in folder:
    # labels = ['Roof', 'Canyon street', 'East Wall', 'West Wall', 'North Wall', 'South Wall']
    # labels = ['East-facing wall', 'North-facing wall', 'South-facing wall', 'Roof', 'West-facing wall']

    # location = ['east-facing wall', 'north-facing wall', 'south-facing wall', 'upwards-facing roof', 'west-facing wall']
    labels = ['East-facing wall', 'North-facing wall',     'South-facing wall',     'Roof',     'West-facing wall']
    location = ['upwards-facing roof', 'north-facing wall', 'south-facing wall', 'east-facing wall', 'west-facing wall']
    linestyle = [':', '-.',  (0, (3, 2)), '-', (0, (15, 3))]
        

plots_setup = {
    
    'Incoming SW Radiation at all surfaces \n H/W=1.0, Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': location,
            'variable': ["q_sw['in']", "q_sw_sensors['in']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        # 'ylim': ylim_W,
        'ylim': [-50, 1050],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
    'Incoming direct SW Radiation at all surfaces \n H/W=1.0, Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': location,
            'variable': ["q_sw['in_dir']", "q_sw_sensors['in_dir']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        # 'ylim': ylim_W,
        'ylim': [-50, 1050],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,        
        'print daily sum': True,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
    'Incoming diffuse SW Radiation at all surfaces \n H/W=1.0, Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': location,
            'variable': ["q_sw['in_diff']", "q_sw_sensors['in_diff']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        # 'ylim': ylim_W,
        'ylim': [-50, 1050],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        'print daily sum': True,
        # 'labels': labels,
        'linestyle': 'location',
        # 'linewidth': [1, 1, 1, 1, 1,1,1,1,1,1],
        },
        
        
    'Absorbed SW Radiation at all surfaces \n H/W=1.0, Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': location,
            'variable': ["q_sw['in']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        'multiply': 'albedo',
        # 'ylim': ylim_W,
        'ylim': [-50, 1050],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        },
        
    'Incoming SW Radiation from els at all surfaces \n H/W=1.0, Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': location,
            'variable': ["q_sw['in_from_els']", "q_sw_sensors['in_from_els']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        # 'ylim': ylim_W,
        # 'ylim': [-50, 800],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        },
        
    'Incoming SW Radiation from sky at all surfaces \n H/W=1.0, Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': location,
            'variable': ["q_sw['in_dir']", "q_sw_sensors['in_dir']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        # 'ylim': ylim_W,
        'ylim': [-50, 1050],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        'linewidth': [1, 1, 1, 1, 1],
        },
    
        
        
    'Incoming LW Radiation from sky at all surfaces \n H/W=1.0, Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': location,
            'variable': ["q_lw['in_from_sky']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': [-50, 1050],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        'print daily sum': True,
        # 'labels': labels,
        'linestyle': 'location',
        },
        
    'Incoming LW Radiation from elements at all surfaces \n H/W=1.0, Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': location,
            'variable': ["q_lw['in_from_els']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        # 'ylim': [300, 500],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        'print daily sum': True,
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        },
        
    'Outgoing LW Radiation at all surfaces \n H/W=1.0, Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': location,
            'variable': ["q_lw['out_rad']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': [375, 550],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        # 'labels': labels,
        
        'print daily sum': True,
        'linestyle': 'location',
        },
        
        
    
        
    'Net All-wave Radiation at all surfaces \n H/W=1.0, Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': location,
            'variable': ["q_lw['net'] + q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': [-125, 500],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location'],
        
        # 'labels': labels,
        'print daily sum': True,
        'linestyle': 'location',
        },
        
    }
    
plot_lines(df, plots_setup, line_properties)

    
# %% SW vs t
    
# %%% SW vs t (H/W=0)

plots_setup = {
    
    'Absorbed SW Radiation at canyon street\nH/W=0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0],
            'albedo': [0.2],
            'emissivity': 0.85, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linestyle': 'albedo',
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
        
    }
    
plot_lines(df, plots_setup, line_properties)    
# %%% SW vs t (H/W=1.0)
    
plots_setup = {
    
    'Absorbed SW Radiation at canyon street \n H/W=1.0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.0, 0.25, 0.5, 1.0, 0.4, 0.2],
            'emissivity': [0.85, 0.95, 1.0], 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linestyle': 'albedo',    
        
        'x': 't',
        'ylabel': y_label1,
        'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
    
    'Absorbed SW Radiation at canyon walls \n H/W=1.0': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.0, 0.25, 0.5, 1.0],            
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['walls'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linestyle': 'albedo',
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    'Absorbed SW Radiation at canyon walls, normalised \n H/W=1.0'  : 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.0, 0.25, 0.5, 1.0],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['walls'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
            
        'linestyle': 'albedo',
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
    
    'Reflected SW Radiation out from canyon \n H/W=1.0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.0, 0.25, 0.5, 1.0],
            'emissivity': [0.85, 1.0], 
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'bottom{slice_name}'],
            'variable': ["q_sw_sensors['in']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
            
        'linestyle': 'albedo',
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    }

plot_lines(df, plots_setup, line_properties)

# %%% Eff albedo vs. t

hw = 0.5

plots_setup = {
    f'Canyon effective albedo \n (canyon reflected/incoming) \n H/W={hw}, albedo=0.2 \n Summer (15 jun.)': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [0.2],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            # 'location': [f'bottom{slice_name} / top{slice_name}'],
            'location': [f'bottom{slice_name} / top{slice_name}'],
            'variable': ["q_sw_sensors['in']"],
            'res': 1.0,
            },
            
        'linestyle': 'albedo',
            
        'x': 't',
        'ylabel': 'Canyon Effective albedo [-]',
        'ylim': [0.2, 0.6],
        'date': datetime.date(2022, 6, 15),
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    f'Canyon effective albedo \n (canyon reflected/incoming) \n H/W={hw} albedo=0.2 \n Winter (3 dec.)': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [0.2],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            # 'location': [f'bottom{slice_name} / top{slice_name}'],
            'location': [f'bottom{slice_name} / top{slice_name}'],
            'variable': ["q_sw_sensors['in']"],
            'res': 1.0,
            },
            
        'linestyle': 'albedo',
            
        'x': 't',
        'ylabel': 'Canyon Effective albedo [-]',
        'ylim': [0.2, 0.6],
        'date': datetime.date(2022, 12, 3),
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    f'Canyon SW absorption enhancement factor \n (surf. alb. / canyon eff. alb.) \n H/W={hw}': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [hw],
            'albedo': [0, 0.25, 0.5, 1.0],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'top{slice_name} / bottom{slice_name}'],
            'variable': ["q_sw_sensors['in']"],
            'res': 1.0,
            },
            
        'linestyle': 'albedo',
            
        'multiply': 'albedo',
        'x': 't',
        'ylabel': 'Absorption enhancement factor [-]',
        'ylim': [0, 3],
        'date': datetime.date(2022, 6, 15),
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    # 'Canyon SW absorption enhancement factor \n (surf. alb. / canyon eff. alb.) \n H/W=1.0, all albedo': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': slice(None),
    #         'emissivity': 1.0, 
    #         'skin thickness': [0.1],
    #         #'density': slice(None),
    #         'vol. heat cap.': slice(None),
    #         'wind speed': [0],
    #         'location': [f'top{slice_name} / bottom{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         'res': 1.0,
    #         },
            
    #     'linestyle': 'albedo',
            
    #     'multiply': 'albedo',
    #     'x': 't',
    #     'ylabel': 'Absorption enhancement factor [-]',
    #     'ylim': [-0.2, 3],
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
    }
        
plot_lines(df, plots_setup, line_properties)
        
    # %%% SW vs t (alb=0.25)
    
plots_setup = {
    'Absorbed SW Radiation at canyon street \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linewidth': 'hw',
        
        'x': 't',
        'ylabel': y_label1,
        'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
    
    'Absorbed SW Radiation at canyon walls \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['walls'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linewidth': 'hw',
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    'Absorbed SW Radiation at canyon walls, normalised \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['walls'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linewidth': 'hw',
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
    
    'Reflected SW Radiation out from canyon \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'bottom{slice_name}'],
            'variable': ["q_sw_sensors['in']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linewidth': 'hw',
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    'Canyon effective albedo \n (canyon reflected/incoming) \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'bottom{slice_name} / top{slice_name}'],
            'variable': ["q_sw_sensors['in']"],
            'res': 1.0,
            },
            
        'linewidth': 'hw',
            
        'x': 't',
        'ylabel': 'Canyon Effective albedo [-]',
        'ylim': [-0.2, 1.5],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    # f'Albedo=0.25 \n Canyon refl. efficiency (Canyon effective alb./surf. alb)': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name} / top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         'res': 1.0,
    #         },
            
    #     'divide': 'albedo',
    #     'x': 't',
    #     'ylabel': 'Canyon Effective albedo [-]',
    #     'ylim': [-0.2, 1.5],
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
    
    'Canyon SW absorption enhancement factor \n (surf. alb. / daily canyon eff. alb.) \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'top{slice_name} / bottom{slice_name})'],
            'variable': ["q_sw_sensors['in']"],
            'res': 1.0,
            },
            
        'linewidth': 'hw',
            
        'multiply': 'albedo',
        'x': 't',
        'ylabel': 'Absorption enhancement factor [-]',
        'ylim': [-0.2, 10],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    
        
    
        
    # !!! FIX!
    # f' albedo=0.25 \n Absorbed SW Radiation enhancement factor, at canyon street': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         # 'hw': ['1.0 / 0.0', '0.5 / 0.0'],
    #         'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
    #         # 'albedo': [0.25],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["q_sw['abs']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 't',
    #     'ylabel': y_label1,
    #     # 'ylim': ylim_W,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
    }
    
    
plot_lines(df, plots_setup, line_properties)
  
# %%% Delta (vs h/w=0) absorbed SW trapping vs t


# problem is that operationd can have only one key 'hw'. fix

plots_setup = {
    'Delta Absorbed SW Radiation at canyon street \n (H/W>0 - H/W=0) \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': ['0.5 - 0.0', '1.0 - 0.0', '2.0 - 0.0', '4.0 - 0.0'],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linewidth': 'hw',
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': [-725, 25],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    'Delta Absorbed SW Radiation at all canyon surfaces \n (H/W>0 - H/W=0) \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': ['0.5 - 0.0', '1.0 - 0.0', '2.0 - 0.0', '4.0 - 0.0'],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['total'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linewidth': 'hw',
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': [-725, 25],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    
    'Delta Absorbed SW Radiation at canyon street \n (H/W>0 - H/W=0) \n Albedo=0.0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': ['0.5 - 0.0', '1.0 - 0.0', '2.0 - 0.0', '4.0 - 0.0'],
            'albedo': [0.0],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linewidth': 'hw',
            
        'x': 't',
        'ylabel': y_label1,
        'ylim': [-725, 25],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    
    'Delta Absorbed SW Radiation at canyon street \n (H/W>0 - H/W=0) \n Albedo=0.5': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': ['0.5 - 0.0', '1.0 - 0.0', '2.0 - 0.0', '4.0 - 0.0'],
            'albedo': [0.5],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
            
        'linewidth': 'hw',
        'x': 't',
        'ylabel': y_label1,
        'ylim': [-725, 25],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    }
    
    
# plots_setup = {
#     'Daily Delta Absorbed SW Radiation at canyon street \n (H/W!=0 - H/W=0)': 
#         {'xs': {
#             'category': 'results, per m2 surface',
#             'emissivity sky': slice(None),
#             'hw': ['0.5 - 0.0', '1.0 - 0.0', '2.0 - 0.0', '4.0 - 0.0'],
#             'albedo': [0.0, 0.25, 0.5, 1.0],
#             'emissivity': 1.0, 
#             'location': [f'canyon street{slice_name}'],
#             'variable': ["q_sw['abs']"],
#             # 'res': [1.0, 2.0],1
#             'res': 1.0,
#             },
            
#         'x': 'hw',
#         'ylabel': y_label2,
#         'ylim': ylim_MJ_normalised,
#         'date': date,
#         'droplevel': ('zenith', 'azimuth'),
#         'legend': ['hw'],
#         },
        
#     }
    
plot_lines(df, plots_setup, line_properties)

# %% SW vs h/w

# %%% Daily absorbed SW vs h/w
    
plots_setup = {
    'Daily Absorbed SW Radiation at street': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': [0.0, 0.25, 0.5],
            'emissivity': slice(None), 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linestyle':'albedo',
            
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,
        'ylim': ylim_MJ_normalised,
        'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
    
    
    'Daily Absorbed SW Radiation at walls': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': [0.0, 0.25, 0.5],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['walls'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linestyle':'albedo',
            
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,
        'ylim': ylim_MJ_normalised,
        'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
        
    'Daily Absorbed SW Radiation at walls, normalised': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': [0.0, 0.25, 0.5],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['walls'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linestyle':'albedo',
            
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,        
        'ylim': ylim_MJ_normalised,
        'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
        
    'Daily Reflected SW Radiation outwards from canyon': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': [0.0, 0.25, 0.5],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'bottom{slice_name}'],
            'variable': ["q_sw_sensors['in']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linestyle':'albedo',
            
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,
        'ylim': ylim_MJ_normalised,
        'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    
        
    }
    
plot_lines(df, plots_setup, line_properties)

# %%% Daily effective albedo / reflective efficiency / enhancement factor

plots_setup = {
        
    'Daily Canyon effective albedo \n (canyon reflected/incoming)': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': slice(None),
            'emissivity': slice(None), 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'bottom{slice_name} / top{slice_name}'],
            'variable': ["q_sw_sensors['in']"],
            'res': 1.0,
            },
            
        'linestyle':'albedo',
            
        'x': 'hw',
        'ylabel': 'Canyon Effective albedo [-]',
        'ylim': [-0.2, 1.5],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
        
        
    # f'Daily Canyon refl. efficiency \n (Canyon effective alb./surf. alb)': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': slice(None),
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name} / top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         'res': 1.0,
    #         },
            
    #     'divide': 'albedo',
    #     'x': 'hw',
    #     'ylabel': 'Canyon Effective albedo [-]',
    #     'ylim': [-0.2, 1.5],
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
        
    'Canyon SW absorption enhancement factor \n (surf. alb / daily canyon eff. albedo)': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': slice(None),
            'emissivity': slice(None), 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'top{slice_name} / bottom{slice_name})'],
            'variable': ["q_sw_sensors['in']"],
            'res': 1.0,
            },
            
        'linestyle':'albedo',
            
        'multiply': 'albedo',
        'x': 'hw',
        'ylabel': 'Absorption enhancement factor [-]',
        'ylim': [-0.2, 5],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    }


plot_lines(df, plots_setup, line_properties)
  



# %%% Delta (vs h/w=0) Daily absorbed SW vs h/w


plots_setup = {
    'Delta Daily absorbed SW Radiation at canyon street \n (H/W>0 - H/W=0)': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            # 'hw': ['0.5 - 0.0', '1.0 - 0.0', '2.0 - 0.0', '4.0 - 0.0'],
            'hw': slice(None),
            'albedo': slice(None),
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'linestyle':'albedo',
            
        'subtract': 'hw=0',
            
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,
        'y_scale': 1e-6,
        # 'ylim': [-725, 25],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    
    }
    
plot_lines(df, plots_setup, line_properties)
    


# %% LW vs t

plots_setup = {
    
    
    # f'LW Radiation \n divided by the ground area': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}', 'walls'],
    #         'variable': ["q_lw['in_from_els']", "q_lw['in_from_sky'] + q_lw['in_from_els']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'marker': 'variable',
    #         \
    #     'x': 't',
    #     'ylabel': y_label1,
    #     # 'xlim': xlim_hw,
    #     'ylim': ylim_W,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location', 'variable'],
    #     },
        
    'LW Radiation on street \n divided by the surface area \n H/W=0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0],
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_sky']", "q_lw['out_rad']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'marker': 'variable',
        'color': 'variable',
            
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['location', 'variable'],
        },
        
        
    'LW Radiation on street \n divided by the surface area \n H/W=1.0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_els']", "q_lw['in_from_sky'] + q_lw['in_from_els']", "q_lw['out_rad']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        # linestyle=['']
            
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': [-20, 430],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
        
    'LW Radiation on street from sky \n various H/w': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': hws,
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_sky']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        # linestyle=['']
        'linewidth': 'hw',
            
        # 'color': colors[2],
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': [-20, 430],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    'LW Radiation on street from elements \n various H/w': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': hws,
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_els']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        # linestyle=['']
        # 'color': colors[1],
        'linewidth': 'hw',

        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': [-20, 430],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    'LW Radiation on street from sky + elements \n various H/w': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': hws,
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_els'] + q_lw['in_from_sky']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        # linestyle=['']
        # 'color': colors[3],
        'linewidth': 'hw',
            
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': [320, 450],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
    
    'LW Radiation on street out emmited \n various H/w': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': hws,
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['out_rad']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        # linestyle=['']
        # 'color': colors[5],
        'linewidth': 'hw',
        
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        # 'ylim': [320, 450],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },    
    
        
        
    'LW Radiation on street net \n various H/w': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': hws,
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['net']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        # linestyle=['']
        # 'color': colors[4],
        'linewidth': 'hw',
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': [-200, 5],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
    'LW Radiation on street net \n H/W=0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0],
            'albedo': [0.2],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_sky']", "q_lw['in_from_els']", "q_lw['in_from_sky'] + q_lw['in_from_els']", "q_lw['out_rad']", "q_lw['net']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        # linestyle=['']
        # 'color': [colors[1], colors[2], colors[3], colors[4], colors[5], colors[6]],
        'color': 'variable',
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        # 'ylim': [-165, 5],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
        
        
    # 'LW Radiation on street net \n various H/W, Albedo=0.0': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
    #         'albedo': [0.0],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["q_lw['net']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     # 'marker': 'variable',
    #     # linestyle=['']
            
    #     'x': 't',
    #     'ylabel': y_label1,
    #     # 'xlim': xlim_hw,
    #     'ylim': [-165, 5],
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
        
    # f'Daily LW Radiation emitted \n divided by the surface area': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}', 'walls'],
    #         'variable': ["q_lw['out_rad']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'marker': 'variable',
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'xlim': xlim_hw,
    #     'ylim': ylim_MJ_normalised,
    #     'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location', 'variable'],
    #     },
    
    }

plot_lines(df, plots_setup, line_properties)


# %% LW vs h/w


# %%% Daily LW vs hw

plots_setup = {
    
    
    # 'Daily LW Radiation \n divided by the ground area': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}', 'walls'],
    #         'variable': ["q_lw['in_from_sky']", "q_lw['in_from_els']", "q_lw['in_from_sky'] + q_lw['in_from_els']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'marker': 'variable',
    #         \
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'xlim': xlim_hw,
    #     # 'ylim': ylim_MJ_normalised,
    #     'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location', 'variable'],
    #     },
        
        
        
    # 'Daily LW Radiation on walls \n divided by the surface area': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': ['walls'],
    #         'variable': ["q_lw['in_from_sky']", "q_lw['in_from_els']", "q_lw['in_from_sky'] + q_lw['in_from_els']", "q_lw['out_rad']", "q_lw['net']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'marker': 'variable',
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'xlim': xlim_hw,
    #     'ylim': ylim_MJ_normalised,
    #     'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location', 'variable'],
    #     },
        
    
    'Daily LW Radiation on street \n divided by the surface area': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_sky']", "q_lw['in_from_els']", "q_lw['in_from_sky'] + q_lw['in_from_els']", "q_lw['out_rad']", "q_lw['net']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        'color': [colors[1], colors[2], colors[3], colors[4], colors[5], colors[6]],
            
        
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,
        'ylim': ylim_MJ_normalised,
        'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
        
        
    # f'Daily LW Radiation emitted \n divided by the surface area': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}', 'walls'],
    #         'variable': ["q_lw['out_rad']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'marker': 'variable',
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'xlim': xlim_hw,
    #     'ylim': ylim_MJ_normalised,
    #     'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location', 'variable'],
    #     },
    
    
    
    
    }

plot_lines(df, plots_setup, line_properties)


# %%% Delta (vs h/w=0) Daily absorbed LW vs h/w


plots_setup = {
    'Delta Daily net LW Radiation at canyon street \n (H/W>0 - H/W=0)': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            # 'hw': ['0.5 - 0.0', '1.0 - 0.0', '2.0 - 0.0', '4.0 - 0.0'],
            'hw': slice(None),
            'albedo': slice(None),
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['net']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'subtract': 'hw=0',
        'linestyle': 'albedo',
            
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,
        'y_scale': 1e-6,
        # 'ylim': [-725, 25],
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    
    }
    
    
plot_lines(df, plots_setup, line_properties)


# %% SW+LW vs t

# %%% SW+LW vs t (alb = 0.25)

 
plots_setup = {
    
    'SW Abs radiation, on canyon street \n divided by the surface area \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['abs']"],
            'res': 1.0,
            },
            
        'linewidth': 'hw',
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
        
    'LW net radiation, on canyon street \n divided by the surface area \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['net']"],
            'res': 1.0,
            },
            
        'linewidth': 'hw',
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
        
    'SW Abs radiation, on all canyon surfaces \n divided by the surface area \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['total'],
            'variable': ["q_sw['abs']"],
            'res': 1.0,
            },
            
        'linewidth': 'hw',
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
        
        
    'LW net radiation, on all canyon surfaces \n divided by the surface area \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['total'],
            'variable': ["q_lw['net']"],
            'res': 1.0,
            },
            
        'linewidth': 'hw',
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
        
    'SW Abs radiation, on canyon walls \n divided by the surface area \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['walls'],
            'variable': ["q_sw['abs']"],
            'res': 1.0,
            },
            
        'linewidth': 'hw',
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
        
        
    'LW net radiation, on canyon walls \n divided by the surface area \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['walls'],
            'variable': ["q_lw['net']"],
            'res': 1.0,
            },
            
        
        'linewidth': 'hw',
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
    
    
    'Net allwave radiation, on canyon walls \n divided by the surface area \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['walls'],
            'variable': ["q_lw['net'] + q_sw['abs']"],
            'res': 1.0,
            },
        
        'linewidth': 'hw',

        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
    
    
    'Net allwave radiation, on canyon street \n divided by the surface area \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['net'] + q_sw['abs']"],
            'res': 1.0,
            },
            
        'linewidth': 'hw',
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
        
    'Net allwave radiation, on all canyon surfaces \n divided by the surface area \n Albedo=0.25': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['total'],
            'variable': ["q_lw['net'] + q_sw['abs']"],
            'res': 1.0,
            },
            
            
        'linewidth': 'hw',
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
    
    
    
    
    }

plot_lines(df, plots_setup, line_properties)

#%%% SW & LW, als Oke

plots_setup = {
    'SW & LW Radiation on canyon street \n H/W=0.0 Albedo=0.25 \n divided by the surface area': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            'res': 1.0,
            },
            
        'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
        'annotations': True,
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
        
        
    'SW & LW Radiation on canyon street \n H/W=1.0 Albedo=0.0 \n divided by the surface area': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.0],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            'res': 1.0,
            },
            
        'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
        
        
    'SW & LW Radiation on canyon street \n H/W=1.0 Albedo=0.25 \n divided by the surface area': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            'res': 1.0,
            },
            
        'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
        
    'SW & LW Radiation on canyon street \n H/W=1.0 Albedo=0.5 \n divided by the surface area': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [1.0],
            'albedo': [0.5],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_sw['in']", "q_sw['in'] - q_sw['abs']", "q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"],
            'res': 1.0,
            },
            
        'fill_between': [["q_lw['in_from_els'] + q_lw['in_from_sky']", "q_lw['out_rad']"], ["q_sw['in']", "q_sw['in'] - q_sw['abs']"]],
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
    
    
    }

plot_lines(df, plots_setup, line_properties)

# %%% SW+LW vs t (alb != 0.25)

plots_setup = {
    # f'Albedo=0.0 \n Net allwave radiation, on canyon walls \n divided by the surface area': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
    #         'albedo': [0.0],
    #         'emissivity': 1.0, 
    #         'location': ['walls'],
    #         'variable': ["q_lw['net'] + q_sw['abs']"],
    #         'res': 1.0,
    #         },
            
    #     'x': 't',
    #     'ylabel': y_label1,
    #     # 'xlim': xlim_hw,
    #     # 'ylim': ylim_MJ_normalised,
    #     # 'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw', 'variable'],
    #     },
    
    
    # f'Albedo=0.5 \n Net allwave radiation, on canyon street \n divided by the surface area': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
    #         'albedo': [0.5],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["q_lw['net'] + q_sw['abs']"],
    #         'res': 1.0,
    #         },
            
            
    #     'x': 't',
    #     'ylabel': y_label1,
    #     # 'xlim': xlim_hw,
    #     # 'ylim': ylim_MJ_normalised,
    #     # 'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
        
    'Net allwave radiation, on all canyon surfaces \n divided by the surface area \n Albedo=0.0': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.0],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['total'],
            'variable': ["q_lw['net'] + q_sw['abs']"],
            'res': 1.0,
            },
            
            
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
        
        
    'Net allwave radiation, on all canyon surfaces \n divided by the surface area \n Albedo=0.5': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': [0.0, 0.5, 1.0, 2.0, 3.0],
            'albedo': [0.5],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['total'],
            'variable': ["q_lw['net'] + q_sw['abs']"],
            'res': 1.0,
            },
            
            
        'x': 't',
        'ylabel': y_label1,
        # 'xlim': xlim_hw,
        'ylim': ylim_W,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['hw'],
        },
    
    }

plot_lines(df, plots_setup, line_properties)






#%% Daily LW and SW vs hw

alb=0.2

plots_setup = {
    
    f'SW and LW Radiation on street 14:00 \n Albedo={alb}': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': [
                "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs']",
                "q_sw['abs_from_els'] + q_lw['abs_from_els']",
                "q_sw['abs_dir'] + q_sw['abs_diff'] + q_lw['abs_from_sky']",
                "q_convection",
                ],
            'res': 2.0,
            },
            
        # 'marker': 'variable',
            
        'between_time': ('12:10', '16:11'),
        'linestyle': 'variable',
        'color': 'variable',
        'x': 'hw',
        'ylabel': y_label1,
        'xlim': xlim_hw,
        'ylim': (-20, 1050),
        # 'y_unit': 'MJ/m2/d',
        # 'ylim': ylim_MJ_normalised,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
        
    f'SW Radiation on street 14:00 \n Albedo={alb}': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': [
                "q_sw['abs_dir'] + q_sw['abs_diff']", 
                "q_sw['abs_from_els']",
                "q_sw['abs_from_els'] + q_sw['abs_dir'] + q_sw['abs_diff']", 
                ],
            'res': 2.0,
            },
            
        # 'marker': 'variable',
            
        'between_time': ('12:10', '16:11'),
        'linestyle': 'variable',
        'color': 'variable',
        'x': 'hw',
        'ylabel': y_label1,
        'xlim': xlim_hw,
        # 'y_unit': 'MJ/m2/d',
        'ylim': (-20, 1050),
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
        
    f'LW Radiation on street 14:00 \n Albedo={alb}': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': [alb],
            'emissivity': slice(None),
            'skin thickness': slice(None),
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': slice(None),
            'location': [f'canyon street{slice_name}'],
            'variable': [
                "q_lw['abs_from_els']", 
                "q_lw['abs_from_sky']", 
                "q_lw['abs_from_els'] + q_lw['abs_from_sky']", 
                ],
            'res': 2.0,
            },
            
        # 'marker': 'variable',
            
        'between_time': ('12:10', '16:11'),
        'linestyle': 'variable',
        'color': 'variable',
        'x': 'hw',
        'ylabel': y_label1,
        'xlim': xlim_hw,
        'ylim': (-20, 1050),
        # 'y_unit': 'MJ/m2/d',
        # 'ylim': ylim_MJ_normalised,
        # 'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
        
    }
    

plot_lines(df, plots_setup, line_properties)

    
#%% Daily Allwave vs hw


plots_setup = {
    'Daily Allwave Radiation, on canyon walls \n divided by the surface area': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': slice(None),
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['walls'],
            'variable': ["q_lw['net'] + q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
            
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,
        'ylim': ylim_MJ_normalised,
        'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
        
    'Daily Allwave Radiation, on canyon street \n divided by the surface area': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': slice(None),
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['net'] + q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        'linestyle': 'albedo',

        
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,
        'ylim': ylim_MJ_normalised,
        'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    'Daily Allwave Radiation, on all canyon surfaces \n divided by the surface area': 
        {'xs': {
            'category': 'results, per m2 surface',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': slice(None),
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['total'],
            'variable': ["q_lw['net'] + q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
        'linestyle': 'albedo',
            
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,
        'ylim': ylim_MJ_normalised,
        'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
        
    'Daily Allwave Radiation, on all canyon surfaces \n divided by the ground area': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': slice(None),
            'albedo': slice(None),
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': ['total'],
            'variable': ["q_lw['net'] + q_sw['abs']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        # 'marker': 'variable',
            
        'x': 'hw',
        'ylabel': y_label2,
        'xlim': xlim_hw,
        'ylim': ylim_MJ_normalised,
        'y_scale': 1e-6,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['albedo'],
        },
        
    }
    
plot_lines(df, plots_setup, line_properties)


# %% ?? Flat area, Daily LW vs hw

plots_setup = {
    
    'LW Radiation at flat area \n H/W=0': 
        {'xs': {
            'category': 'results, per m2 ground',
            'emissivity sky': slice(None),
            'hw': [0.0],
            'albedo': [0.25],
            'emissivity': 1.0, 
            'skin thickness': [0.1],
            #'density': slice(None),
            'vol. heat cap.': slice(None),
            'wind speed': [0],
            'location': [f'canyon street{slice_name}'],
            'variable': ["q_lw['in_from_els']", "q_lw['in_from_sky']", "q_lw['net']", "q_lw['out_rad']"],
            # 'res': [1.0, 2.0],1
            'res': 1.0,
            },
            
        'x': 't',
        'ylabel': y_label1,
        'date': date,
        'droplevel': ('zenith', 'azimuth'),
        'legend': ['variable'],
        },
    
    }
    
plot_lines(df, plots_setup, line_properties)


# %%% ?


# plots_setup = {
    
    
    
    # 'db East-West Urban Canyon \n H/W=1.0 \n Absorbed SW Radiation at street': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0, 0.25, 0.5, 1.0],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["q_sw['abs']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Absorbed SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
    # 'db East-West Urban Canyon \n H/W=1.0 \n Absorbed SW Radiation at walls': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0, 0.25, 0.5, 1.0],
    #         'emissivity': 1.0, 
    #         'location': ['walls'],
    #         'variable': ["q_sw['abs']"],
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Absorbed SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
    # 'East-West Urban Canyon \n H/W=1.0 \n Absorbed SW Radiation at walls (normalised)': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0, 0.25, 0.5, 1.0],
    #         'emissivity': 1.0, 
    #         'location': ['walls'],
    #         'variable': ["q_sw['abs']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Absorbed SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
    # 'East-West Urban Canyon \n H/W=1.0 \n Reflected SW Radiation outwards from canyon': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0, 0.25, 0.5, 1.0],
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Outgoing SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
        
    # # various h/w
        
    # 'East-West Urban Canyon \n albedo=0.25 \n Absorbed SW Radiation at street': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [0.5, 1.0, 2.0, 4.0],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["q_sw['abs']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Absorbed SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
        
    # 'East-West Urban Canyon \n albedo=0.25 \n Absorbed SW Radiation at walls': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [0.5, 1.0, 2.0, 4.0],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': ['walls'],
    #         'variable': ["q_sw['abs']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Absorbed SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
    # 'East-West Urban Canyon \n albedo=0.25 \n Absorbed SW Radiation at walls (normalised)': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': [0.5, 1.0, 2.0, 4.0],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': ['walls'],
    #         'variable': ["q_sw['abs']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Absorbed SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
    # 'East-West Urban Canyon \n albedo=0.25 \n Reflected SW Radiation outwards from canyon': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [0.5, 1.0, 2.0, 4.0],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': ['bottom'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Outgoing SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        
        
    # # various location, fixed alb, fixed h/w 
    
    # 'East-West Urban Canyon \n albedo=0.25, H/W=1.0 \n SW Radiation on various surfaces': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}', 'walls', f'bottom{slice_name}'],
    #         'variable': ["q_sw['abs']", "q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location', 'variable'],
    #     },
        
    # 'East-West Urban Canyon \n albedo=1.0, various \n SW Radiation on various surfaces': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [0.5, 1.0, 2.0, 4.0],
    #         'albedo': [1.0],
    #         'emissivity': 1.0, 
    #         # 'location': [f'canyon street{slice_name}', 'walls', f'bottom{slice_name}', 'total'],
    #         'location': [f'bottom{slice_name}'],
    #         'variable': ["q_sw['in']", "q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     # 'legend': ['location', 'variable'],
    #     'legend': ['hw'],
    #     },
    
    
    # 'East-West Urban Canyon \n albedo=1.0, H/W=1.0 \n SW Radiation on various surfaces': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [1.0],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}', 'walls', f'bottom{slice_name}', 'total'],
    #         'variable': ["q_sw['abs']", "q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location', 'variable'],
    #     },
        
    # 'East-West Urban Canyon \n albedo=0.5, H/W=1.0 \n SW Radiation on various surfaces': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0.5],
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}', 'walls', f'bottom{slice_name}', 'total'],
    #         'variable': ["q_sw['abs']", "q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location', 'variable'],
    #     },
        
        
    # 'East-West Urban Canyon \n albedo=0.5, H/W=1.0 \n SW Radiation on sensors faces': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0.0],
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name}', f'top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location', 'variable'],
    #     },
        
        
    # 'db East-West Urban Canyon \n H/W=1.0 \n SW Radiation outwards from canyon': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0, 0.25, 0.5, 0.75, 1.0],
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name}', f'top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Outgoing SW radiation flux density [W/m2]',
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo', 'variable', 'location'],
    #     },
        

    # 'East-West Urban Canyon \n albedo=0.25 \n Canyon effective albedo': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [0.5, 1.0, 2.0],
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name} / top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Canyon Effective albedo [-]',
    #     'ylim': [0, 1.5],
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['hw'],
    #     },
        

    # 'East-West Urban Canyon \n H/W=1.0 \n Canyon effective albedo': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0, 0.25, 0.5, 1.0],
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name} / top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Canyon Effective albedo [-]',
    #     'ylim': [0, 1.5],
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
    # 'East-West Urban Canyon \n H/W=2.0 \n Canyon effective albedo': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [2.0],
    #         'albedo': [0, 0.25, 0.5, 1.0],
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name} / top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Canyon Effective albedo [-]',
    #     'ylim': [0, 1.5],
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
        
    # 'East-West Urban Canyon \n H/W=1.0 \n Canyon effective albedo / surface albedo': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [1.0],
    #         'albedo': [0.0, 0.25, 0.5, 1.0],
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name} / top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'divide': 'albedo',
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Canyon reflective efficiency [-]',
    #     'ylim': [0, 1.5],
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
        
    # 'db East-West Urban Canyon \n H/W=2.0 \n Canyon effective albedo / surface albedo': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': [2.0],
    #         'albedo': [0.0, 0.25, 0.5, 1.0],
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name} / top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'divide': 'albedo',
            
    #     'x': 'Local time [H]',
    #     'ylabel': 'Canyon reflective efficiency [-]',
    #     'ylim': [0, 1.5],
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
        
        
    # H/W x-axis
    
    # 'East-West Urban Canyon \n Absorbed SW Radiation at street': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': slice(None),
    #         'emissivity': 1.0, 
    #         'location': [f'canyon street{slice_name}'],
    #         'variable': ["q_sw['abs']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
    
    
    # 'East-West Urban Canyon \n Absorbed SW Radiation at walls': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': slice(None),
    #         'emissivity': 1.0, 
    #         'location': ['walls'],
    #         'variable': ["q_sw['abs']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
        
    # 'East-West Urban Canyon \n Reflected SW Radiation outwards from canyon': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': slice(None),
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
        
    #   'East-West Urban Canyon \n sensors ()': 
    #       {'xs': {
    #           'category': 'results, per m2 ground',
    #           'emissivity sky': slice(None),
    #           'hw': slice(None),
    #           'albedo': slice(None),
    #           'emissivity': 1.0, 
    #           'location': [f'bottom{slice_name}',f'top{slice_name}'],
    #           'variable': ["q_sw_sensors['in']"],
    #           # 'res': [1.0, 2.0],1
    #           'res': 1.0,
    #           },
             
    #       'x': 'hw',
    #       'ylabel': 'Albedo [-]',
    #       # 'y_scale': 1e-6,
    #       'date': date,
    #       'droplevel': ('zenith', 'azimuth'),
    #       'legend': ['albedo'],
    #       },
        
    # 'East-West Urban Canyon \n Daily effective canyon albedo (= Reflected / Incoming)': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': slice(None),
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name} / top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'hw',
    #     'ylabel': 'Albedo [-]',
    #     # 'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
    
    # 'East-West Urban Canyon \n Canyon Reflection Efficiency (= Daily effective canyon albedo / Surface albedo)': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': slice(None),
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name} / top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'divide': 'albedo',
    #     'x': 'hw',
    #     'ylabel': 'Albedo [-]',
    #     # 'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
    # 'db East-West Urban Canyon \n Canyon Reflection Efficiency (= Daily effective canyon albedo / Surface albedo)': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': slice(None),
    #         'emissivity': 1.0, 
    #         'location': [f'bottom{slice_name} / top{slice_name}'],
    #         'variable': ["q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'divide': 'albedo',
    #     'x': 'hw',
    #     'ylabel': 'Albedo [-]',
    #     # 'y_scale': 1e-6,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['albedo'],
    #     },
        
        
    # 'East-West Urban Canyon \n emissivity surface = 1.0 \n LW radiation on various surfaces':
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': ['walls', f'canyon street{slice_name}'],
    #         'variable': ["q_lw['in_from_els']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'y_scale': 1e-6,
    #     'ylim': [0, None],
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location'],
    #     },
        
        
    # 'East-West Urban Canyon \n emissivity surface = 1.0 \n LW radiation on various surfaces, normalised':
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': ['walls', f'canyon street{slice_name}'],
    #         'variable': ["q_lw['in_from_els']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'y_scale': 1e-6,
    #     'ylim': ylim_MJ_normalised,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location'],
    #     },
        
        
    # 'East-West Urban Canyon \n albedo = 0.25 \n SW Radiation Absorbed, at various surfaces ': 
    #     {'xs': {
    #         'category': 'results, per m2 ground',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         # 'location': [f'bottom{slice_name}'],
    #         'location': ['walls', f'canyon street{slice_name}', f'bottom{slice_name}'],
    #         'variable': ["q_sw['abs']", "q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'y_scale': 1e-6,
    #     'ylim': ylim_MJ_normalised,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location'],
    #     },
    
    
    # 'East-West Urban Canyon \n albedo = 0.25 \n SW Radiation Absorbed, at various surfaces, normalised ': 
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         # 'location': [f'bottom{slice_name}'],
    #         'location': ['walls', f'canyon street{slice_name}', f'bottom{slice_name}'],
    #         'variable': ["q_sw['abs']", "q_sw_sensors['in']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'y_scale': 1e-6,
    #     'ylim': ylim_MJ_normalised,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location'],
    #     },
        
        
    # 'db East-West Urban Canyon \n emissivity surface = 1.0 \n LW radiation on various surfaces, normalised':
    #     {'xs': {
    #         'category': 'results, per m2 surface',
    #         'emissivity sky': slice(None),
    #         'hw': slice(None),
    #         'albedo': [0.25],
    #         'emissivity': 1.0, 
    #         'location': ['walls', f'canyon street{slice_name}'],
    #         'variable': ["q_lw['in_from_els']"],
    #         # 'res': [1.0, 2.0],1
    #         'res': 1.0,
    #         },
            
    #     'x': 'hw',
    #     'ylabel': y_label2,
    #     'y_scale': 1e-6,
    #     'ylim': ylim_MJ_normalised,
    #     'date': date,
    #     'droplevel': ('zenith', 'azimuth'),
    #     'legend': ['location'],
    #     },
        
# }
    
    
        
    


plot_lines(df, plots_setup, line_properties)



